import io
from io import BytesIO  
import os  
import pdfkit
import json

from urllib.parse import urlencode
from datetime import datetime, timedelta, date
from dateutil.relativedelta import relativedelta 
from calendar import month_name

# Django imports
from django.shortcuts import render, redirect
from django.http import HttpResponse, JsonResponse
from django.db.models import Count, F, Sum, Q
from django.db.models.functions import ExtractWeek, ExtractYear, ExtractDay, ExtractMonth, TruncDay, TruncDate, TruncMonth, TruncQuarter, TruncYear
from django.utils import timezone
from django.core.paginator import Paginator, PageNotAnInteger, EmptyPage
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth import get_user_model, logout as auth_logout
from django.contrib import messages
from django.contrib.admin.models import LogEntry
from django.contrib.auth.models import Group, User as AuthUser
from django.template.loader import get_template, render_to_string
from django.core.serializers import serialize
from django.db import models
from django.contrib.gis.geos import GEOSGeometry
from django.template import loader

# Third-party libraries
import pandas as pd
import openpyxl  # Import the openpyxl library
from openpyxl import Workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
import PyPDF2
from fpdf import FPDF
import xhtml2pdf.pisa as pisa
#from weasyprint import HTML

# Local imports
from .models import Patient, History, Treatment, Barangay, Municipality, User,Doctor
from .forms import FormatForm
from .admin import PatientResource

# Python collections
from collections import Counter

# Django views
from django.views.generic import ListView, FormView

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
WKHTMLTOPDF_PATH = os.path.join(BASE_DIR, 'bin', 'wkhtmltopdf.exe')


class PatientListView(ListView, FormView):
    model = Patient
    template_name = 'monitoring/listview.html'
    form_class = FormatForm

    def post(self, request, *args, **kwargs):
        qs = self.get_queryset()
        dataset = PatientResource().export(qs)

        # Retrieve the selected format
        format = request.POST.get('format')
        
        # Select the format for the dataset export
        if format == 'xls':
            ds = dataset.export('xlsx')  # Export to xlsx (for Excel)
            content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            file_extension = 'xlsx'
        elif format == 'csv':
            ds = dataset.export('csv')  # Export to CSV
            content_type = 'text/csv'
            file_extension = 'csv'
        else:
            ds = dataset.export('json')  # Export to JSON
            content_type = 'application/json'
            file_extension = 'json'

        # Set the content type and download file name
        response = HttpResponse(ds, content_type=content_type)
        response['Content-Disposition'] = f'attachment; filename="patients.{file_extension}"'
        return response

def paginate_histories(histories):
    pages = []
    current_page = []
    max_rows_13 = 10  # Maximum rows for pages meant to hold smaller groups
    max_rows_15 = 15  # Maximum rows for other pages
    current_row_count = 0  # Counter to keep track of rows in the current page

    for history in histories:
        if current_row_count < max_rows_13:
            max_rows = max_rows_13
        else:
            max_rows = max_rows_15

        current_page.append(history)
        current_row_count += 1

        if len(current_page) >= max_rows:
            pages.append(current_page)
            current_page = []
            current_row_count = 0  # Reset row counter for the new page

    if current_page:  # Add remaining rows to the last page
        pages.append(current_page)

    return pages


def excel_masterlist_create(request):
    # Get the current user
    user = request.user

    # Get selected filters from request
    selected_municipality = request.GET.get('municipality')
    selected_barangay = request.GET.get('barangay')
    selected_user = request.GET.get('searchUsername') 
    start_month = request.GET.get('startMonth')
    end_month = request.GET.get('endMonth')
    search_name = request.GET.get('searchName')
    
    # Fetch the histories with related patient, municipality, and barangay data
    histories = History.objects.select_related('patient_id', 'muni_id', 'brgy_id').order_by('-registration_no')

    if not user.is_superuser:
        # Filter histories for the current user if not a superuser
        histories = histories.filter(patient_id__user=user)
    
    # Apply filters based on selected municipality and barangay
    if selected_municipality:
        histories = histories.filter(muni_id=selected_municipality)
    if selected_barangay:
        histories = histories.filter(brgy_id=selected_barangay)

    # Apply filter based on username search only if user is a superuser
    if user.is_superuser and selected_user:
        histories = histories.filter(patient_id__user__username=selected_user)
    
    # Apply filter based on selected start and end months
    if start_month and end_month:
        try:
            current_year = timezone.now().year
            start_date = datetime.strptime(f"{start_month} {current_year}", "%B %Y").replace(day=1)
            end_date = datetime.strptime(f"{end_month} {current_year}", "%B %Y").replace(day=1) + relativedelta(months=1) - relativedelta(days=1)
            histories = histories.filter(date_registered__gte=start_date, date_registered__lte=end_date)
        except Exception as e:
            print(f"Error parsing date: {e}")

    # Apply filter based on name search
    if search_name:
        histories = histories.filter(Q(patient_id__first_name__icontains=search_name) | Q(patient_id__last_name__icontains=search_name))

    # Prepare data for Excel
    data = []
    for history in histories:
        age = calculate_age(history.patient_id.birthday)
        treatment = Treatment.objects.filter(patient_id=history.patient_id).first()
        data.append({
            'Registration No': history.registration_no,
            'Patient Name': f"{history.patient_id.first_name} {history.patient_id.last_name}",
            'Municipality': history.muni_id.muni_name,
            'Barangay': history.brgy_id.brgy_name,
            'Date Registered': history.date_registered,
            'Age': age,
            'Treatment': treatment.vaccine_generic_name if treatment else '',
            'Sex': history.patient_id.sex,
            'Source of Exposure': history.source_of_exposure
        })

    # Create a DataFrame from the data
    df = pd.DataFrame(data)

    # Create an Excel response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = ' filename="masterlist.xlsx"'#attachment;

    # Write the DataFrame to the Excel response
    with pd.ExcelWriter(response, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Masterlist')

    return response

# Function to read data from Excel
def read_excel(file_path):
    df = pd.read_excel(file_path)
    return df

# Function to read data from PDF
def read_pdf(file_path):
    pdf_reader = PyPDF2.PdfFileReader(file_path)
    text = ""
    for page_num in range(pdf_reader.numPages):
        page = pdf_reader.getPage(page_num)
        text += page.extract_text()
    return text

# Function to generate PDF report
def generate_pdf_report(excel_data, pdf_data, output_path):
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", size=12)

    # Add Excel data to PDF
    pdf.cell(200, 10, txt="Excel Data", ln=True, align='C')
    for i in range(len(excel_data)):
        pdf.cell(200, 10, txt=str(excel_data.iloc[i].to_dict()), ln=True)

    # Add PDF data to PDF
    pdf.add_page()
    pdf.cell(200, 10, txt="PDF Data", ln=True, align='C')
    pdf.multi_cell(0, 10, pdf_data)

    pdf.output(output_path)

def index(request):
    # Retrieve all totals
    total_patients = Patient.objects.count()
    total_history = History.objects.count()
    total_treatments = Treatment.objects.count()

    # Count male and female patients
    male_patients = Patient.objects.filter(sex='male').count()
    female_patients = Patient.objects.filter(sex='female').count()

    # Calculate counts for different animal bites
    dog_bites = History.objects.filter(source_of_exposure='Dog').count()
    cat_bites = History.objects.filter(source_of_exposure='Cat').count()
    
    # Perform a query to count the occurrences of each municipality of exposure
    municipality_exposures = History.objects.values('muni_id__muni_name').annotate(count=Count('muni_id')).order_by('-count').first()

    # Determine which animal has the highest count
    animal_counts = History.objects.values('source_of_exposure').annotate(count=Count('source_of_exposure')).order_by('-count').first()
    most_cases_animal = animal_counts

    gender_counts = Patient.objects.values('sex').annotate(count=Count('sex')).order_by('-count').first()

    histories = History.objects.all()  # Retrieve all Histories from the database

    # Count the number of male and female patients
    gender = Patient.objects.values('sex').annotate(count=Count('sex')).order_by('sex')
    
    gen = [data['sex'].capitalize() for data in gender]
    dataGender = [data['count'] for data in gender]

    # Count the number animal source of exposure
    source_exposure = History.objects.values('source_of_exposure').annotate(count=Count('source_of_exposure')).order_by('source_of_exposure')
    
    animal = [data['source_of_exposure'].capitalize() for data in source_exposure]
    dataAnimal = [data['count'] for data in source_exposure]

    # Count occurrences of each municipality using Django ORM
    municipality_counts = History.objects.values('muni_id__muni_name').annotate(count=Count('muni_id')).order_by('muni_id__muni_name')
    """ print(municipality_counts.muni_name) """
    # Prepare data for chart
    municipalities = [data['muni_id__muni_name'] for data in municipality_counts]
    municipality_case_counts = [data['count'] for data in municipality_counts]

    # Group cases by month
    monthly_cases = History.objects.annotate(month=TruncMonth('date_registered')).values('month').annotate(count=Count('history_id')).order_by('month')

    months = [data['month'].strftime('%B') for data in monthly_cases]
    case_counts = [data['count'] for data in monthly_cases]

    # Calculate the number of cases per week
    weekly_cases = History.objects.annotate(
        week=ExtractWeek('date_registered'),
        year=ExtractYear('date_registered')
    ).values('year', 'week').annotate(count=Count('history_id')).order_by('year', 'week')

    # Prepare the data for the chart with start dates of weeks
    weeks = []
    weekly_case_counts = []
    for entry in weekly_cases:
        year = entry['year']
        week = entry['week']
        start_date = datetime.strptime(f'{year}-W{week}-1', "%Y-W%U-%w")  # Ensure the week starts from Sunday
        week_label = start_date.strftime('%Y-%m-%d')
        weeks.append(week_label)
        weekly_case_counts.append(entry['count'])

    # Calculate the number of cases per day
    daily_cases = (
        History.objects
        .annotate(day=TruncDate('date_registered'))
        .values('day')
        .annotate(count=Count('history_id'))
        .order_by('day')
    )

    days = []
    daily_case_counts = []

    for data in daily_cases:
        day = data['day']
        count = data['count']
        
        if day is not None:
            days.append(day.strftime('%Y-%m-%d'))
        else:
            days.append('Unknown')  # Handle None case with a fallback
        
        daily_case_counts.append(count)

    # Calculate the number of cases per quarter
    quarterly_cases = (
        History.objects
        .annotate(quarter=TruncQuarter('date_registered'))
        .values('quarter')
        .annotate(count=Count('history_id'))
        .order_by('quarter')
    )

    # Prepare quarters and quarterly counts
    quarters = []
    quarterly_case_counts = []
    
    for data in quarterly_cases:
        quarter_str = f"Quarter {((data['quarter'].month - 1) // 3) + 1} {data['quarter'].year}"
        quarters.append(quarter_str)
        quarterly_case_counts.append(data['count'])

        
    # Calculate the number of cases per year
    annual_cases = (
        History.objects
        .annotate(year=TruncYear('date_registered'))
        .values('year')
        .annotate(count=Count('history_id'))
        .order_by('year')
    )

    years = [data['year'].strftime('%Y') for data in annual_cases]
    annual_case_counts = [data['count'] for data in annual_cases]


    total_cases = History.objects.count()

    # Get the start and end dates from the GET parameters
    start_date = request.GET.get('startDate')
    end_date = request.GET.get('endDate')

    # Filter history objects based on date range
    if start_date and end_date:
        # Convert string dates to datetime objects
        start_date = datetime.strptime(start_date, '%Y-%m-%d')
        end_date = datetime.strptime(end_date, '%Y-%m-%d')

        # Filter heatmap data based on the selected date range
        heatmap_data = History.objects.filter(
            date_registered__range=(start_date, end_date)  # Ensure to use the correct field
        ).values('latitude', 'longitude').annotate(count=Count('history_id'))

        # Filter human rabies cases based on the selected date range
        rabies_heatmap_data = History.objects.filter(
            date_registered__range=(start_date, end_date),
            human_rabies=True
        ).values('latitude', 'longitude').annotate(count=Count('history_id'))

        # Get the total rabies cases within the date range
        total_rabies_cases = History.objects.filter(
            date_registered__range=(start_date, end_date),
            human_rabies=True
        ).count()

        # Get the total cases within the date range
        total_cases = History.objects.filter(
            date_registered__range=(start_date, end_date)
        ).count()

    else:
        # If no dates provided, get all history
        heatmap_data = History.objects.values('latitude', 'longitude').annotate(count=Count('history_id'))
        total_cases = History.objects.count()  # Total cases without filter

        # For human rabies data without date filter
        rabies_heatmap_data = History.objects.filter(human_rabies=True).values('latitude', 'longitude').annotate(count=Count('history_id'))
        total_rabies_cases = History.objects.filter(human_rabies=True).count()

    # Prepare the heatmap data for JSON
    heatmap_data = [[entry['latitude'], entry['longitude'], entry['count']] for entry in heatmap_data]

    # Prepare the rabies heatmap data for JSON
    rabies_heatmap_data = [[entry['latitude'], entry['longitude'], entry['count']] for entry in rabies_heatmap_data]

    # Get human rabies cases by municipality
    rabies_municipality_counts = History.objects.filter(human_rabies=True).values('muni_id__muni_name').annotate(count=Count('muni_id')).order_by('-count')

    # Prepare data for human rabies by municipality
    rabies_municipalities = [data['muni_id__muni_name'] for data in rabies_municipality_counts]
    rabies_case_counts = [data['count'] for data in rabies_municipality_counts]

    context = {
        'total_patients': total_patients,
        'total_history': total_history,
        'total_treatments': total_treatments,
        'male_patients': male_patients,
        'female_patients': female_patients,
        'dog_bites': dog_bites,
        'cat_bites': cat_bites,
        'most_cases_animal': most_cases_animal,
        'municipality': municipality_exposures,
        'gender_counts': gender_counts,
        'histories': histories,
        'gen': gen,
        'dataGender': dataGender,
        'animal': animal,
        'dataAnimal': dataAnimal,
        'municipalities': municipalities,
        'municipality_case_counts': municipality_case_counts,
        'months': months,
        'case_counts': case_counts,
        'weeks': weeks,
        'weekly_case_counts': weekly_case_counts,
        'days': days,
        'daily_case_counts': daily_case_counts,
        'quarters': quarters,
        'quarterly_case_counts': quarterly_case_counts,
        'years': years,
        'annual_case_counts': annual_case_counts,
        'total_cases': total_cases,
        'heatmap_data': json.dumps(heatmap_data),
        'total_rabies_cases': total_rabies_cases,
        'rabies_municipalities': rabies_municipalities,
        'rabies_case_counts': rabies_case_counts,
        'rabies_heatmap_data': json.dumps(rabies_heatmap_data),
    }
    return render(request, 'monitoring/index.html', context)


def choropleth_map(request):
    # Choropleth Map
    barangay_density = History.objects.values('brgy_id').annotate(patient_count=Count('patient_id')).order_by('brgy_id')

    # Convert this to a dictionary where the key is brgy_id and the value is the patient_count
    density_dict = {str(d['brgy_id']): d['patient_count'] for d in barangay_density}

    # Get human rabies cases by municipality
    rabies_municipality_counts = History.objects.filter(human_rabies=True).values('muni_id__muni_name').annotate(count=Count('muni_id')).order_by('-count')
    
    # Total rabies cases
    total_rabies_cases = History.objects.filter(human_rabies=True).count()

    # Prepare data for human rabies by municipality
    rabies_municipalities = [data['muni_id__muni_name'] for data in rabies_municipality_counts]
    rabies_case_counts = [data['count'] for data in rabies_municipality_counts]

    # Prepare heatmap data for human rabies cases
    rabies_heatmap_data = History.objects.filter(human_rabies=True).values('latitude', 'longitude').annotate(count=Count('history_id'))

    # Prepare the heatmap data for JSON
    rabies_heatmap_data = [[entry['latitude'], entry['longitude'], entry['count']] for entry in rabies_heatmap_data]
    
    all_municipalities = Municipality.objects.all()

    if request.method == 'POST':
        selected_municipality = request.POST.get('municipality', None)
        barangay_search = request.POST.get('barangay', '').strip()
    else:
        selected_municipality = "ALL"
        barangay_search = ""

    total_barangays = 0
    total_patients = 0
    barangay_summary = []
    municipality_summary = []

    # Check for barangay search input
    if barangay_search:
        # Fetch barangays matching search across all municipalities
        barangays_in_municipality = Barangay.objects.filter(brgy_name__icontains=barangay_search)
        
        # Get barangay-level case summary
        cases_in_barangays = History.objects.filter(brgy_id__brgy_name__icontains=barangay_search) \
            .values('brgy_id__brgy_name', 'muni_id__muni_name') \
            .annotate(total_patients=Count('history_id')) \
            .order_by('brgy_id__brgy_name')

        cases_dict = {(record['muni_id__muni_name'], record['brgy_id__brgy_name']): record['total_patients'] for record in cases_in_barangays}

        barangay_summary = [
            {
                'muni_name': barangay.muni_id.muni_name,
                'brgy_name': barangay.brgy_name,
                'total_patients': cases_dict.get((barangay.muni_id.muni_name, barangay.brgy_name), 0)
            }
            for barangay in barangays_in_municipality
        ]

        total_barangays = barangays_in_municipality.count()
        total_patients = sum(record['total_patients'] for record in barangay_summary)
    elif selected_municipality == "ALL" or selected_municipality is None:
        municipality_summary = (
            History.objects.values('muni_id__muni_name')
            .annotate(total_barangays=Count('brgy_id', distinct=True), total_cases=Count('history_id'))
            .order_by('muni_id__muni_name')
        )

        total_barangays = sum(record['total_barangays'] for record in municipality_summary)
        total_patients = sum(record['total_cases'] for record in municipality_summary)
    else:
        barangays_in_municipality = Barangay.objects.filter(muni_id__muni_name=selected_municipality)
        cases_in_barangays = History.objects.filter(muni_id__muni_name=selected_municipality) \
            .values('brgy_id__brgy_name') \
            .annotate(total_patients=Count('history_id')) \
            .order_by('brgy_id__brgy_name')

        cases_dict = {record['brgy_id__brgy_name']: record['total_patients'] for record in cases_in_barangays}
        barangay_summary = [
            {
                'brgy_name': barangay.brgy_name,
                'total_patients': cases_dict.get(barangay.brgy_name, 0)
            }
            for barangay in barangays_in_municipality
        ]

        total_barangays = barangays_in_municipality.count()
        total_patients = sum(record['total_patients'] for record in barangay_summary)

        municipality_summary = [
            {
                'muni_id__muni_name': selected_municipality,
                'total_barangays': total_barangays,
                'total_cases': total_patients,
            }
        ]
    
    # Calculate total patients for all municipalities or barangays
    total_patients_all = total_patients  # This already includes all patients for the selected filter

    # For Barangay or Municipality summary, calculate percentage
    for record in barangay_summary:
        record['percentage'] = (record['total_patients'] / total_patients_all) * 100

    for record in municipality_summary:
        record['percentage'] = (record['total_cases'] / total_patients_all) * 100

    # Pass the data to the template as part of the context
    context = {
        'total_rabies_cases': total_rabies_cases,
        'rabies_case_counts': rabies_case_counts,
        'density_dict': density_dict,
        'rabies_heatmap_data': json.dumps(rabies_heatmap_data),
        'barangay_summary': barangay_summary,
        'municipality_summary': municipality_summary,
        'municipalities': all_municipalities,
        'selected_municipality': selected_municipality,
        'barangay_search': barangay_search,  # Pass the search input to the template
        'total_barangays': total_barangays,
        'total_patients': total_patients,
        'all_municipalities': all_municipalities,
    }

    return render(request, 'monitoring/choropleth.html', context)


@login_required
def choro(request):
    # Determine if the user is a superuser
    is_superuser = request.user.is_superuser

    # Query all history data if the user is superuser, otherwise filter by specific patient linked to the logged-in user
    if is_superuser:
        history_queryset = History.objects.all()
    else:
        # Filter history data by the specific patient linked to the logged-in user
        history_queryset = History.objects.filter(patient_id__user=request.user)

    # Get barangay densities: Count the number of patients per barangay for the relevant histories
    barangay_density = history_queryset.values('brgy_id').annotate(patient_count=Count('patient_id')).order_by('brgy_id')

    # Convert this to a dictionary where the key is brgy_id and the value is the patient_count
    density_dict = {str(d['brgy_id']): d['patient_count'] for d in barangay_density}

    # Get human rabies cases by municipality
    rabies_municipality_counts = history_queryset.filter(human_rabies=True).values('muni_id__muni_name').annotate(count=Count('muni_id')).order_by('-count')
    
    # Total rabies cases
    total_rabies_cases = history_queryset.filter(human_rabies=True).count()

    # Prepare data for human rabies by municipality
    rabies_municipalities = [data['muni_id__muni_name'] for data in rabies_municipality_counts]
    rabies_case_counts = [data['count'] for data in rabies_municipality_counts]

    # Prepare heatmap data for human rabies cases
    rabies_heatmap_data = history_queryset.filter(human_rabies=True).values('latitude', 'longitude').annotate(count=Count('history_id'))

    # Prepare the heatmap data for JSON
    rabies_heatmap_data = [[entry['latitude'], entry['longitude'], entry['count']] for entry in rabies_heatmap_data]

    # Pass the data to the template as part of the context
    context = {
        'density_dict': density_dict,  # For barangay patient density
        'rabies_case_counts': rabies_case_counts,
        'total_rabies_cases': total_rabies_cases,
        'rabies_heatmap_data': json.dumps(rabies_heatmap_data),
    }
    
    return render(request, 'monitoring/choro.html', context)

def admin_redirect(request):
    return redirect('/admin/')

def logout(request):
    auth_logout(request) 
    return render(request, 'registration/logged_out.html')

# Function to calculate the age of the patient
def calculate_age(birthday):
    today = date.today()
    return today.year - birthday.year - ((today.month, today.day) < (birthday.month, birthday.day))

def staff_or_superuser_required(view_func):
    def _wrapped_view_func(request, *args, **kwargs):
        if not (request.user.is_staff or request.user.is_superuser):
            messages.warning(request, "Only staff members or superusers can access this page.")
            return redirect('/admin/')
        return view_func(request, *args, **kwargs)
    return login_required(_wrapped_view_func)

def superuser_required(view_func):
    def _wrapped_view_func(request, *args, **kwargs):
        if not request.user.is_superuser:
            messages.warning(request, "Only superusers or admins can access this page.")
            return redirect('/admin/')  # Replace 'some_view_name' with the name of the view you want to redirect to
        return view_func(request, *args, **kwargs)
    return login_required(_wrapped_view_func)

@staff_member_required
@login_required
def overview(request):
    staff = request.user
    
    CODE_TO_MUNICIPALITY = {
        'ALM': 'ALMERIA',
        'BIL': 'BILIRAN',
        'CABUC': 'CABUCGAYAN',
        'CAIB': 'CAIBIRAN',
        'CUL': 'CULABA',
        'KAW': 'KAWAYAN',
        'MAR': 'MARIPIPI',
        'NAV': 'NAVAL',
    }


    if staff.is_superuser:
        municipality = "BPH"
        treatment_center = "Animal Bite Treatment Center: Biliran Province Hospital"
    else:
        municipality = CODE_TO_MUNICIPALITY.get(staff.code, "Unknown Municipality")
        treatment_center = f"Animal Bite Treatment Center: {municipality}"

    fname = staff.first_name
    lname = staff.last_name

    # Get the start and end dates from the GET parameters
    start_date = request.GET.get('startDate')
    end_date = request.GET.get('endDate')

    # Check if the current user is a superuser
    is_superuser = request.user.is_superuser

    if is_superuser:
        history_queryset = History.objects.all()
    else:
        # Filter history data by the specific patient linked to the logged-in user's username
        history_queryset = History.objects.filter(patient_id__user__username=request.user.username)

    # Apply date filtering if start and end dates are provided
    if start_date and end_date:
        start_date = datetime.strptime(start_date, '%Y-%m-%d')
        end_date = datetime.strptime(end_date, '%Y-%m-%d')
        history_queryset = history_queryset.filter(date_registered__range=(start_date, end_date))

    # Total number of cases
    total_cases = history_queryset.count()

    # Total human rabies cases
    total_rabies_cases = history_queryset.filter(human_rabies=True).count()

    # Heatmap data for all cases
    heatmap_data = (
        history_queryset
        .values('latitude', 'longitude')
        .annotate(count=Count('history_id'))
    )
    heatmap_data = [[entry['latitude'], entry['longitude'], entry['count']] for entry in heatmap_data]

    # Heatmap data specifically for human rabies cases
    rabies_heatmap_data = (
        history_queryset
        .filter(human_rabies=True)
        .values('latitude', 'longitude')
        .annotate(count=Count('history_id'))
    )
    rabies_heatmap_data = [[entry['latitude'], entry['longitude'], entry['count']] for entry in rabies_heatmap_data]

    # Human rabies cases by municipality
    rabies_municipality_counts = (
        history_queryset
        .filter(human_rabies=True)
        .values('muni_id__muni_name')
        .annotate(count=Count('muni_id'))
        .order_by('-count')
    )
    rabies_municipalities = [data['muni_id__muni_name'] for data in rabies_municipality_counts]
    rabies_case_counts = [data['count'] for data in rabies_municipality_counts]

    # Municipality counts
    municipality_counts = (
        history_queryset
        .values('muni_id__muni_name')
        .annotate(count=Count('muni_id'))
        .order_by('muni_id__muni_name')
    )
    municipalities = [data['muni_id__muni_name'] for data in municipality_counts]
    municipality_case_counts = [data['count'] for data in municipality_counts]

    # Gender distribution
    gender = history_queryset.values('patient_id__sex').annotate(count=Count('patient_id__sex')).order_by('patient_id__sex')
    gen = [data['patient_id__sex'].capitalize() for data in gender]
    dataGender = [data['count'] for data in gender]

    # Animal sources of exposure
    source_exposure = history_queryset.values('source_of_exposure').annotate(count=Count('source_of_exposure')).order_by('source_of_exposure')
    animal = [data['source_of_exposure'].capitalize() for data in source_exposure]
    dataAnimal = [data['count'] for data in source_exposure]

    # Monthly cases
    monthly_cases = history_queryset.annotate(month=TruncMonth('date_registered')).values('month').annotate(count=Count('history_id')).order_by('month')
    months = [data['month'].strftime('%B') for data in monthly_cases]
    case_counts = [data['count'] for data in monthly_cases]

    # Weekly cases
    weekly_cases = history_queryset.annotate(
        week=ExtractWeek('date_registered'),
        year=ExtractYear('date_registered')
    ).values('year', 'week').annotate(count=Count('history_id')).order_by('year', 'week')

    weeks = []
    weekly_case_counts = []
    for entry in weekly_cases:
        year = entry['year']
        week = entry['week']
        start_date = datetime.strptime(f'{year}-W{week}-1', "%Y-W%U-%w")
        week_label = start_date.strftime('%Y-%m-%d')
        weeks.append(week_label)
        weekly_case_counts.append(entry['count'])

    # Daily cases
    daily_cases = history_queryset.annotate(day=TruncDate('date_registered')).values('day').annotate(count=Count('history_id')).order_by('day')
    days = [data['day'].strftime('%Y-%m-%d') if data['day'] else 'Unknown' for data in daily_cases]
    daily_case_counts = [data['count'] for data in daily_cases]

    # Quarterly cases
    quarterly_cases = history_queryset.annotate(quarter=TruncQuarter('date_registered')).values('quarter').annotate(count=Count('history_id')).order_by('quarter')
    quarters = [f"Quarter {((data['quarter'].month - 1) // 3) + 1} {data['quarter'].year}" for data in quarterly_cases]
    quarterly_case_counts = [data['count'] for data in quarterly_cases]

    # Annual cases
    annual_cases = history_queryset.annotate(year=TruncYear('date_registered')).values('year').annotate(count=Count('history_id')).order_by('year')
    years = [data['year'].strftime('%Y') for data in annual_cases]
    annual_case_counts = [data['count'] for data in annual_cases]

    # Pass all data to the template
    context = {
        'heatmap_data': json.dumps(heatmap_data),
        'total_cases': total_cases,
        'total_rabies_cases': total_rabies_cases,
        'rabies_municipalities': rabies_municipalities,
        'rabies_case_counts': rabies_case_counts,
        'rabies_heatmap_data': json.dumps(rabies_heatmap_data),
        'municipalities': municipalities,
        'municipality_case_counts': municipality_case_counts,
        'gen': gen,
        'dataGender': dataGender,
        'animal': animal,
        'dataAnimal': dataAnimal,
        'days': days,
        'daily_case_counts': daily_case_counts,
        'weeks': weeks,
        'weekly_case_counts': weekly_case_counts,
        'months': months,
        'case_counts': case_counts,
        'quarters': quarters,
        'quarterly_case_counts': quarterly_case_counts,
        'years': years,
        'annual_case_counts': annual_case_counts,
        'username': request.user.username,
        'fname': fname,
        'lname': lname,
        'municipality': municipality,
        'treatment_center': treatment_center,
    }

    return render(request, 'monitoring/overview.html', context)

def calculate_age(birthday):
    today = datetime.today()
    age = today.year - birthday.year - ((today.month, today.day) < (birthday.month, birthday.day))
    return age

@login_required
def reports(request):
    user = request.user

    municipality_map = {
        "MAR": "Maripipi",
        "KAW": "Kawayan",
        "NAV": "Naval",
        "CAIB": "Caibiran",
        "ALM":"Almeria",
        "BIL":"Biliran",
        "CUL":"Culaba",
        "CABUC":"Cabucgayan"
    }

    if user.is_superuser and user.code == "NAV":
        municipality_name = "BPH"
    else:
        municipality_name = municipality_map.get(user.code, "Naval")

    selected_quarter = request.GET.get('quarter', '1')  # Default to '1' if no quarter is selected
    year = date.today().year

    quarter_ranges = {
        '1': (date(year, 1, 1), date(year, 3, 31)),
        '2': (date(year, 4, 1), date(year, 6, 30)),
        '3': (date(year, 7, 1), date(year, 9, 30)),
        '4': (date(year, 10, 1), date(year, 12, 31)),
    }

    # Generate the title and date range based on the selected quarter
    if selected_quarter == 'annual':
        start_date, end_date = date(year, 1, 1), date(year, 12, 31)
        quarter_select = "Annual"
        report_title = f"{quarter_select} Report {year}"  # Annual report title
    else:
        start_date, end_date = quarter_ranges[selected_quarter]
        if selected_quarter == '1':
            quarter_select = '1st'
        elif selected_quarter == '2':
            quarter_select = '2nd'
        elif selected_quarter == '3':
            quarter_select = '3rd'
        elif selected_quarter == '4':
            quarter_select = '4th'
        report_title = f"{quarter_select} Quarter Report {year}"

    data = []
    total_male = 0
    total_female = 0
    total_all = 0
    total_age_15_below = 0
    total_age_above_15 = 0
    total_sex_percentage = 0
    total_animal_bite_I = 0
    total_animal_bite_II = 0
    total_animal_bite_III = 0
    total_category_percentage = 0
    total_tcv_given = 0
    total_hrig_given = 0
    total_erig_given = 0  
    total_tcv_percentage = 0
    total_hrig_percentage = 0
    total_rig_percentage = 0
    total_dog_bites = 0
    total_cat_bites = 0
    total_other_bites = 0
    total_animal_type_percentage = 0
    total_animal_bite_I_percentage = 0
    total_animal_bite_II_percentage = 0
    total_animal_bite_III_percentage = 0
    total_immunized = 0
    total_unimmunized = 0


    if user.is_superuser:
        abtcs = User.objects.filter(is_superuser=False).distinct()
        for abtc_user in abtcs:
            if not abtc_user.code:
                continue

            male_count = History.objects.filter(
                patient_id__user=abtc_user,
                date_registered__range=(start_date, end_date),
                patient_id__sex='male'
            ).count()
            female_count = History.objects.filter(
                patient_id__user=abtc_user,
                date_registered__range=(start_date, end_date),
                patient_id__sex='female'
            ).count()

            patients = Patient.objects.filter(
                user=abtc_user,  
                histories__date_registered__range=(start_date, end_date)  
            ).distinct()
            age_15_below_count = sum(1 for patient in patients if calculate_age(patient.birthday) <= 15)
            age_above_15_count = sum(1 for patient in patients if calculate_age(patient.birthday) > 15)

            user_animal_bite_I = 0
            user_animal_bite_II = 0
            user_animal_bite_III = 0

            animal_bite_counts = History.objects.filter(
                patient_id__user=abtc_user,
                date_registered__range=(start_date, end_date)
            ).values('category_of_exposure').annotate(count=models.Count('category_of_exposure'))

            for count in animal_bite_counts:
                if count['category_of_exposure'] == 'I':
                    user_animal_bite_I = count['count']
                elif count['category_of_exposure'] == 'II':
                    user_animal_bite_II = count['count']
                elif count['category_of_exposure'] == 'III':
                    user_animal_bite_III = count['count']

            tcv_count = Treatment.objects.filter(
                patient_id__user=abtc_user,
                tcv_given__range=(start_date, end_date)
            ).count()

            hrig_count = Treatment.objects.filter(
                patient_id__user=abtc_user,
                hrig_given__range=(start_date, end_date)
            ).count()

            erig_count = Treatment.objects.filter(
                patient_id__user=abtc_user,
                rig_given__range=(start_date, end_date) 
            ).count()

            animal_type_counts = History.objects.filter(
                patient_id__user=abtc_user,
                date_registered__range=(start_date, end_date)
            ).values('source_of_exposure').annotate(count=models.Count('source_of_exposure'))

            user_dog_bites = 0
            user_cat_bites = 0
            user_other_bites = 0

            for count in animal_type_counts:
                if count['source_of_exposure'] == 'Dog':
                    user_dog_bites = count['count']
                elif count['source_of_exposure'] == 'Cat':
                    user_cat_bites = count['count']
                elif count['source_of_exposure'] == 'Others':
                    user_other_bites = count['count']

            user_immunized_count = History.objects.filter(
                patient_id__user=abtc_user,
                immunization_status='Immunized',
                date_registered__range=(start_date, end_date)
            ).count()
            
            user_unimmunized_count = History.objects.filter(
                patient_id__user=abtc_user,
                immunization_status='Unimmunized',
                date_registered__range=(start_date, end_date)
            ).count()

            user_human_rabies_count = History.objects.filter(
                patient_id__user=abtc_user,
                human_rabies=True,
                date_registered__range=(start_date, end_date)
            ).count()
                    
            total_immunized += user_immunized_count
            total_unimmunized += user_unimmunized_count
            total_dog_bites += user_dog_bites
            total_cat_bites += user_cat_bites
            total_other_bites += user_other_bites
            total_tcv_given += tcv_count
            total_hrig_given += hrig_count
            total_erig_given += erig_count

            total_count = male_count + female_count
            total_male += male_count
            total_female += female_count
            total_all += total_count
            total_age_15_below += age_15_below_count
            total_age_above_15 += age_above_15_count
            total_animal_bite_I += user_animal_bite_I
            total_animal_bite_II += user_animal_bite_II
            total_animal_bite_III += user_animal_bite_III

            # Determine barangay field for this ABTC user
            if abtc_user.code == "NAV":
                barangay_name = "BPH-ABTC"
            else:
                barangay_name = f"{municipality_map.get(abtc_user.code, 'Unknown')}-ABTC"

            data.append({
                'barangay': barangay_name,
                'data_male': male_count,
                'data_female': female_count,
                'data_total': total_count,
                'age_15_below': age_15_below_count,
                'age_above_15': age_above_15_count,
                'age_total': age_15_below_count + age_above_15_count,
                'total_animal_bite_I':user_animal_bite_I,
                'total_animal_bite_II':user_animal_bite_II,
                'total_animal_bite_III':user_animal_bite_III,
                'total_animal':user_animal_bite_I + user_animal_bite_II + user_animal_bite_III,
                'total_tcv_given': tcv_count,
                'total_hrig_given': hrig_count,
                'total_erig_given': erig_count,
                'total_dog_bites': user_dog_bites,
                'total_cat_bites': user_cat_bites,
                'total_other_bites': user_other_bites,
                'total_animal_bites': user_dog_bites + user_cat_bites + user_other_bites,
                'immunized_count': user_immunized_count,
                'unimmunized_count': user_unimmunized_count,  
                'human_rabies_count':user_human_rabies_count,

            })
    else:
        # For non-superuser
        patients = Patient.objects.filter(user=user)
        barangays = Barangay.objects.filter(patients_brgy__in=patients).distinct()

        for barangay in barangays:
            # Filter by date range (selected quarter)
            male_count = History.objects.filter(
                patient_id__brgy_id=barangay,
                patient_id__user=user,
                date_registered__range=(start_date, end_date),
                patient_id__sex='male'
            ).count()

            female_count = History.objects.filter(
                patient_id__brgy_id=barangay,
                patient_id__user=user,
                date_registered__range=(start_date, end_date),
                patient_id__sex='female'
            ).count()

            # Skip barangay if there is no data for the selected quarter
            if male_count == 0 and female_count == 0:
                continue

            # Calculate age breakdown (15 and below, above 15)
            patients_in_quarter = Patient.objects.filter(
                brgy_id=barangay,
                user=user,
                histories__date_registered__range=(start_date, end_date)
            ).distinct()

            age_15_below_count = sum(1 for patient in patients_in_quarter if calculate_age(patient.birthday) <= 15)
            age_above_15_count = sum(1 for patient in patients_in_quarter if calculate_age(patient.birthday) > 15)

            # Initialize animal bite categories and other counters
            barangay_animal_bite_I = 0
            barangay_animal_bite_II = 0
            barangay_animal_bite_III = 0

            animal_bite_counts = History.objects.filter(
                patient_id__brgy_id=barangay,
                patient_id__user=user,
                date_registered__range=(start_date, end_date)
            ).values('category_of_exposure').annotate(count=models.Count('category_of_exposure'))

            tcv_count = Treatment.objects.filter(
                patient_id__brgy_id=barangay,
                patient_id__user=user,
                tcv_given__range=(start_date, end_date)
            ).count()

            hrig_count = Treatment.objects.filter(
                patient_id__brgy_id=barangay,
                patient_id__user=user,
                hrig_given__range=(start_date, end_date)
            ).count()

            erig_count = Treatment.objects.filter(
                patient_id__brgy_id=barangay,
                patient_id__user=user,
                rig_given__range=(start_date, end_date)
            ).count()

            # Count animal bites for different categories
            for count in animal_bite_counts:
                if count['category_of_exposure'] == 'I':
                    barangay_animal_bite_I = count['count']
                elif count['category_of_exposure'] == 'II':
                    barangay_animal_bite_II = count['count']
                elif count['category_of_exposure'] == 'III':
                    barangay_animal_bite_III = count['count']

            # Count different sources of animal bites (e.g., dog, cat)
            animal_type_counts = History.objects.filter(
                patient_id__brgy_id=barangay,
                patient_id__user=user,
                date_registered__range=(start_date, end_date)
            ).values('source_of_exposure').annotate(count=models.Count('source_of_exposure'))

            barangay_dog_bites = 0
            barangay_cat_bites = 0
            barangay_other_bites = 0

            for count in animal_type_counts:
                if count['source_of_exposure'] == 'Dog':
                    barangay_dog_bites = count['count']
                elif count['source_of_exposure'] == 'Cat':
                    barangay_cat_bites = count['count']
                elif count['source_of_exposure'] == 'Others':
                    barangay_other_bites = count['count']

            # Immunization counts
            barangay_immunized_count = History.objects.filter(
                patient_id__brgy_id=barangay,
                patient_id__user=user,
                immunization_status='Immunized',
                date_registered__range=(start_date, end_date)
            ).count()

            barangay_unimmunized_count = History.objects.filter(
                patient_id__brgy_id=barangay,
                patient_id__user=user,
                immunization_status='Unimmunized',
                date_registered__range=(start_date, end_date)
            ).count()

            # Human rabies count
            barangay_human_rabies_count = History.objects.filter(
                patient_id__brgy_id=barangay,
                patient_id__user=user,
                human_rabies=True,
                date_registered__range=(start_date, end_date)
            ).count()

            total_immunized += barangay_immunized_count
            total_unimmunized += barangay_unimmunized_count
            total_dog_bites += barangay_dog_bites
            total_cat_bites += barangay_cat_bites
            total_other_bites += barangay_other_bites
            total_dog_bites += barangay_dog_bites
            total_cat_bites += barangay_cat_bites
            total_other_bites += barangay_other_bites
            total_animal_bite_I += barangay_animal_bite_I
            total_animal_bite_II += barangay_animal_bite_II
            total_animal_bite_III += barangay_animal_bite_III

            # Add to total counts
            total_count = male_count + female_count
            total_male += male_count
            total_female += female_count
            total_all += total_count
            total_age_15_below += age_15_below_count
            total_age_above_15 += age_above_15_count
            data.append({
                'barangay': barangay.brgy_name,
                'data_male': male_count,
                'data_female': female_count,
                'data_total': total_count,
                'age_15_below': age_15_below_count,
                'age_above_15': age_above_15_count,
                'age_total': age_15_below_count + age_above_15_count,
                'total_animal_bite_I': barangay_animal_bite_I,
                'total_animal_bite_II': barangay_animal_bite_II,
                'total_animal_bite_III': barangay_animal_bite_III,
                'total_animal': barangay_animal_bite_I + barangay_animal_bite_II + barangay_animal_bite_III,
                'total_tcv_given': tcv_count,
                'total_hrig_given': hrig_count,
                'total_erig_given': erig_count,
                'total_dog_bites': barangay_dog_bites,
                'total_cat_bites': barangay_cat_bites,
                'total_other_bites': barangay_other_bites,
                'total_animal_bites': barangay_dog_bites + barangay_cat_bites + barangay_other_bites,
                'immunized_count': barangay_immunized_count,
                'unimmunized_count': barangay_unimmunized_count,
                'human_rabies_count': barangay_human_rabies_count,
            })

    if total_all > 0:
        male_percentage = (total_male / total_all) * 100
        female_percentage = (total_female / total_all) * 100
        total_sex_percentage = (male_percentage + female_percentage)
        age_15_below_percentage = (total_age_15_below / total_all) * 100
        age_above_15_percentage = (total_age_above_15 / total_all) * 100
        total_age_percentage = (total_age_15_below + total_age_above_15) / total_all * 100
        total_animal_bite_I_percentage = (total_animal_bite_I / total_all ) * 100
        total_animal_bite_II_percentage = (total_animal_bite_II / total_all ) * 100
        total_animal_bite_III_percentage = (total_animal_bite_III / total_all ) * 100
        total_category_percentage = (total_animal_bite_I_percentage + total_animal_bite_II_percentage + total_animal_bite_III_percentage)
        total_tcv_percentage = (total_tcv_given / total_all) * 100
        total_hrig_percentage = (total_hrig_given / total_all) * 100
        total_rig_percentage = (total_erig_given / total_all) * 100  
        dog_bite_percentage = (total_dog_bites / total_all) * 100
        cat_bite_percentage = (total_cat_bites / total_all) * 100
        other_bite_percentage = (total_other_bites / total_all) * 100
        total_animal_type_percentage = dog_bite_percentage + cat_bite_percentage + other_bite_percentage
        if total_immunized + total_unimmunized > 0:
            immunized_percentage = (total_immunized / (total_immunized + total_unimmunized)) * 100
            unimmunized_percentage = (total_unimmunized / (total_immunized + total_unimmunized)) * 100
        else:
            immunized_percentage = 0
            unimmunized_percentage = 0
    else:
        male_percentage = female_percentage = age_15_below_percentage = age_above_15_percentage = total_age_percentage = 0
        total_tcv_percentage = total_hrig_percentage = total_rig_percentage = 0
        dog_bite_percentage = cat_bite_percentage = other_bite_percentage = 0
        immunized_percentage = 0
        unimmunized_percentage = 0

    overall_total = sum(entry.get('data_total', 0) for entry in data)
    overall_total_tcv = sum(entry.get('total_tcv_given', 0) for entry in data)
    overall_total_hrig = sum(entry.get('total_hrig_given', 0) for entry in data)
    overall_total_erig = sum(entry.get('total_erig_given', 0) for entry in data)

    for entry in data:
        entry['percent_total'] = round((entry['data_total'] / overall_total) * 100, 1) if overall_total > 0 else 0
        entry['percent_tcv'] = round((entry.get('total_tcv_given', 0) / overall_total_tcv) * 100, 1) if overall_total_tcv > 0 else 0
        entry['percent_hrig'] = round((entry.get('total_hrig_given', 0) / overall_total_hrig) * 100, 1) if overall_total_hrig > 0 else 0
        entry['percent_erig'] = round((entry.get('total_erig_given', 0) / overall_total_erig) * 100, 1) if overall_total_erig > 0 else 0
    #all percent totals
    total_percent = sum(entry['percent_total'] for entry in data)
    total_tcv_percent = sum(entry['percent_tcv'] for entry in data)
    total_hrig_percent = sum(entry['percent_hrig'] for entry in data)
    total_erig_percent = sum(entry['percent_erig'] for entry in data)   
    total_human_rabies = sum(entry.get('human_rabies_count', 0) for entry in data)  # Add this line

    municipalities = Municipality.objects.all()
    selected_municipality_id = request.GET.get('municipality_id')
    selected_municipality = municipalities.filter(muni_id=selected_municipality_id).first() if selected_municipality_id else None
    
    if user.is_superuser:
        table = "ABTC"
    else:
        table = "Barangay"

    if user.first_name or user.last_name:
        signature_name = f"{user.first_name} {user.last_name}".strip()
    else:
        signature_name = user.username

    if user.logo_image:  # Ensure the user has a logo image
        logo_url = user.logo_image.url  # Access the image URL directly
    else:
        logo_url = None  # If no logo is available, set logo_url to None
    
    code_to_muni_name = {
        'ALM': 'Almeria',
        'BIL': 'Biliran',
        'CABUC': 'Cabucgayan',
        'CAIB': 'Caibiran',
        'CUL': 'Culaba',
        'KAW': 'Kawayan',
        'MAR': 'Maripipi',
        'NAV': 'Naval',
    }
    
    if user.is_superuser:
        coordinator = "Provincial Rabies Coordinator"
        pho = "PHO II"
        # Get the first lead doctor
        doct = Doctor.objects.filter(is_superdoctor=True).first()
        doctor = doct.full_name() if doct else "No Doctor Assigned"
    else:
        # Map user code to municipality name
        muni_name = code_to_muni_name.get(user.code)

        if muni_name:
            try:
                # Find the municipality with this name
                municipality = Municipality.objects.get(muni_name=muni_name)
                # Filter doctors by this municipality
                doct = Doctor.objects.filter(muni_id=municipality).first()
                doctor = doct.full_name() if doct else "No Doctor Assigned"
                coordinator = f"Municipality of {municipality.muni_name}"
                pho = f"Doctor of {municipality.muni_name}"
            except Municipality.DoesNotExist:
                # Handle case where the municipality name does not exist
                coordinator = "Unknown Municipality"
                pho = "No Doctor Assigned"
                doctor = "No Doctor Assigned"
        else:
            # Handle case where the user's code does not match any municipality
            coordinator = "Unknown Municipality"
            pho = "No Doctor Assigned"
            doctor = "No Doctor Assigned"
        
        

    karon = date.today().year
    context = {
        'doctor': doctor,
        'coordinator': coordinator,
        'pho': pho,
        'report_title': report_title,
        'karon': karon,
        'quarter_select': quarter_select,
        'logo_url': logo_url,
        'signature_name':signature_name,
        'table': table,
        'municipalities': municipalities,
        'selected_municipality': selected_municipality or municipalities.first(),  # Default to the first municipality if none is selected
        'municipality_name': municipality_name,
        'selected_quarter': selected_quarter,
        'barangay_list': [d['barangay'] for d in data],
        'data': data,
        'total_male': total_male,
        'total_female': total_female,
        'total_all': total_all,
        'total_age_15_below': total_age_15_below,
        'total_age_above_15': total_age_above_15,
        'male_percentage': round(male_percentage, 1),
        'female_percentage': round(female_percentage, 1),
        'total_sex_percentage':round(total_sex_percentage),
        'age_15_below_percentage': round(age_15_below_percentage, 1),
        'age_above_15_percentage': round(age_above_15_percentage, 1),
        'total_age_percentage': round(total_age_percentage, 1),
        'total_animal_bite_I': total_animal_bite_I,
        'total_animal_bite_II': total_animal_bite_II,
        'total_animal_bite_III': total_animal_bite_III,
        'total_animal_bite_I_percentage': round(total_animal_bite_I_percentage,1),
        'total_animal_bite_II_percentage': round(total_animal_bite_II_percentage,1),
        'total_animal_bite_III_percentage': round(total_animal_bite_III_percentage,1),
        'total_category_percentage': round(total_category_percentage,1),
        'overall_total':overall_total,
        'total_percent':round(total_percent, ),
        'total_tcv_given': total_tcv_given,
        'total_hrig_given': total_hrig_given,
        'total_rig_given': total_erig_given,
        'total_tcv_percentage': round(total_tcv_percentage, 1),
        'total_hrig_percentage': round(total_hrig_percentage, 1),
        'total_rig_percentage': round(total_rig_percentage, 1),
        'total_tcv_percent':round(total_tcv_percent, ),
        'total_erig_percent':round(total_erig_percent, ),
        'total_hrig_percent':round(total_hrig_percent, ),
        'total_dog_bites': total_dog_bites,
        'total_cat_bites': total_cat_bites,
        'total_other_bites': total_other_bites,
        'dog_bite_percentage': round(dog_bite_percentage, 1),
        'cat_bite_percentage': round(cat_bite_percentage, 1),
        'other_bite_percentage': round(other_bite_percentage, 1),
        'total_animal_type_percentage': round(total_animal_type_percentage, 1),
        'total_tcv_given': sum(entry.get('total_tcv_given', 0) for entry in data),
        'total_hrig_given': sum(entry.get('total_hrig_given', 0) for entry in data),
        'total_erig_given': sum(entry.get('total_erig_given', 0) for entry in data),
        'total_immunized': total_immunized,
        'total_unimmunized': total_unimmunized,
        'immunized_percentage': round(immunized_percentage, 1), 
        'unimmunized_percentage': round(unimmunized_percentage, 1),  
        'total_human_rabies':total_human_rabies,
    }
    return render(request, 'monitoring/report.html', context)

def pdf_report_create(request):
    user = request.user
    municipality_map = {
        "MAR": "Maripipi",
        "KAW": "Kawayan",
        "NAV": "Naval",
        "CAIB": "Caibiran",
        "ALM": "Almeria",
        "BIL": "Biliran",
        "CUL": "Culaba",
        "CABUC": "Cabucgayan"
    }
    if user.is_superuser and user.code == "NAV":
        municipality_name = "BPH"
    else:
        municipality_name = municipality_map.get(user.code, "Province")

    if user.is_superuser:
        municipal = "Province of Biliran"
    else:
        munici = municipality_map.get(user.code)
        municipal = (f"Municipality of {munici}")


    # Only set for 2nd quarter (April 1 to June 30)
    selected_quarter = '1'
    year = date.today().year
    start_date = date(year, 1, 1)
    end_date = date(year, 3, 31)


    data = []
    total_male = 0
    total_female = 0
    total_all = 0
    total_age_15_below = 0
    total_age_above_15 = 0
    total_sex_percentage = 0
    total_animal_bite_I = 0
    total_animal_bite_II = 0
    total_animal_bite_III = 0
    total_category_percentage = 0
    total_tcv_given = 0
    total_hrig_given = 0
    total_erig_given = 0  
    total_tcv_percentage = 0
    total_hrig_percentage = 0
    total_rig_percentage = 0
    total_dog_bites = 0
    total_cat_bites = 0
    total_other_bites = 0
    total_animal_type_percentage = 0
    total_animal_bite_I_percentage = 0
    total_animal_bite_II_percentage = 0
    total_animal_bite_III_percentage = 0
    total_immunized = 0
    total_unimmunized = 0

    if user.is_superuser:
        abtcs = User.objects.filter(is_superuser=False).distinct()
        for abtc_user in abtcs:
            if not abtc_user.code:
                continue

            male_count = History.objects.filter(
                patient_id__user=abtc_user,
                date_registered__range=(start_date, end_date),
                patient_id__sex='male'
            ).count()
            female_count = History.objects.filter(
                patient_id__user=abtc_user,
                date_registered__range=(start_date, end_date),
                patient_id__sex='female'
            ).count()

            patients = Patient.objects.filter(
                user=abtc_user,  
                histories__date_registered__range=(start_date, end_date)  
            ).distinct()
            age_15_below_count = sum(1 for patient in patients if calculate_age(patient.birthday) <= 15)
            age_above_15_count = sum(1 for patient in patients if calculate_age(patient.birthday) > 15)

            user_animal_bite_I = 0
            user_animal_bite_II = 0
            user_animal_bite_III = 0

            animal_bite_counts = History.objects.filter(
                patient_id__user=abtc_user,
                date_registered__range=(start_date, end_date)
            ).values('category_of_exposure').annotate(count=models.Count('category_of_exposure'))

            for count in animal_bite_counts:
                if count['category_of_exposure'] == 'I':
                    user_animal_bite_I = count['count']
                elif count['category_of_exposure'] == 'II':
                    user_animal_bite_II = count['count']
                elif count['category_of_exposure'] == 'III':
                    user_animal_bite_III = count['count']

            tcv_count = Treatment.objects.filter(
                patient_id__user=abtc_user,
                tcv_given__range=(start_date, end_date)
            ).count()

            hrig_count = Treatment.objects.filter(
                patient_id__user=abtc_user,
                hrig_given__range=(start_date, end_date)
            ).count()

            erig_count = Treatment.objects.filter(
                patient_id__user=abtc_user,
                rig_given__range=(start_date, end_date) 
            ).count()

            animal_type_counts = History.objects.filter(
                patient_id__user=abtc_user,
                date_registered__range=(start_date, end_date)
            ).values('source_of_exposure').annotate(count=models.Count('source_of_exposure'))

            user_dog_bites = 0
            user_cat_bites = 0
            user_other_bites = 0

            for count in animal_type_counts:
                if count['source_of_exposure'] == 'Dog':
                    user_dog_bites = count['count']
                elif count['source_of_exposure'] == 'Cat':
                    user_cat_bites = count['count']
                elif count['source_of_exposure'] == 'Others':
                    user_other_bites = count['count']

            user_immunized_count = History.objects.filter(
                patient_id__user=abtc_user,
                immunization_status='Immunized',
                date_registered__range=(start_date, end_date)
            ).count()
            
            user_unimmunized_count = History.objects.filter(
                patient_id__user=abtc_user,
                immunization_status='Unimmunized',
                date_registered__range=(start_date, end_date)
            ).count()

            user_human_rabies_count = History.objects.filter(
                patient_id__user=abtc_user,
                human_rabies=True,
                date_registered__range=(start_date, end_date)
            ).count()
                    
            total_immunized += user_immunized_count
            total_unimmunized += user_unimmunized_count
            total_dog_bites += user_dog_bites
            total_cat_bites += user_cat_bites
            total_other_bites += user_other_bites
            total_tcv_given += tcv_count
            total_hrig_given += hrig_count
            total_erig_given += erig_count

            total_count = male_count + female_count
            total_male += male_count
            total_female += female_count
            total_all += total_count
            total_age_15_below += age_15_below_count
            total_age_above_15 += age_above_15_count
            total_animal_bite_I += user_animal_bite_I
            total_animal_bite_II += user_animal_bite_II
            total_animal_bite_III += user_animal_bite_III

            # Determine barangay field for this ABTC user
            if abtc_user.code == "NAV":
                barangay_name = "BPH-ABTC"
            else:
                barangay_name = f"{municipality_map.get(abtc_user.code, 'Unknown')}-ABTC"

            data.append({
                'barangay': barangay_name,
                'data_male': male_count,
                'data_female': female_count,
                'data_total': total_count,
                'age_15_below': age_15_below_count,
                'age_above_15': age_above_15_count,
                'age_total': age_15_below_count + age_above_15_count,
                'total_animal_bite_I': user_animal_bite_I,
                'total_animal_bite_II': user_animal_bite_II,
                'total_animal_bite_III': user_animal_bite_III,
                'total_animal': user_animal_bite_I + user_animal_bite_II + user_animal_bite_III,
                'total_tcv_given': tcv_count,
                'total_hrig_given': hrig_count,
                'total_erig_given': erig_count,
                'total_dog_bites': user_dog_bites,
                'total_cat_bites': user_cat_bites,
                'total_other_bites': user_other_bites,
                'total_animal_bites': user_dog_bites + user_cat_bites + user_other_bites,
                'immunized_count': user_immunized_count,
                'unimmunized_count': user_unimmunized_count,  
                'human_rabies_count': user_human_rabies_count,
            })
    else:
        # For non-superuser
        patients = Patient.objects.filter(user=user)
        barangays = Barangay.objects.filter(patients_brgy__in=patients).distinct()

        # Filter barangays that have records in the 2nd quarter
        barangays_with_data = []
        for barangay in barangays:
            # Check if there's any patient in the barangay with a history in the 2nd quarter
            patient_data_exists = History.objects.filter(
                patient_id__brgy_id=barangay,
                patient_id__user=user,
                date_registered__range=(start_date, end_date)
            ).exists()
            
            if patient_data_exists:
                barangays_with_data.append(barangay)

        # Loop through barangays that have data in the 2nd quarter
        for barangay in barangays_with_data:
            male_count = History.objects.filter(
                patient_id__brgy_id=barangay,
                patient_id__user=user,
                date_registered__range=(start_date, end_date),
                patient_id__sex='male'
            ).count()
            female_count = History.objects.filter(
                patient_id__brgy_id=barangay,
                patient_id__user=user,
                date_registered__range=(start_date, end_date),
                patient_id__sex='female'
            ).count()

            patients = Patient.objects.filter(
                brgy_id=barangay,
                user=user,
                histories__date_registered__range=(start_date, end_date)  # Filter by registration date within the second quarter
            ).distinct()
            
            age_15_below_count = sum(1 for patient in patients if calculate_age(patient.birthday) <= 15)
            age_above_15_count = sum(1 for patient in patients if calculate_age(patient.birthday) > 15)

            # Initialize counts for animal bite categories
            barangay_animal_bite_I = 0
            barangay_animal_bite_II = 0
            barangay_animal_bite_III = 0

            animal_bite_counts = History.objects.filter(
                patient_id__brgy_id=barangay,
                patient_id__user=user,
                date_registered__range=(start_date, end_date)
            ).values('category_of_exposure').annotate(count=models.Count('category_of_exposure'))

            tcv_count = Treatment.objects.filter(
                patient_id__brgy_id=barangay,
                patient_id__user=user,
                tcv_given__range=(start_date, end_date)
            ).count()

            hrig_count = Treatment.objects.filter(
                patient_id__brgy_id=barangay,
                patient_id__user=user,
                hrig_given__range=(start_date, end_date)
            ).count()

            erig_count = Treatment.objects.filter(
                patient_id__brgy_id=barangay,
                patient_id__user=user,
                rig_given__range=(start_date, end_date)
            ).count()

            total_tcv_given += tcv_count
            total_hrig_given += hrig_count
            total_erig_given += erig_count

            for count in animal_bite_counts:
                if count['category_of_exposure'] == 'I':
                    barangay_animal_bite_I = count['count']
                elif count['category_of_exposure'] == 'II':
                    barangay_animal_bite_II = count['count']
                elif count['category_of_exposure'] == 'III':
                    barangay_animal_bite_III = count['count']

            animal_type_counts = History.objects.filter(
                patient_id__brgy_id=barangay,
                patient_id__user=user,
                date_registered__range=(start_date, end_date)
            ).values('source_of_exposure').annotate(count=models.Count('source_of_exposure'))

            barangay_dog_bites = 0
            barangay_cat_bites = 0
            barangay_other_bites = 0

            for count in animal_type_counts:
                if count['source_of_exposure'] == 'Dog':
                    barangay_dog_bites = count['count']
                elif count['source_of_exposure'] == 'Cat':
                    barangay_cat_bites = count['count']
                elif count['source_of_exposure'] == 'Others':
                    barangay_other_bites = count['count']

            barangay_immunized_count = History.objects.filter(
                patient_id__brgy_id=barangay,
                patient_id__user=user,
                immunization_status='Immunized',
                date_registered__range=(start_date, end_date)
            ).count()

            barangay_unimmunized_count = History.objects.filter(
                patient_id__brgy_id=barangay,
                patient_id__user=user,
                immunization_status='Unimmunized',
                date_registered__range=(start_date, end_date)
            ).count()

            barangay_human_rabies_count = History.objects.filter(
                patient_id__brgy_id=barangay,
                patient_id__user=user,
                human_rabies=True,
                date_registered__range=(start_date, end_date)
            ).count()

            total_immunized += barangay_immunized_count
            total_unimmunized += barangay_unimmunized_count
            total_dog_bites += barangay_dog_bites
            total_cat_bites += barangay_cat_bites
            total_other_bites += barangay_other_bites
            total_animal_bite_I += barangay_animal_bite_I
            total_animal_bite_II += barangay_animal_bite_II
            total_animal_bite_III += barangay_animal_bite_III

            # Add to total counts
            total_count = male_count + female_count
            total_male += male_count
            total_female += female_count
            total_all += total_count
            total_age_15_below += age_15_below_count
            total_age_above_15 += age_above_15_count

            data.append({
                'barangay': barangay.brgy_name,
                'data_male': male_count,
                'data_female': female_count,
                'data_total': total_count,
                'age_15_below': age_15_below_count,
                'age_above_15': age_above_15_count,
                'age_total': age_15_below_count + age_above_15_count,
                'total_animal_bite_I': barangay_animal_bite_I,
                'total_animal_bite_II': barangay_animal_bite_II,
                'total_animal_bite_III': barangay_animal_bite_III,
                'total_animal': barangay_animal_bite_I + barangay_animal_bite_II + barangay_animal_bite_III,
                'total_tcv_given': tcv_count,
                'total_hrig_given': hrig_count,
                'total_erig_given': erig_count,
                'total_dog_bites': barangay_dog_bites,
                'total_cat_bites': barangay_cat_bites,
                'total_other_bites': barangay_other_bites,
                'total_animal_bites': barangay_dog_bites + barangay_cat_bites + barangay_other_bites,
                'immunized_count': barangay_immunized_count,
                'unimmunized_count': barangay_unimmunized_count,
                'human_rabies_count': barangay_human_rabies_count,
            })
    if total_all > 0:
        male_percentage = (total_male / total_all) * 100
        female_percentage = (total_female / total_all) * 100
        total_sex_percentage = (male_percentage + female_percentage)
        age_15_below_percentage = (total_age_15_below / total_all) * 100
        age_above_15_percentage = (total_age_above_15 / total_all) * 100
        total_age_percentage = (total_age_15_below + total_age_above_15) / total_all * 100
        total_animal_bite_I_percentage = (total_animal_bite_I / total_all ) * 100
        total_animal_bite_II_percentage = (total_animal_bite_II / total_all ) * 100
        total_animal_bite_III_percentage = (total_animal_bite_III / total_all ) * 100
        total_category_percentage = (total_animal_bite_I_percentage + total_animal_bite_II_percentage + total_animal_bite_III_percentage)
        total_tcv_percentage = (total_tcv_given / total_all) * 100
        total_hrig_percentage = (total_hrig_given / total_all) * 100
        total_rig_percentage = (total_erig_given / total_all) * 100  
        dog_bite_percentage = (total_dog_bites / total_all) * 100
        cat_bite_percentage = (total_cat_bites / total_all) * 100
        other_bite_percentage = (total_other_bites / total_all) * 100
        total_animal_type_percentage = dog_bite_percentage + cat_bite_percentage + other_bite_percentage
        if total_immunized + total_unimmunized > 0:
            immunized_percentage = (total_immunized / (total_immunized + total_unimmunized)) * 100
            unimmunized_percentage = (total_unimmunized / (total_immunized + total_unimmunized)) * 100
        else:
            immunized_percentage = 0
            unimmunized_percentage = 0
    else:
        male_percentage = female_percentage = age_15_below_percentage = age_above_15_percentage = total_age_percentage = 0
        total_tcv_percentage = total_hrig_percentage = total_rig_percentage = 0
        dog_bite_percentage = cat_bite_percentage = other_bite_percentage = 0
        immunized_percentage = 0
        unimmunized_percentage = 0

    overall_total = sum(entry.get('data_total', 0) for entry in data)
    overall_total_tcv = sum(entry.get('total_tcv_given', 0) for entry in data)
    overall_total_hrig = sum(entry.get('total_hrig_given', 0) for entry in data)
    overall_total_erig = sum(entry.get('total_erig_given', 0) for entry in data)

    for entry in data:
        entry['percent_total'] = round((entry['data_total'] / overall_total) * 100, 1) if overall_total > 0 else 0
        entry['percent_tcv'] = round((entry.get('total_tcv_given', 0) / overall_total_tcv) * 100, 1) if overall_total_tcv > 0 else 0
        entry['percent_hrig'] = round((entry.get('total_hrig_given', 0) / overall_total_hrig) * 100, 1) if overall_total_hrig > 0 else 0
        entry['percent_erig'] = round((entry.get('total_erig_given', 0) / overall_total_erig) * 100, 1) if overall_total_erig > 0 else 0
    total_percent = sum(entry['percent_total'] for entry in data)
    total_tcv_percent = sum(entry['percent_tcv'] for entry in data)
    total_hrig_percent = sum(entry['percent_hrig'] for entry in data)
    total_erig_percent = sum(entry['percent_erig'] for entry in data)   
    total_human_rabies = sum(entry.get('human_rabies_count', 0) for entry in data)  # Add this line
    municipalities = Municipality.objects.all()
    selected_municipality_id = request.GET.get('municipality_id')
    selected_municipality = municipalities.filter(muni_id=selected_municipality_id).first() if selected_municipality_id else None
    if user.is_superuser:
        table = "ABTC"
    else:
        table = "Barangay"

    if user.first_name or user.last_name:
        signature_name = f"{user.first_name} {user.last_name}".strip()
    else:
        signature_name = user.username

    if user.logo_image:  # Ensure the user has a logo image
        logo_url = request.build_absolute_uri(user.logo_image.url)
    else:
        logo_url = None  # If no logo is available, set logo_url to None

    karon = date.today().year

    code_to_muni_name = {
        'ALM': 'Almeria',
        'BIL': 'Biliran',
        'CABUC': 'Cabucgayan',
        'CAIB': 'Caibiran',
        'CUL': 'Culaba',
        'KAW': 'Kawayan',
        'MAR': 'Maripipi',
        'NAV': 'Naval',
    }

    if user.is_superuser:
        coordinator = "Provincial Rabies Coordinator"
        pho = "PHO II"
        # Get the first lead doctor
        doct = Doctor.objects.filter(is_superdoctor=True).first()
        doctor = doct.full_name() if doct else "No Doctor Assigned"
        center = "Biliran Province Hospital"
        center_label = f"Animal Bite Treatment Center: {center}"  # Add the label
    else:
        # Map user code to municipality name
        muni_name = code_to_muni_name.get(user.code)

        if muni_name:
            try:
                # Find the municipality with this name
                municipality = Municipality.objects.get(muni_name=muni_name)
                # Filter doctors by this municipality
                doct = Doctor.objects.filter(muni_id=municipality).first()
                doctor = doct.full_name() if doct else "No Doctor Assigned"
                coordinator = f"Municipality of {municipality.muni_name}"
                pho = f"Doctor of {municipality.muni_name}"
                center = muni_name  # Use the user's code to determine the center
                center_label = f"Animal Bite Treatment Center: {center} Animal Bite Treatment Center"  # Add the label
            except Municipality.DoesNotExist:
                # Handle case where the municipality name does not exist
                coordinator = "Unknown Municipality"
                pho = "No Doctor Assigned"
                doctor = "No Doctor Assigned"
                center = "Unknown Center"
                center_label = f"Animal Bite Treatment Center: {center}"  # Add the label
        else:
            # Handle case where the user's code does not match any municipality
            coordinator = "Unknown Municipality"
            pho = "No Doctor Assigned"
            doctor = "No Doctor Assigned"
            center = "Unknown Center"
            center_label = f"Animal Bite Treatment Center: {center} Animal Bite Treatment Center"  # Add the label

        
    
    context = {
        'doctor': doctor,
        'center': center,
        'center_label': center_label,
        'coordinator': coordinator,
        'pho': pho,
        'karon': karon,    
        'logo_url': logo_url,    
        'signature_name':signature_name,
        'table': table,
        'municipalities': municipalities,
        'selected_municipality': selected_municipality or municipalities.first(),  # Default to the first municipality if none is selected
        'municipality_name': municipality_name,
        'municipal': municipal,
        'selected_quarter': selected_quarter,
        'barangay_list': [d['barangay'] for d in data],
        'data': data,
        'total_male': total_male,
        'total_female': total_female,
        'total_all': total_all,
        'total_age_15_below': total_age_15_below,
        'total_age_above_15': total_age_above_15,
        'male_percentage': round(male_percentage, 1),
        'female_percentage': round(female_percentage, 1),
        'total_sex_percentage':round(total_sex_percentage),
        'age_15_below_percentage': round(age_15_below_percentage, 1),
        'age_above_15_percentage': round(age_above_15_percentage, 1),
        'total_age_percentage': round(total_age_percentage, 1),
        'total_animal_bite_I': total_animal_bite_I,
        'total_animal_bite_II': total_animal_bite_II,
        'total_animal_bite_III': total_animal_bite_III,
        'total_animal_bite_I_percentage': round(total_animal_bite_I_percentage,1),
        'total_animal_bite_II_percentage': round(total_animal_bite_II_percentage,1),
        'total_animal_bite_III_percentage': round(total_animal_bite_III_percentage,1),
        'total_category_percentage': round(total_category_percentage,1),
        'overall_total':overall_total,
        'total_percent':round(total_percent, ),
        'total_tcv_given': total_tcv_given,
        'total_hrig_given': total_hrig_given,
        'total_rig_given': total_erig_given,
        'total_tcv_percentage': round(total_tcv_percentage, 1),
        'total_hrig_percentage': round(total_hrig_percentage, 1),
        'total_rig_percentage': round(total_rig_percentage, 1),
        'total_tcv_percent':round(total_tcv_percent, ),
        'total_erig_percent':round(total_erig_percent, ),
        'total_hrig_percent':round(total_hrig_percent, ),
        'total_dog_bites': total_dog_bites,
        'total_cat_bites': total_cat_bites,
        'total_other_bites': total_other_bites,
        'dog_bite_percentage': round(dog_bite_percentage, 1),
        'cat_bite_percentage': round(cat_bite_percentage, 1),
        'other_bite_percentage': round(other_bite_percentage, 1),
        'total_animal_type_percentage': round(total_animal_type_percentage, 1),
        'total_tcv_given': sum(entry.get('total_tcv_given', 0) for entry in data),
        'total_hrig_given': sum(entry.get('total_hrig_given', 0) for entry in data),
        'total_erig_given': sum(entry.get('total_erig_given', 0) for entry in data),
        'total_immunized': total_immunized,
        'total_unimmunized': total_unimmunized,
        'immunized_percentage': round(immunized_percentage, 1), 
        'unimmunized_percentage': round(unimmunized_percentage, 1),  
        'total_human_rabies':total_human_rabies,

    }
    template = loader.get_template('monitoring/report_pdf.html')
    html = template.render(context)
    
    """ # Specify path to wkhtmltopdf executable
    path_to_wkhtmltopdf = r"C:\wkhtmltox\bin\wkhtmltopdf.exe"
    
    # Configure pdfkit with the path
    config = pdfkit.configuration(wkhtmltopdf=path_to_wkhtmltopdf) """

    config = pdfkit.configuration(wkhtmltopdf=WKHTMLTOPDF_PATH)

    # Options for landscape orientation
    options = {
        'orientation': 'Landscape',  # Set to Landscape
        'page-size': 'A4',  # A4 paper size, you can adjust if needed
        'footer-left': center_label,  # Add page numbers
        'footer-right': 'Page [page] of [toPage]',  # Add page numbers
        'footer-font-size': '10',  # Adjust font size for the footer
        'footer-spacing': '5',  # Space between footer and content
        'margin-bottom': '15mm',  # Ensure space for the footer
    }
    
    # Generate PDF from HTML using pdfkit with configuration
    pdf = pdfkit.from_string(html, configuration=config,options=options)

    # Return the PDF as a responsev 
    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment;filename="1st Quarter Report.pdf"'

    return response

def pdf_report_create2(request):
    user = request.user
    municipality_map = {
        "MAR": "Maripipi",
        "KAW": "Kawayan",
        "NAV": "Naval",
        "CAIB": "Caibiran",
        "ALM": "Almeria",
        "BIL": "Biliran",
        "CUL": "Culaba",
        "CABUC": "Cabucgayan"
    }
    if user.is_superuser and user.code == "NAV":
        municipality_name = "BPH"
    else:
        municipality_name = municipality_map.get(user.code, "Province")

    if user.is_superuser:
        municipal = "Province of Biliran"
    else:
        munici = municipality_map.get(user.code)
        municipal = (f"Municipality of {munici}")

    # Only set for 2nd quarter (April 1 to June 30)
    selected_quarter = '2'
    year = date.today().year
    start_date = date(year, 4, 1)
    end_date = date(year, 6, 30)

    data = []
    total_male = 0
    total_female = 0
    total_all = 0
    total_age_15_below = 0
    total_age_above_15 = 0
    total_sex_percentage = 0
    total_animal_bite_I = 0
    total_animal_bite_II = 0
    total_animal_bite_III = 0
    total_category_percentage = 0
    total_tcv_given = 0
    total_hrig_given = 0
    total_erig_given = 0  
    total_tcv_percentage = 0
    total_hrig_percentage = 0
    total_rig_percentage = 0
    total_dog_bites = 0
    total_cat_bites = 0
    total_other_bites = 0
    total_animal_type_percentage = 0
    total_animal_bite_I_percentage = 0
    total_animal_bite_II_percentage = 0
    total_animal_bite_III_percentage = 0
    total_immunized = 0
    total_unimmunized = 0

    if user.is_superuser:
        abtcs = User.objects.filter(is_superuser=False).distinct()
        for abtc_user in abtcs:
            if not abtc_user.code:
                continue

            male_count = History.objects.filter(
                patient_id__user=abtc_user,
                date_registered__range=(start_date, end_date),
                patient_id__sex='male'
            ).count()
            female_count = History.objects.filter(
                patient_id__user=abtc_user,
                date_registered__range=(start_date, end_date),
                patient_id__sex='female'
            ).count()

            patients = Patient.objects.filter(
                user=abtc_user,  
                histories__date_registered__range=(start_date, end_date)  
            ).distinct()
            age_15_below_count = sum(1 for patient in patients if calculate_age(patient.birthday) <= 15)
            age_above_15_count = sum(1 for patient in patients if calculate_age(patient.birthday) > 15)

            user_animal_bite_I = 0
            user_animal_bite_II = 0
            user_animal_bite_III = 0

            animal_bite_counts = History.objects.filter(
                patient_id__user=abtc_user,
                date_registered__range=(start_date, end_date)
            ).values('category_of_exposure').annotate(count=models.Count('category_of_exposure'))

            for count in animal_bite_counts:
                if count['category_of_exposure'] == 'I':
                    user_animal_bite_I = count['count']
                elif count['category_of_exposure'] == 'II':
                    user_animal_bite_II = count['count']
                elif count['category_of_exposure'] == 'III':
                    user_animal_bite_III = count['count']

            tcv_count = Treatment.objects.filter(
                patient_id__user=abtc_user,
                tcv_given__range=(start_date, end_date)
            ).count()

            hrig_count = Treatment.objects.filter(
                patient_id__user=abtc_user,
                hrig_given__range=(start_date, end_date)
            ).count()

            erig_count = Treatment.objects.filter(
                patient_id__user=abtc_user,
                rig_given__range=(start_date, end_date) 
            ).count()

            animal_type_counts = History.objects.filter(
                patient_id__user=abtc_user,
                date_registered__range=(start_date, end_date)
            ).values('source_of_exposure').annotate(count=models.Count('source_of_exposure'))

            user_dog_bites = 0
            user_cat_bites = 0
            user_other_bites = 0

            for count in animal_type_counts:
                if count['source_of_exposure'] == 'Dog':
                    user_dog_bites = count['count']
                elif count['source_of_exposure'] == 'Cat':
                    user_cat_bites = count['count']
                elif count['source_of_exposure'] == 'Others':
                    user_other_bites = count['count']

            user_immunized_count = History.objects.filter(
                patient_id__user=abtc_user,
                immunization_status='Immunized',
                date_registered__range=(start_date, end_date)
            ).count()
            
            user_unimmunized_count = History.objects.filter(
                patient_id__user=abtc_user,
                immunization_status='Unimmunized',
                date_registered__range=(start_date, end_date)
            ).count()

            user_human_rabies_count = History.objects.filter(
                patient_id__user=abtc_user,
                human_rabies=True,
                date_registered__range=(start_date, end_date)
            ).count()
                    
            total_immunized += user_immunized_count
            total_unimmunized += user_unimmunized_count
            total_dog_bites += user_dog_bites
            total_cat_bites += user_cat_bites
            total_other_bites += user_other_bites
            total_tcv_given += tcv_count
            total_hrig_given += hrig_count
            total_erig_given += erig_count

            total_count = male_count + female_count
            total_male += male_count
            total_female += female_count
            total_all += total_count
            total_age_15_below += age_15_below_count
            total_age_above_15 += age_above_15_count
            total_animal_bite_I += user_animal_bite_I
            total_animal_bite_II += user_animal_bite_II
            total_animal_bite_III += user_animal_bite_III

            # Determine barangay field for this ABTC user
            if abtc_user.code == "NAV":
                barangay_name = "BPH-ABTC"
            else:
                barangay_name = f"{municipality_map.get(abtc_user.code, 'Unknown')}-ABTC"

            data.append({
                'barangay': barangay_name,
                'data_male': male_count,
                'data_female': female_count,
                'data_total': total_count,
                'age_15_below': age_15_below_count,
                'age_above_15': age_above_15_count,
                'age_total': age_15_below_count + age_above_15_count,
                'total_animal_bite_I': user_animal_bite_I,
                'total_animal_bite_II': user_animal_bite_II,
                'total_animal_bite_III': user_animal_bite_III,
                'total_animal': user_animal_bite_I + user_animal_bite_II + user_animal_bite_III,
                'total_tcv_given': tcv_count,
                'total_hrig_given': hrig_count,
                'total_erig_given': erig_count,
                'total_dog_bites': user_dog_bites,
                'total_cat_bites': user_cat_bites,
                'total_other_bites': user_other_bites,
                'total_animal_bites': user_dog_bites + user_cat_bites + user_other_bites,
                'immunized_count': user_immunized_count,
                'unimmunized_count': user_unimmunized_count,  
                'human_rabies_count': user_human_rabies_count,
            })
    else:
        # For non-superuser
        patients = Patient.objects.filter(user=user)
        barangays = Barangay.objects.filter(patients_brgy__in=patients).distinct()

        # Filter barangays that have records in the 2nd quarter
        barangays_with_data = []
        for barangay in barangays:
            # Check if there's any patient in the barangay with a history in the 2nd quarter
            patient_data_exists = History.objects.filter(
                patient_id__brgy_id=barangay,
                patient_id__user=user,
                date_registered__range=(start_date, end_date)
            ).exists()
            
            if patient_data_exists:
                barangays_with_data.append(barangay)

        # Loop through barangays that have data in the 2nd quarter
        for barangay in barangays_with_data:
            male_count = History.objects.filter(
                patient_id__brgy_id=barangay,
                patient_id__user=user,
                date_registered__range=(start_date, end_date),
                patient_id__sex='male'
            ).count()
            female_count = History.objects.filter(
                patient_id__brgy_id=barangay,
                patient_id__user=user,
                date_registered__range=(start_date, end_date),
                patient_id__sex='female'
            ).count()

            patients = Patient.objects.filter(
                brgy_id=barangay,
                user=user,
                histories__date_registered__range=(start_date, end_date)  # Filter by registration date within the second quarter
            ).distinct()
            
            age_15_below_count = sum(1 for patient in patients if calculate_age(patient.birthday) <= 15)
            age_above_15_count = sum(1 for patient in patients if calculate_age(patient.birthday) > 15)

            # Initialize counts for animal bite categories
            barangay_animal_bite_I = 0
            barangay_animal_bite_II = 0
            barangay_animal_bite_III = 0

            animal_bite_counts = History.objects.filter(
                patient_id__brgy_id=barangay,
                patient_id__user=user,
                date_registered__range=(start_date, end_date)
            ).values('category_of_exposure').annotate(count=models.Count('category_of_exposure'))

            tcv_count = Treatment.objects.filter(
                patient_id__brgy_id=barangay,
                patient_id__user=user,
                tcv_given__range=(start_date, end_date)
            ).count()

            hrig_count = Treatment.objects.filter(
                patient_id__brgy_id=barangay,
                patient_id__user=user,
                hrig_given__range=(start_date, end_date)
            ).count()

            erig_count = Treatment.objects.filter(
                patient_id__brgy_id=barangay,
                patient_id__user=user,
                rig_given__range=(start_date, end_date)
            ).count()

            total_tcv_given += tcv_count
            total_hrig_given += hrig_count
            total_erig_given += erig_count

            for count in animal_bite_counts:
                if count['category_of_exposure'] == 'I':
                    barangay_animal_bite_I = count['count']
                elif count['category_of_exposure'] == 'II':
                    barangay_animal_bite_II = count['count']
                elif count['category_of_exposure'] == 'III':
                    barangay_animal_bite_III = count['count']

            animal_type_counts = History.objects.filter(
                patient_id__brgy_id=barangay,
                patient_id__user=user,
                date_registered__range=(start_date, end_date)
            ).values('source_of_exposure').annotate(count=models.Count('source_of_exposure'))

            barangay_dog_bites = 0
            barangay_cat_bites = 0
            barangay_other_bites = 0

            for count in animal_type_counts:
                if count['source_of_exposure'] == 'Dog':
                    barangay_dog_bites = count['count']
                elif count['source_of_exposure'] == 'Cat':
                    barangay_cat_bites = count['count']
                elif count['source_of_exposure'] == 'Others':
                    barangay_other_bites = count['count']

            barangay_immunized_count = History.objects.filter(
                patient_id__brgy_id=barangay,
                patient_id__user=user,
                immunization_status='Immunized',
                date_registered__range=(start_date, end_date)
            ).count()

            barangay_unimmunized_count = History.objects.filter(
                patient_id__brgy_id=barangay,
                patient_id__user=user,
                immunization_status='Unimmunized',
                date_registered__range=(start_date, end_date)
            ).count()

            barangay_human_rabies_count = History.objects.filter(
                patient_id__brgy_id=barangay,
                patient_id__user=user,
                human_rabies=True,
                date_registered__range=(start_date, end_date)
            ).count()

            total_immunized += barangay_immunized_count
            total_unimmunized += barangay_unimmunized_count
            total_dog_bites += barangay_dog_bites
            total_cat_bites += barangay_cat_bites
            total_other_bites += barangay_other_bites
            total_animal_bite_I += barangay_animal_bite_I
            total_animal_bite_II += barangay_animal_bite_II
            total_animal_bite_III += barangay_animal_bite_III

            # Add to total counts
            total_count = male_count + female_count
            total_male += male_count
            total_female += female_count
            total_all += total_count
            total_age_15_below += age_15_below_count
            total_age_above_15 += age_above_15_count

            data.append({
                'barangay': barangay.brgy_name,
                'data_male': male_count,
                'data_female': female_count,
                'data_total': total_count,
                'age_15_below': age_15_below_count,
                'age_above_15': age_above_15_count,
                'age_total': age_15_below_count + age_above_15_count,
                'total_animal_bite_I': barangay_animal_bite_I,
                'total_animal_bite_II': barangay_animal_bite_II,
                'total_animal_bite_III': barangay_animal_bite_III,
                'total_animal': barangay_animal_bite_I + barangay_animal_bite_II + barangay_animal_bite_III,
                'total_tcv_given': tcv_count,
                'total_hrig_given': hrig_count,
                'total_erig_given': erig_count,
                'total_dog_bites': barangay_dog_bites,
                'total_cat_bites': barangay_cat_bites,
                'total_other_bites': barangay_other_bites,
                'total_animal_bites': barangay_dog_bites + barangay_cat_bites + barangay_other_bites,
                'immunized_count': barangay_immunized_count,
                'unimmunized_count': barangay_unimmunized_count,
                'human_rabies_count': barangay_human_rabies_count,
            })
    if total_all > 0:
        male_percentage = (total_male / total_all) * 100
        female_percentage = (total_female / total_all) * 100
        total_sex_percentage = (male_percentage + female_percentage)
        age_15_below_percentage = (total_age_15_below / total_all) * 100
        age_above_15_percentage = (total_age_above_15 / total_all) * 100
        total_age_percentage = (total_age_15_below + total_age_above_15) / total_all * 100
        total_animal_bite_I_percentage = (total_animal_bite_I / total_all ) * 100
        total_animal_bite_II_percentage = (total_animal_bite_II / total_all ) * 100
        total_animal_bite_III_percentage = (total_animal_bite_III / total_all ) * 100
        total_category_percentage = (total_animal_bite_I_percentage + total_animal_bite_II_percentage + total_animal_bite_III_percentage)
        total_tcv_percentage = (total_tcv_given / total_all) * 100
        total_hrig_percentage = (total_hrig_given / total_all) * 100
        total_rig_percentage = (total_erig_given / total_all) * 100  
        dog_bite_percentage = (total_dog_bites / total_all) * 100
        cat_bite_percentage = (total_cat_bites / total_all) * 100
        other_bite_percentage = (total_other_bites / total_all) * 100
        total_animal_type_percentage = dog_bite_percentage + cat_bite_percentage + other_bite_percentage
        if total_immunized + total_unimmunized > 0:
            immunized_percentage = (total_immunized / (total_immunized + total_unimmunized)) * 100
            unimmunized_percentage = (total_unimmunized / (total_immunized + total_unimmunized)) * 100
        else:
            immunized_percentage = 0
            unimmunized_percentage = 0
    else:
        male_percentage = female_percentage = age_15_below_percentage = age_above_15_percentage = total_age_percentage = 0
        total_tcv_percentage = total_hrig_percentage = total_rig_percentage = 0
        dog_bite_percentage = cat_bite_percentage = other_bite_percentage = 0
        immunized_percentage = 0
        unimmunized_percentage = 0

    overall_total = sum(entry.get('data_total', 0) for entry in data)
    overall_total_tcv = sum(entry.get('total_tcv_given', 0) for entry in data)
    overall_total_hrig = sum(entry.get('total_hrig_given', 0) for entry in data)
    overall_total_erig = sum(entry.get('total_erig_given', 0) for entry in data)

    for entry in data:
        entry['percent_total'] = round((entry['data_total'] / overall_total) * 100, 1) if overall_total > 0 else 0
        entry['percent_tcv'] = round((entry.get('total_tcv_given', 0) / overall_total_tcv) * 100, 1) if overall_total_tcv > 0 else 0
        entry['percent_hrig'] = round((entry.get('total_hrig_given', 0) / overall_total_hrig) * 100, 1) if overall_total_hrig > 0 else 0
        entry['percent_erig'] = round((entry.get('total_erig_given', 0) / overall_total_erig) * 100, 1) if overall_total_erig > 0 else 0
    total_percent = sum(entry['percent_total'] for entry in data)
    total_tcv_percent = sum(entry['percent_tcv'] for entry in data)
    total_hrig_percent = sum(entry['percent_hrig'] for entry in data)
    total_erig_percent = sum(entry['percent_erig'] for entry in data)   
    total_human_rabies = sum(entry.get('human_rabies_count', 0) for entry in data)  # Add this line
    municipalities = Municipality.objects.all()
    selected_municipality_id = request.GET.get('municipality_id')
    selected_municipality = municipalities.filter(muni_id=selected_municipality_id).first() if selected_municipality_id else None
    if user.is_superuser:
        table = "ABTC"
    else:
        table = "Barangay"

    if user.first_name or user.last_name:
        signature_name = f"{user.first_name} {user.last_name}".strip()
    else:
        signature_name = user.username

    if user.logo_image:  # Ensure the user has a logo image
        logo_url = request.build_absolute_uri(user.logo_image.url)
    else:
        logo_url = None  # If no logo is available, set logo_url to None

    karon = date.today().year

    code_to_muni_name = {
        'ALM': 'Almeria',
        'BIL': 'Biliran',
        'CABUC': 'Cabucgayan',
        'CAIB': 'Caibiran',
        'CUL': 'Culaba',
        'KAW': 'Kawayan',
        'MAR': 'Maripipi',
        'NAV': 'Naval',
    }

    if user.is_superuser:
        coordinator = "Provincial Rabies Coordinator"
        pho = "PHO II"
        # Get the first lead doctor
        doct = Doctor.objects.filter(is_superdoctor=True).first()
        doctor = doct.full_name() if doct else "No Doctor Assigned"
        center = "Biliran Province Hospital"
        center_label = f"Animal Bite Treatment Center: {center}"  # Add the label
    else:
        # Map user code to municipality name
        muni_name = code_to_muni_name.get(user.code)

        if muni_name:
            try:
                # Find the municipality with this name
                municipality = Municipality.objects.get(muni_name=muni_name)
                # Filter doctors by this municipality
                doct = Doctor.objects.filter(muni_id=municipality).first()
                doctor = doct.full_name() if doct else "No Doctor Assigned"
                coordinator = f"Municipality of {municipality.muni_name}"
                pho = f"Doctor of {municipality.muni_name}"
                center = muni_name  # Use the user's code to determine the center
                center_label = f"Animal Bite Treatment Center: {center} Animal Bite Treatment Center"  # Add the label
            except Municipality.DoesNotExist:
                # Handle case where the municipality name does not exist
                coordinator = "Unknown Municipality"
                pho = "No Doctor Assigned"
                doctor = "No Doctor Assigned"
                center = "Unknown Center"
                center_label = f"Animal Bite Treatment Center: {center}"  # Add the label
        else:
            # Handle case where the user's code does not match any municipality
            coordinator = "Unknown Municipality"
            pho = "No Doctor Assigned"
            doctor = "No Doctor Assigned"
            center = "Unknown Center"
            center_label = f"Animal Bite Treatment Center: {center} Animal Bite Treatment Center"  # Add the label

    context = {
        'doctor': doctor,
        'center': center,
        'center_label': center_label,
        'coordinator': coordinator,
        'pho': pho,
        'karon': karon,    
        'logo_url': logo_url,    
        'signature_name':signature_name,
        'table': table,
        'municipalities': municipalities,
        'selected_municipality': selected_municipality or municipalities.first(),  # Default to the first municipality if none is selected
        'municipality_name': municipality_name,
        'municipal': municipal,
        'selected_quarter': selected_quarter,
        'barangay_list': [d['barangay'] for d in data],
        'data': data,
        'total_male': total_male,
        'total_female': total_female,
        'total_all': total_all,
        'total_age_15_below': total_age_15_below,
        'total_age_above_15': total_age_above_15,
        'male_percentage': round(male_percentage, 1),
        'female_percentage': round(female_percentage, 1),
        'total_sex_percentage':round(total_sex_percentage),
        'age_15_below_percentage': round(age_15_below_percentage, 1),
        'age_above_15_percentage': round(age_above_15_percentage, 1),
        'total_age_percentage': round(total_age_percentage, 1),
        'total_animal_bite_I': total_animal_bite_I,
        'total_animal_bite_II': total_animal_bite_II,
        'total_animal_bite_III': total_animal_bite_III,
        'total_animal_bite_I_percentage': round(total_animal_bite_I_percentage,1),
        'total_animal_bite_II_percentage': round(total_animal_bite_II_percentage,1),
        'total_animal_bite_III_percentage': round(total_animal_bite_III_percentage,1),
        'total_category_percentage': round(total_category_percentage,1),
        'overall_total':overall_total,
        'total_percent':round(total_percent, ),
        'total_tcv_given': total_tcv_given,
        'total_hrig_given': total_hrig_given,
        'total_rig_given': total_erig_given,
        'total_tcv_percentage': round(total_tcv_percentage, 1),
        'total_hrig_percentage': round(total_hrig_percentage, 1),
        'total_rig_percentage': round(total_rig_percentage, 1),
        'total_tcv_percent':round(total_tcv_percent, ),
        'total_erig_percent':round(total_erig_percent, ),
        'total_hrig_percent':round(total_hrig_percent, ),
        'total_dog_bites': total_dog_bites,
        'total_cat_bites': total_cat_bites,
        'total_other_bites': total_other_bites,
        'dog_bite_percentage': round(dog_bite_percentage, 1),
        'cat_bite_percentage': round(cat_bite_percentage, 1),
        'other_bite_percentage': round(other_bite_percentage, 1),
        'total_animal_type_percentage': round(total_animal_type_percentage, 1),
        'total_tcv_given': sum(entry.get('total_tcv_given', 0) for entry in data),
        'total_hrig_given': sum(entry.get('total_hrig_given', 0) for entry in data),
        'total_erig_given': sum(entry.get('total_erig_given', 0) for entry in data),
        'total_immunized': total_immunized,
        'total_unimmunized': total_unimmunized,
        'immunized_percentage': round(immunized_percentage, 1), 
        'unimmunized_percentage': round(unimmunized_percentage, 1),  
        'total_human_rabies':total_human_rabies,

    }
    template = loader.get_template('monitoring/report_pdf2.html')
    html = template.render(context)

    """ # Specify path to wkhtmltopdf executable
    path_to_wkhtmltopdf = r"C:\wkhtmltox\bin\wkhtmltopdf.exe"
    
    # Configure pdfkit with the path
    config = pdfkit.configuration(wkhtmltopdf=path_to_wkhtmltopdf) """

    config = pdfkit.configuration(wkhtmltopdf=WKHTMLTOPDF_PATH)

    # Options for landscape orientation
    options = {
        'orientation': 'Landscape',  # Set to Landscape
        'page-size': 'A4',  # A4 paper size, you can adjust if needed
        'footer-left': center_label,  # Add page numbers
        'footer-right': 'Page [page] of [toPage]',  # Add page numbers
        'footer-font-size': '10',  # Adjust font size for the footer
        'footer-spacing': '5',  # Space between footer and content
        'margin-bottom': '15mm',  # Ensure space for the footer
    }
    
    # Generate PDF from HTML using pdfkit with configuration
    pdf = pdfkit.from_string(html, configuration=config,options=options)

    # Return the PDF as a responsev 
    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="2nd Quarter Report.pdf"'

    return response

def pdf_report_create3(request):
    user = request.user
    municipality_map = {
        "MAR": "Maripipi",
        "KAW": "Kawayan",
        "NAV": "Naval",
        "CAIB": "Caibiran",
        "ALM": "Almeria",
        "BIL": "Biliran",
        "CUL": "Culaba",
        "CABUC": "Cabucgayan"
    }
    if user.is_superuser and user.code == "NAV":
        municipality_name = "BPH"
    else:
        municipality_name = municipality_map.get(user.code, "Province")

    if user.is_superuser:
        municipal = "Province of Biliran"
    else:
        munici = municipality_map.get(user.code)
        municipal = (f"Municipality of {munici}")

    # Only set for 2nd quarter (April 1 to June 30)
    selected_quarter = '3'
    year = date.today().year
    # 3rd Quarter
    start_date = date(year, 7, 1)
    end_date = date(year, 9, 30)
    


    data = []
    total_male = 0
    total_female = 0
    total_all = 0
    total_age_15_below = 0
    total_age_above_15 = 0
    total_sex_percentage = 0
    total_animal_bite_I = 0
    total_animal_bite_II = 0
    total_animal_bite_III = 0
    total_category_percentage = 0
    total_tcv_given = 0
    total_hrig_given = 0
    total_erig_given = 0  
    total_tcv_percentage = 0
    total_hrig_percentage = 0
    total_rig_percentage = 0
    total_dog_bites = 0
    total_cat_bites = 0
    total_other_bites = 0
    total_animal_type_percentage = 0
    total_animal_bite_I_percentage = 0
    total_animal_bite_II_percentage = 0
    total_animal_bite_III_percentage = 0
    total_immunized = 0
    total_unimmunized = 0

    if user.is_superuser:
        abtcs = User.objects.filter(is_superuser=False).distinct()
        for abtc_user in abtcs:
            if not abtc_user.code:
                continue

            male_count = History.objects.filter(
                patient_id__user=abtc_user,
                date_registered__range=(start_date, end_date),
                patient_id__sex='male'
            ).count()
            female_count = History.objects.filter(
                patient_id__user=abtc_user,
                date_registered__range=(start_date, end_date),
                patient_id__sex='female'
            ).count()

            patients = Patient.objects.filter(
                user=abtc_user,  
                histories__date_registered__range=(start_date, end_date)  
            ).distinct()
            age_15_below_count = sum(1 for patient in patients if calculate_age(patient.birthday) <= 15)
            age_above_15_count = sum(1 for patient in patients if calculate_age(patient.birthday) > 15)

            user_animal_bite_I = 0
            user_animal_bite_II = 0
            user_animal_bite_III = 0

            animal_bite_counts = History.objects.filter(
                patient_id__user=abtc_user,
                date_registered__range=(start_date, end_date)
            ).values('category_of_exposure').annotate(count=models.Count('category_of_exposure'))

            for count in animal_bite_counts:
                if count['category_of_exposure'] == 'I':
                    user_animal_bite_I = count['count']
                elif count['category_of_exposure'] == 'II':
                    user_animal_bite_II = count['count']
                elif count['category_of_exposure'] == 'III':
                    user_animal_bite_III = count['count']

            tcv_count = Treatment.objects.filter(
                patient_id__user=abtc_user,
                tcv_given__range=(start_date, end_date)
            ).count()

            hrig_count = Treatment.objects.filter(
                patient_id__user=abtc_user,
                hrig_given__range=(start_date, end_date)
            ).count()

            erig_count = Treatment.objects.filter(
                patient_id__user=abtc_user,
                rig_given__range=(start_date, end_date) 
            ).count()

            animal_type_counts = History.objects.filter(
                patient_id__user=abtc_user,
                date_registered__range=(start_date, end_date)
            ).values('source_of_exposure').annotate(count=models.Count('source_of_exposure'))

            user_dog_bites = 0
            user_cat_bites = 0
            user_other_bites = 0

            for count in animal_type_counts:
                if count['source_of_exposure'] == 'Dog':
                    user_dog_bites = count['count']
                elif count['source_of_exposure'] == 'Cat':
                    user_cat_bites = count['count']
                elif count['source_of_exposure'] == 'Others':
                    user_other_bites = count['count']

            user_immunized_count = History.objects.filter(
                patient_id__user=abtc_user,
                immunization_status='Immunized',
                date_registered__range=(start_date, end_date)
            ).count()
            
            user_unimmunized_count = History.objects.filter(
                patient_id__user=abtc_user,
                immunization_status='Unimmunized',
                date_registered__range=(start_date, end_date)
            ).count()

            user_human_rabies_count = History.objects.filter(
                patient_id__user=abtc_user,
                human_rabies=True,
                date_registered__range=(start_date, end_date)
            ).count()
                    
            total_immunized += user_immunized_count
            total_unimmunized += user_unimmunized_count
            total_dog_bites += user_dog_bites
            total_cat_bites += user_cat_bites
            total_other_bites += user_other_bites
            total_tcv_given += tcv_count
            total_hrig_given += hrig_count
            total_erig_given += erig_count

            total_count = male_count + female_count
            total_male += male_count
            total_female += female_count
            total_all += total_count
            total_age_15_below += age_15_below_count
            total_age_above_15 += age_above_15_count
            total_animal_bite_I += user_animal_bite_I
            total_animal_bite_II += user_animal_bite_II
            total_animal_bite_III += user_animal_bite_III

            # Determine barangay field for this ABTC user
            if abtc_user.code == "NAV":
                barangay_name = "BPH-ABTC"
            else:
                barangay_name = f"{municipality_map.get(abtc_user.code, 'Unknown')}-ABTC"

            data.append({
                'barangay': barangay_name,
                'data_male': male_count,
                'data_female': female_count,
                'data_total': total_count,
                'age_15_below': age_15_below_count,
                'age_above_15': age_above_15_count,
                'age_total': age_15_below_count + age_above_15_count,
                'total_animal_bite_I': user_animal_bite_I,
                'total_animal_bite_II': user_animal_bite_II,
                'total_animal_bite_III': user_animal_bite_III,
                'total_animal': user_animal_bite_I + user_animal_bite_II + user_animal_bite_III,
                'total_tcv_given': tcv_count,
                'total_hrig_given': hrig_count,
                'total_erig_given': erig_count,
                'total_dog_bites': user_dog_bites,
                'total_cat_bites': user_cat_bites,
                'total_other_bites': user_other_bites,
                'total_animal_bites': user_dog_bites + user_cat_bites + user_other_bites,
                'immunized_count': user_immunized_count,
                'unimmunized_count': user_unimmunized_count,  
                'human_rabies_count': user_human_rabies_count,
            })
    else:
        # For non-superuser
        patients = Patient.objects.filter(user=user)
        barangays = Barangay.objects.filter(patients_brgy__in=patients).distinct()

        # Filter barangays that have records in the 2nd quarter
        barangays_with_data = []
        for barangay in barangays:
            # Check if there's any patient in the barangay with a history in the 2nd quarter
            patient_data_exists = History.objects.filter(
                patient_id__brgy_id=barangay,
                patient_id__user=user,
                date_registered__range=(start_date, end_date)
            ).exists()
            
            if patient_data_exists:
                barangays_with_data.append(barangay)

        # Loop through barangays that have data in the 2nd quarter
        for barangay in barangays_with_data:
            male_count = History.objects.filter(
                patient_id__brgy_id=barangay,
                patient_id__user=user,
                date_registered__range=(start_date, end_date),
                patient_id__sex='male'
            ).count()
            female_count = History.objects.filter(
                patient_id__brgy_id=barangay,
                patient_id__user=user,
                date_registered__range=(start_date, end_date),
                patient_id__sex='female'
            ).count()

            patients = Patient.objects.filter(
                brgy_id=barangay,
                user=user,
                histories__date_registered__range=(start_date, end_date)  # Filter by registration date within the second quarter
            ).distinct()
            
            age_15_below_count = sum(1 for patient in patients if calculate_age(patient.birthday) <= 15)
            age_above_15_count = sum(1 for patient in patients if calculate_age(patient.birthday) > 15)

            # Initialize counts for animal bite categories
            barangay_animal_bite_I = 0
            barangay_animal_bite_II = 0
            barangay_animal_bite_III = 0

            animal_bite_counts = History.objects.filter(
                patient_id__brgy_id=barangay,
                patient_id__user=user,
                date_registered__range=(start_date, end_date)
            ).values('category_of_exposure').annotate(count=models.Count('category_of_exposure'))

            tcv_count = Treatment.objects.filter(
                patient_id__brgy_id=barangay,
                patient_id__user=user,
                tcv_given__range=(start_date, end_date)
            ).count()

            hrig_count = Treatment.objects.filter(
                patient_id__brgy_id=barangay,
                patient_id__user=user,
                hrig_given__range=(start_date, end_date)
            ).count()

            erig_count = Treatment.objects.filter(
                patient_id__brgy_id=barangay,
                patient_id__user=user,
                rig_given__range=(start_date, end_date)
            ).count()

            total_tcv_given += tcv_count
            total_hrig_given += hrig_count
            total_erig_given += erig_count

            for count in animal_bite_counts:
                if count['category_of_exposure'] == 'I':
                    barangay_animal_bite_I = count['count']
                elif count['category_of_exposure'] == 'II':
                    barangay_animal_bite_II = count['count']
                elif count['category_of_exposure'] == 'III':
                    barangay_animal_bite_III = count['count']

            animal_type_counts = History.objects.filter(
                patient_id__brgy_id=barangay,
                patient_id__user=user,
                date_registered__range=(start_date, end_date)
            ).values('source_of_exposure').annotate(count=models.Count('source_of_exposure'))

            barangay_dog_bites = 0
            barangay_cat_bites = 0
            barangay_other_bites = 0

            for count in animal_type_counts:
                if count['source_of_exposure'] == 'Dog':
                    barangay_dog_bites = count['count']
                elif count['source_of_exposure'] == 'Cat':
                    barangay_cat_bites = count['count']
                elif count['source_of_exposure'] == 'Others':
                    barangay_other_bites = count['count']

            barangay_immunized_count = History.objects.filter(
                patient_id__brgy_id=barangay,
                patient_id__user=user,
                immunization_status='Immunized',
                date_registered__range=(start_date, end_date)
            ).count()

            barangay_unimmunized_count = History.objects.filter(
                patient_id__brgy_id=barangay,
                patient_id__user=user,
                immunization_status='Unimmunized',
                date_registered__range=(start_date, end_date)
            ).count()

            barangay_human_rabies_count = History.objects.filter(
                patient_id__brgy_id=barangay,
                patient_id__user=user,
                human_rabies=True,
                date_registered__range=(start_date, end_date)
            ).count()

            total_immunized += barangay_immunized_count
            total_unimmunized += barangay_unimmunized_count
            total_dog_bites += barangay_dog_bites
            total_cat_bites += barangay_cat_bites
            total_other_bites += barangay_other_bites
            total_animal_bite_I += barangay_animal_bite_I
            total_animal_bite_II += barangay_animal_bite_II
            total_animal_bite_III += barangay_animal_bite_III

            # Add to total counts
            total_count = male_count + female_count
            total_male += male_count
            total_female += female_count
            total_all += total_count
            total_age_15_below += age_15_below_count
            total_age_above_15 += age_above_15_count

            data.append({
                'barangay': barangay.brgy_name,
                'data_male': male_count,
                'data_female': female_count,
                'data_total': total_count,
                'age_15_below': age_15_below_count,
                'age_above_15': age_above_15_count,
                'age_total': age_15_below_count + age_above_15_count,
                'total_animal_bite_I': barangay_animal_bite_I,
                'total_animal_bite_II': barangay_animal_bite_II,
                'total_animal_bite_III': barangay_animal_bite_III,
                'total_animal': barangay_animal_bite_I + barangay_animal_bite_II + barangay_animal_bite_III,
                'total_tcv_given': tcv_count,
                'total_hrig_given': hrig_count,
                'total_erig_given': erig_count,
                'total_dog_bites': barangay_dog_bites,
                'total_cat_bites': barangay_cat_bites,
                'total_other_bites': barangay_other_bites,
                'total_animal_bites': barangay_dog_bites + barangay_cat_bites + barangay_other_bites,
                'immunized_count': barangay_immunized_count,
                'unimmunized_count': barangay_unimmunized_count,
                'human_rabies_count': barangay_human_rabies_count,
            })
    if total_all > 0:
        male_percentage = (total_male / total_all) * 100
        female_percentage = (total_female / total_all) * 100
        total_sex_percentage = (male_percentage + female_percentage)
        age_15_below_percentage = (total_age_15_below / total_all) * 100
        age_above_15_percentage = (total_age_above_15 / total_all) * 100
        total_age_percentage = (total_age_15_below + total_age_above_15) / total_all * 100
        total_animal_bite_I_percentage = (total_animal_bite_I / total_all ) * 100
        total_animal_bite_II_percentage = (total_animal_bite_II / total_all ) * 100
        total_animal_bite_III_percentage = (total_animal_bite_III / total_all ) * 100
        total_category_percentage = (total_animal_bite_I_percentage + total_animal_bite_II_percentage + total_animal_bite_III_percentage)
        total_tcv_percentage = (total_tcv_given / total_all) * 100
        total_hrig_percentage = (total_hrig_given / total_all) * 100
        total_rig_percentage = (total_erig_given / total_all) * 100  
        dog_bite_percentage = (total_dog_bites / total_all) * 100
        cat_bite_percentage = (total_cat_bites / total_all) * 100
        other_bite_percentage = (total_other_bites / total_all) * 100
        total_animal_type_percentage = dog_bite_percentage + cat_bite_percentage + other_bite_percentage
        if total_immunized + total_unimmunized > 0:
            immunized_percentage = (total_immunized / (total_immunized + total_unimmunized)) * 100
            unimmunized_percentage = (total_unimmunized / (total_immunized + total_unimmunized)) * 100
        else:
            immunized_percentage = 0
            unimmunized_percentage = 0
    else:
        male_percentage = female_percentage = age_15_below_percentage = age_above_15_percentage = total_age_percentage = 0
        total_tcv_percentage = total_hrig_percentage = total_rig_percentage = 0
        dog_bite_percentage = cat_bite_percentage = other_bite_percentage = 0
        immunized_percentage = 0
        unimmunized_percentage = 0

    overall_total = sum(entry.get('data_total', 0) for entry in data)
    overall_total_tcv = sum(entry.get('total_tcv_given', 0) for entry in data)
    overall_total_hrig = sum(entry.get('total_hrig_given', 0) for entry in data)
    overall_total_erig = sum(entry.get('total_erig_given', 0) for entry in data)

    for entry in data:
        entry['percent_total'] = round((entry['data_total'] / overall_total) * 100, 1) if overall_total > 0 else 0
        entry['percent_tcv'] = round((entry.get('total_tcv_given', 0) / overall_total_tcv) * 100, 1) if overall_total_tcv > 0 else 0
        entry['percent_hrig'] = round((entry.get('total_hrig_given', 0) / overall_total_hrig) * 100, 1) if overall_total_hrig > 0 else 0
        entry['percent_erig'] = round((entry.get('total_erig_given', 0) / overall_total_erig) * 100, 1) if overall_total_erig > 0 else 0
    total_percent = sum(entry['percent_total'] for entry in data)
    total_tcv_percent = sum(entry['percent_tcv'] for entry in data)
    total_hrig_percent = sum(entry['percent_hrig'] for entry in data)
    total_erig_percent = sum(entry['percent_erig'] for entry in data)   
    total_human_rabies = sum(entry.get('human_rabies_count', 0) for entry in data)  # Add this line
    municipalities = Municipality.objects.all()
    selected_municipality_id = request.GET.get('municipality_id')
    selected_municipality = municipalities.filter(muni_id=selected_municipality_id).first() if selected_municipality_id else None
    if user.is_superuser:
        table = "ABTC"
    else:
        table = "Barangay"

    if user.first_name or user.last_name:
        signature_name = f"{user.first_name} {user.last_name}".strip()
    else:
        signature_name = user.username

    if user.logo_image:  # Ensure the user has a logo image
        logo_url = request.build_absolute_uri(user.logo_image.url)
    else:
        logo_url = None  # If no logo is available, set logo_url to None

    karon = date.today().year

    code_to_muni_name = {
        'ALM': 'Almeria',
        'BIL': 'Biliran',
        'CABUC': 'Cabucgayan',
        'CAIB': 'Caibiran',
        'CUL': 'Culaba',
        'KAW': 'Kawayan',
        'MAR': 'Maripipi',
        'NAV': 'Naval',
    }

    if user.is_superuser:
        coordinator = "Provincial Rabies Coordinator"
        pho = "PHO II"
        # Get the first lead doctor
        doct = Doctor.objects.filter(is_superdoctor=True).first()
        doctor = doct.full_name() if doct else "No Doctor Assigned"
        center = "Biliran Province Hospital"
        center_label = f"Animal Bite Treatment Center: {center}"  # Add the label
    else:
        # Map user code to municipality name
        muni_name = code_to_muni_name.get(user.code)

        if muni_name:
            try:
                # Find the municipality with this name
                municipality = Municipality.objects.get(muni_name=muni_name)
                # Filter doctors by this municipality
                doct = Doctor.objects.filter(muni_id=municipality).first()
                doctor = doct.full_name() if doct else "No Doctor Assigned"
                coordinator = f"Municipality of {municipality.muni_name}"
                pho = f"Doctor of {municipality.muni_name}"
                center = muni_name  # Use the user's code to determine the center
                center_label = f"Animal Bite Treatment Center: {center} Animal Bite Treatment Center"  # Add the label
            except Municipality.DoesNotExist:
                # Handle case where the municipality name does not exist
                coordinator = "Unknown Municipality"
                pho = "No Doctor Assigned"
                doctor = "No Doctor Assigned"
                center = "Unknown Center"
                center_label = f"Animal Bite Treatment Center: {center}"  # Add the label
        else:
            # Handle case where the user's code does not match any municipality
            coordinator = "Unknown Municipality"
            pho = "No Doctor Assigned"
            doctor = "No Doctor Assigned"
            center = "Unknown Center"
            center_label = f"Animal Bite Treatment Center: {center} Animal Bite Treatment Center"  # Add the label


    context = {
        'doctor': doctor,
        'center': center,
        'center_label': center_label,
        'coordinator': coordinator,
        'pho': pho,
        'karon': karon,    
        'logo_url': logo_url,    
        'signature_name':signature_name,
        'table': table,
        'municipalities': municipalities,
        'selected_municipality': selected_municipality or municipalities.first(),  # Default to the first municipality if none is selected
        'municipality_name': municipality_name,
        'municipal': municipal,
        'selected_quarter': selected_quarter,
        'barangay_list': [d['barangay'] for d in data],
        'data': data,
        'total_male': total_male,
        'total_female': total_female,
        'total_all': total_all,
        'total_age_15_below': total_age_15_below,
        'total_age_above_15': total_age_above_15,
        'male_percentage': round(male_percentage, 1),
        'female_percentage': round(female_percentage, 1),
        'total_sex_percentage':round(total_sex_percentage),
        'age_15_below_percentage': round(age_15_below_percentage, 1),
        'age_above_15_percentage': round(age_above_15_percentage, 1),
        'total_age_percentage': round(total_age_percentage, 1),
        'total_animal_bite_I': total_animal_bite_I,
        'total_animal_bite_II': total_animal_bite_II,
        'total_animal_bite_III': total_animal_bite_III,
        'total_animal_bite_I_percentage': round(total_animal_bite_I_percentage,1),
        'total_animal_bite_II_percentage': round(total_animal_bite_II_percentage,1),
        'total_animal_bite_III_percentage': round(total_animal_bite_III_percentage,1),
        'total_category_percentage': round(total_category_percentage,1),
        'overall_total':overall_total,
        'total_percent':round(total_percent, ),
        'total_tcv_given': total_tcv_given,
        'total_hrig_given': total_hrig_given,
        'total_rig_given': total_erig_given,
        'total_tcv_percentage': round(total_tcv_percentage, 1),
        'total_hrig_percentage': round(total_hrig_percentage, 1),
        'total_rig_percentage': round(total_rig_percentage, 1),
        'total_tcv_percent':round(total_tcv_percent, ),
        'total_erig_percent':round(total_erig_percent, ),
        'total_hrig_percent':round(total_hrig_percent, ),
        'total_dog_bites': total_dog_bites,
        'total_cat_bites': total_cat_bites,
        'total_other_bites': total_other_bites,
        'dog_bite_percentage': round(dog_bite_percentage, 1),
        'cat_bite_percentage': round(cat_bite_percentage, 1),
        'other_bite_percentage': round(other_bite_percentage, 1),
        'total_animal_type_percentage': round(total_animal_type_percentage, 1),
        'total_tcv_given': sum(entry.get('total_tcv_given', 0) for entry in data),
        'total_hrig_given': sum(entry.get('total_hrig_given', 0) for entry in data),
        'total_erig_given': sum(entry.get('total_erig_given', 0) for entry in data),
        'total_immunized': total_immunized,
        'total_unimmunized': total_unimmunized,
        'immunized_percentage': round(immunized_percentage, 1), 
        'unimmunized_percentage': round(unimmunized_percentage, 1),  
        'total_human_rabies':total_human_rabies,

    }
    template = loader.get_template('monitoring/report_pdf3.html')
    html = template.render(context)

    """ # Specify path to wkhtmltopdf executable
    path_to_wkhtmltopdf = r"C:\wkhtmltox\bin\wkhtmltopdf.exe"
    
    # Configure pdfkit with the path
    config = pdfkit.configuration(wkhtmltopdf=path_to_wkhtmltopdf) """

    config = pdfkit.configuration(wkhtmltopdf=WKHTMLTOPDF_PATH)

    # Options for landscape orientation
    options = {
        'orientation': 'Landscape',  # Set to Landscape
        'page-size': 'A4',  # A4 paper size, you can adjust if needed
        'footer-left': center_label,  # Add page numbers
        'footer-right': 'Page [page] of [toPage]',  # Add page numbers
        'footer-font-size': '10',  # Adjust font size for the footer
        'footer-spacing': '5',  # Space between footer and content
        'margin-bottom': '15mm',  # Ensure space for the footer
    }
    
    # Generate PDF from HTML using pdfkit with configuration
    pdf = pdfkit.from_string(html, configuration=config,options=options)

    # Return the PDF as a responsev 
    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment;filename="3rd Quarter Report.pdf"'

    return response

def pdf_report_create4(request):
    user = request.user
    municipality_map = {
        "MAR": "Maripipi",
        "KAW": "Kawayan",
        "NAV": "Naval",
        "CAIB": "Caibiran",
        "ALM": "Almeria",
        "BIL": "Biliran",
        "CUL": "Culaba",
        "CABUC": "Cabucgayan"
    }
    if user.is_superuser and user.code == "NAV":
        municipality_name = "BPH"
    else:
        municipality_name = municipality_map.get(user.code, "Province")

    if user.is_superuser:
        municipal = "Province of Biliran"
    else:
        munici = municipality_map.get(user.code)
        municipal = (f"Municipality of {munici}")
    # Only set for 2nd quarter (April 1 to June 30)
    selected_quarter = '4'
    year = date.today().year
    # 4th Quarter
    start_date = date(year, 10, 1)
    end_date = date(year, 12, 31)
    

    data = []
    total_male = 0
    total_female = 0
    total_all = 0
    total_age_15_below = 0
    total_age_above_15 = 0
    total_sex_percentage = 0
    total_animal_bite_I = 0
    total_animal_bite_II = 0
    total_animal_bite_III = 0
    total_category_percentage = 0
    total_tcv_given = 0
    total_hrig_given = 0
    total_erig_given = 0  
    total_tcv_percentage = 0
    total_hrig_percentage = 0
    total_rig_percentage = 0
    total_dog_bites = 0
    total_cat_bites = 0
    total_other_bites = 0
    total_animal_type_percentage = 0
    total_animal_bite_I_percentage = 0
    total_animal_bite_II_percentage = 0
    total_animal_bite_III_percentage = 0
    total_immunized = 0
    total_unimmunized = 0

    if user.is_superuser:
        abtcs = User.objects.filter(is_superuser=False).distinct()
        for abtc_user in abtcs:
            if not abtc_user.code:
                continue

            male_count = History.objects.filter(
                patient_id__user=abtc_user,
                date_registered__range=(start_date, end_date),
                patient_id__sex='male'
            ).count()
            female_count = History.objects.filter(
                patient_id__user=abtc_user,
                date_registered__range=(start_date, end_date),
                patient_id__sex='female'
            ).count()

            patients = Patient.objects.filter(
                user=abtc_user,  
                histories__date_registered__range=(start_date, end_date)  
            ).distinct()
            age_15_below_count = sum(1 for patient in patients if calculate_age(patient.birthday) <= 15)
            age_above_15_count = sum(1 for patient in patients if calculate_age(patient.birthday) > 15)

            user_animal_bite_I = 0
            user_animal_bite_II = 0
            user_animal_bite_III = 0

            animal_bite_counts = History.objects.filter(
                patient_id__user=abtc_user,
                date_registered__range=(start_date, end_date)
            ).values('category_of_exposure').annotate(count=models.Count('category_of_exposure'))

            for count in animal_bite_counts:
                if count['category_of_exposure'] == 'I':
                    user_animal_bite_I = count['count']
                elif count['category_of_exposure'] == 'II':
                    user_animal_bite_II = count['count']
                elif count['category_of_exposure'] == 'III':
                    user_animal_bite_III = count['count']

            tcv_count = Treatment.objects.filter(
                patient_id__user=abtc_user,
                tcv_given__range=(start_date, end_date)
            ).count()

            hrig_count = Treatment.objects.filter(
                patient_id__user=abtc_user,
                hrig_given__range=(start_date, end_date)
            ).count()

            erig_count = Treatment.objects.filter(
                patient_id__user=abtc_user,
                rig_given__range=(start_date, end_date) 
            ).count()

            animal_type_counts = History.objects.filter(
                patient_id__user=abtc_user,
                date_registered__range=(start_date, end_date)
            ).values('source_of_exposure').annotate(count=models.Count('source_of_exposure'))

            user_dog_bites = 0
            user_cat_bites = 0
            user_other_bites = 0

            for count in animal_type_counts:
                if count['source_of_exposure'] == 'Dog':
                    user_dog_bites = count['count']
                elif count['source_of_exposure'] == 'Cat':
                    user_cat_bites = count['count']
                elif count['source_of_exposure'] == 'Others':
                    user_other_bites = count['count']

            user_immunized_count = History.objects.filter(
                patient_id__user=abtc_user,
                immunization_status='Immunized',
                date_registered__range=(start_date, end_date)
            ).count()
            
            user_unimmunized_count = History.objects.filter(
                patient_id__user=abtc_user,
                immunization_status='Unimmunized',
                date_registered__range=(start_date, end_date)
            ).count()

            user_human_rabies_count = History.objects.filter(
                patient_id__user=abtc_user,
                human_rabies=True,
                date_registered__range=(start_date, end_date)
            ).count()
                    
            total_immunized += user_immunized_count
            total_unimmunized += user_unimmunized_count
            total_dog_bites += user_dog_bites
            total_cat_bites += user_cat_bites
            total_other_bites += user_other_bites
            total_tcv_given += tcv_count
            total_hrig_given += hrig_count
            total_erig_given += erig_count

            total_count = male_count + female_count
            total_male += male_count
            total_female += female_count
            total_all += total_count
            total_age_15_below += age_15_below_count
            total_age_above_15 += age_above_15_count
            total_animal_bite_I += user_animal_bite_I
            total_animal_bite_II += user_animal_bite_II
            total_animal_bite_III += user_animal_bite_III

            # Determine barangay field for this ABTC user
            if abtc_user.code == "NAV":
                barangay_name = "BPH-ABTC"
            else:
                barangay_name = f"{municipality_map.get(abtc_user.code, 'Unknown')}-ABTC"

            data.append({
                'barangay': barangay_name,
                'data_male': male_count,
                'data_female': female_count,
                'data_total': total_count,
                'age_15_below': age_15_below_count,
                'age_above_15': age_above_15_count,
                'age_total': age_15_below_count + age_above_15_count,
                'total_animal_bite_I': user_animal_bite_I,
                'total_animal_bite_II': user_animal_bite_II,
                'total_animal_bite_III': user_animal_bite_III,
                'total_animal': user_animal_bite_I + user_animal_bite_II + user_animal_bite_III,
                'total_tcv_given': tcv_count,
                'total_hrig_given': hrig_count,
                'total_erig_given': erig_count,
                'total_dog_bites': user_dog_bites,
                'total_cat_bites': user_cat_bites,
                'total_other_bites': user_other_bites,
                'total_animal_bites': user_dog_bites + user_cat_bites + user_other_bites,
                'immunized_count': user_immunized_count,
                'unimmunized_count': user_unimmunized_count,  
                'human_rabies_count': user_human_rabies_count,
            })
    else:
        # For non-superuser
        patients = Patient.objects.filter(user=user)
        barangays = Barangay.objects.filter(patients_brgy__in=patients).distinct()

        # Filter barangays that have records in the 2nd quarter
        barangays_with_data = []
        for barangay in barangays:
            # Check if there's any patient in the barangay with a history in the 2nd quarter
            patient_data_exists = History.objects.filter(
                patient_id__brgy_id=barangay,
                patient_id__user=user,
                date_registered__range=(start_date, end_date)
            ).exists()
            
            if patient_data_exists:
                barangays_with_data.append(barangay)

        # Loop through barangays that have data in the 2nd quarter
        for barangay in barangays_with_data:
            male_count = History.objects.filter(
                patient_id__brgy_id=barangay,
                patient_id__user=user,
                date_registered__range=(start_date, end_date),
                patient_id__sex='male'
            ).count()
            female_count = History.objects.filter(
                patient_id__brgy_id=barangay,
                patient_id__user=user,
                date_registered__range=(start_date, end_date),
                patient_id__sex='female'
            ).count()

            patients = Patient.objects.filter(
                brgy_id=barangay,
                user=user,
                histories__date_registered__range=(start_date, end_date)  # Filter by registration date within the second quarter
            ).distinct()
            
            age_15_below_count = sum(1 for patient in patients if calculate_age(patient.birthday) <= 15)
            age_above_15_count = sum(1 for patient in patients if calculate_age(patient.birthday) > 15)

            # Initialize counts for animal bite categories
            barangay_animal_bite_I = 0
            barangay_animal_bite_II = 0
            barangay_animal_bite_III = 0

            animal_bite_counts = History.objects.filter(
                patient_id__brgy_id=barangay,
                patient_id__user=user,
                date_registered__range=(start_date, end_date)
            ).values('category_of_exposure').annotate(count=models.Count('category_of_exposure'))

            tcv_count = Treatment.objects.filter(
                patient_id__brgy_id=barangay,
                patient_id__user=user,
                tcv_given__range=(start_date, end_date)
            ).count()

            hrig_count = Treatment.objects.filter(
                patient_id__brgy_id=barangay,
                patient_id__user=user,
                hrig_given__range=(start_date, end_date)
            ).count()

            erig_count = Treatment.objects.filter(
                patient_id__brgy_id=barangay,
                patient_id__user=user,
                rig_given__range=(start_date, end_date)
            ).count()

            total_tcv_given += tcv_count
            total_hrig_given += hrig_count
            total_erig_given += erig_count

            for count in animal_bite_counts:
                if count['category_of_exposure'] == 'I':
                    barangay_animal_bite_I = count['count']
                elif count['category_of_exposure'] == 'II':
                    barangay_animal_bite_II = count['count']
                elif count['category_of_exposure'] == 'III':
                    barangay_animal_bite_III = count['count']

            animal_type_counts = History.objects.filter(
                patient_id__brgy_id=barangay,
                patient_id__user=user,
                date_registered__range=(start_date, end_date)
            ).values('source_of_exposure').annotate(count=models.Count('source_of_exposure'))

            barangay_dog_bites = 0
            barangay_cat_bites = 0
            barangay_other_bites = 0

            for count in animal_type_counts:
                if count['source_of_exposure'] == 'Dog':
                    barangay_dog_bites = count['count']
                elif count['source_of_exposure'] == 'Cat':
                    barangay_cat_bites = count['count']
                elif count['source_of_exposure'] == 'Others':
                    barangay_other_bites = count['count']

            barangay_immunized_count = History.objects.filter(
                patient_id__brgy_id=barangay,
                patient_id__user=user,
                immunization_status='Immunized',
                date_registered__range=(start_date, end_date)
            ).count()

            barangay_unimmunized_count = History.objects.filter(
                patient_id__brgy_id=barangay,
                patient_id__user=user,
                immunization_status='Unimmunized',
                date_registered__range=(start_date, end_date)
            ).count()

            barangay_human_rabies_count = History.objects.filter(
                patient_id__brgy_id=barangay,
                patient_id__user=user,
                human_rabies=True,
                date_registered__range=(start_date, end_date)
            ).count()

            total_immunized += barangay_immunized_count
            total_unimmunized += barangay_unimmunized_count
            total_dog_bites += barangay_dog_bites
            total_cat_bites += barangay_cat_bites
            total_other_bites += barangay_other_bites
            total_animal_bite_I += barangay_animal_bite_I
            total_animal_bite_II += barangay_animal_bite_II
            total_animal_bite_III += barangay_animal_bite_III

            # Add to total counts
            total_count = male_count + female_count
            total_male += male_count
            total_female += female_count
            total_all += total_count
            total_age_15_below += age_15_below_count
            total_age_above_15 += age_above_15_count

            data.append({
                'barangay': barangay.brgy_name,
                'data_male': male_count,
                'data_female': female_count,
                'data_total': total_count,
                'age_15_below': age_15_below_count,
                'age_above_15': age_above_15_count,
                'age_total': age_15_below_count + age_above_15_count,
                'total_animal_bite_I': barangay_animal_bite_I,
                'total_animal_bite_II': barangay_animal_bite_II,
                'total_animal_bite_III': barangay_animal_bite_III,
                'total_animal': barangay_animal_bite_I + barangay_animal_bite_II + barangay_animal_bite_III,
                'total_tcv_given': tcv_count,
                'total_hrig_given': hrig_count,
                'total_erig_given': erig_count,
                'total_dog_bites': barangay_dog_bites,
                'total_cat_bites': barangay_cat_bites,
                'total_other_bites': barangay_other_bites,
                'total_animal_bites': barangay_dog_bites + barangay_cat_bites + barangay_other_bites,
                'immunized_count': barangay_immunized_count,
                'unimmunized_count': barangay_unimmunized_count,
                'human_rabies_count': barangay_human_rabies_count,
            })
    if total_all > 0:
        male_percentage = (total_male / total_all) * 100
        female_percentage = (total_female / total_all) * 100
        total_sex_percentage = (male_percentage + female_percentage)
        age_15_below_percentage = (total_age_15_below / total_all) * 100
        age_above_15_percentage = (total_age_above_15 / total_all) * 100
        total_age_percentage = (total_age_15_below + total_age_above_15) / total_all * 100
        total_animal_bite_I_percentage = (total_animal_bite_I / total_all ) * 100
        total_animal_bite_II_percentage = (total_animal_bite_II / total_all ) * 100
        total_animal_bite_III_percentage = (total_animal_bite_III / total_all ) * 100
        total_category_percentage = (total_animal_bite_I_percentage + total_animal_bite_II_percentage + total_animal_bite_III_percentage)
        total_tcv_percentage = (total_tcv_given / total_all) * 100
        total_hrig_percentage = (total_hrig_given / total_all) * 100
        total_rig_percentage = (total_erig_given / total_all) * 100  
        dog_bite_percentage = (total_dog_bites / total_all) * 100
        cat_bite_percentage = (total_cat_bites / total_all) * 100
        other_bite_percentage = (total_other_bites / total_all) * 100
        total_animal_type_percentage = dog_bite_percentage + cat_bite_percentage + other_bite_percentage
        if total_immunized + total_unimmunized > 0:
            immunized_percentage = (total_immunized / (total_immunized + total_unimmunized)) * 100
            unimmunized_percentage = (total_unimmunized / (total_immunized + total_unimmunized)) * 100
        else:
            immunized_percentage = 0
            unimmunized_percentage = 0
    else:
        male_percentage = female_percentage = age_15_below_percentage = age_above_15_percentage = total_age_percentage = 0
        total_tcv_percentage = total_hrig_percentage = total_rig_percentage = 0
        dog_bite_percentage = cat_bite_percentage = other_bite_percentage = 0
        immunized_percentage = 0
        unimmunized_percentage = 0

    overall_total = sum(entry.get('data_total', 0) for entry in data)
    overall_total_tcv = sum(entry.get('total_tcv_given', 0) for entry in data)
    overall_total_hrig = sum(entry.get('total_hrig_given', 0) for entry in data)
    overall_total_erig = sum(entry.get('total_erig_given', 0) for entry in data)

    for entry in data:
        entry['percent_total'] = round((entry['data_total'] / overall_total) * 100, 1) if overall_total > 0 else 0
        entry['percent_tcv'] = round((entry.get('total_tcv_given', 0) / overall_total_tcv) * 100, 1) if overall_total_tcv > 0 else 0
        entry['percent_hrig'] = round((entry.get('total_hrig_given', 0) / overall_total_hrig) * 100, 1) if overall_total_hrig > 0 else 0
        entry['percent_erig'] = round((entry.get('total_erig_given', 0) / overall_total_erig) * 100, 1) if overall_total_erig > 0 else 0
    total_percent = sum(entry['percent_total'] for entry in data)
    total_tcv_percent = sum(entry['percent_tcv'] for entry in data)
    total_hrig_percent = sum(entry['percent_hrig'] for entry in data)
    total_erig_percent = sum(entry['percent_erig'] for entry in data)   
    total_human_rabies = sum(entry.get('human_rabies_count', 0) for entry in data)  # Add this line
    municipalities = Municipality.objects.all()
    selected_municipality_id = request.GET.get('municipality_id')
    selected_municipality = municipalities.filter(muni_id=selected_municipality_id).first() if selected_municipality_id else None
    if user.is_superuser:
        table = "ABTC"
    else:
        table = "Barangay"

    if user.first_name or user.last_name:
        signature_name = f"{user.first_name} {user.last_name}".strip()
    else:
        signature_name = user.username

    if user.logo_image:  # Ensure the user has a logo image
        logo_url = request.build_absolute_uri(user.logo_image.url)  
    else:
        logo_url = None  # If no logo is available, set logo_url to None

    karon = date.today().year

    code_to_muni_name = {
        'ALM': 'Almeria',
        'BIL': 'Biliran',
        'CABUC': 'Cabucgayan',
        'CAIB': 'Caibiran',
        'CUL': 'Culaba',
        'KAW': 'Kawayan',
        'MAR': 'Maripipi',
        'NAV': 'Naval',
    }

    if user.is_superuser:
        coordinator = "Provincial Rabies Coordinator"
        pho = "PHO II"
        # Get the first lead doctor
        doct = Doctor.objects.filter(is_superdoctor=True).first()
        doctor = doct.full_name() if doct else "No Doctor Assigned"
        center = "Biliran Province Hospital"
        center_label = f"Animal Bite Treatment Center: {center}"  # Add the label
    else:
        # Map user code to municipality name
        muni_name = code_to_muni_name.get(user.code)

        if muni_name:
            try:
                # Find the municipality with this name
                municipality = Municipality.objects.get(muni_name=muni_name)
                # Filter doctors by this municipality
                doct = Doctor.objects.filter(muni_id=municipality).first()
                doctor = doct.full_name() if doct else "No Doctor Assigned"
                coordinator = f"Municipality of {municipality.muni_name}"
                pho = f"Doctor of {municipality.muni_name}"
                center = muni_name  # Use the user's code to determine the center
                center_label = f"Animal Bite Treatment Center: {center} Animal Bite Treatment Center"  # Add the label
            except Municipality.DoesNotExist:
                # Handle case where the municipality name does not exist
                coordinator = "Unknown Municipality"
                pho = "No Doctor Assigned"
                doctor = "No Doctor Assigned"
                center = "Unknown Center"
                center_label = f"Animal Bite Treatment Center: {center}"  # Add the label
        else:
            # Handle case where the user's code does not match any municipality
            coordinator = "Unknown Municipality"
            pho = "No Doctor Assigned"
            doctor = "No Doctor Assigned"
            center = "Unknown Center"
            center_label = f"Animal Bite Treatment Center: {center} Animal Bite Treatment Center"  # Add the label

        
    context = {
        'doctor': doctor,
        'center': center,
        'center_label': center_label,
        'coordinator': coordinator,
        'pho': pho,
        'karon': karon,    
        'logo_url': logo_url,    
        'signature_name':signature_name,
        'table': table,
        'municipalities': municipalities,
        'selected_municipality': selected_municipality or municipalities.first(),  # Default to the first municipality if none is selected
        'municipality_name': municipality_name,
        'municipal': municipal,
        'selected_quarter': selected_quarter,
        'barangay_list': [d['barangay'] for d in data],
        'data': data,
        'total_male': total_male,
        'total_female': total_female,
        'total_all': total_all,
        'total_age_15_below': total_age_15_below,
        'total_age_above_15': total_age_above_15,
        'male_percentage': round(male_percentage, 1),
        'female_percentage': round(female_percentage, 1),
        'total_sex_percentage':round(total_sex_percentage),
        'age_15_below_percentage': round(age_15_below_percentage, 1),
        'age_above_15_percentage': round(age_above_15_percentage, 1),
        'total_age_percentage': round(total_age_percentage, 1),
        'total_animal_bite_I': total_animal_bite_I,
        'total_animal_bite_II': total_animal_bite_II,
        'total_animal_bite_III': total_animal_bite_III,
        'total_animal_bite_I_percentage': round(total_animal_bite_I_percentage,1),
        'total_animal_bite_II_percentage': round(total_animal_bite_II_percentage,1),
        'total_animal_bite_III_percentage': round(total_animal_bite_III_percentage,1),
        'total_category_percentage': round(total_category_percentage,1),
        'overall_total':overall_total,
        'total_percent':round(total_percent, ),
        'total_tcv_given': total_tcv_given,
        'total_hrig_given': total_hrig_given,
        'total_rig_given': total_erig_given,
        'total_tcv_percentage': round(total_tcv_percentage, 1),
        'total_hrig_percentage': round(total_hrig_percentage, 1),
        'total_rig_percentage': round(total_rig_percentage, 1),
        'total_tcv_percent':round(total_tcv_percent, ),
        'total_erig_percent':round(total_erig_percent, ),
        'total_hrig_percent':round(total_hrig_percent, ),
        'total_dog_bites': total_dog_bites,
        'total_cat_bites': total_cat_bites,
        'total_other_bites': total_other_bites,
        'dog_bite_percentage': round(dog_bite_percentage, 1),
        'cat_bite_percentage': round(cat_bite_percentage, 1),
        'other_bite_percentage': round(other_bite_percentage, 1),
        'total_animal_type_percentage': round(total_animal_type_percentage, 1),
        'total_tcv_given': sum(entry.get('total_tcv_given', 0) for entry in data),
        'total_hrig_given': sum(entry.get('total_hrig_given', 0) for entry in data),
        'total_erig_given': sum(entry.get('total_erig_given', 0) for entry in data),
        'total_immunized': total_immunized,
        'total_unimmunized': total_unimmunized,
        'immunized_percentage': round(immunized_percentage, 1), 
        'unimmunized_percentage': round(unimmunized_percentage, 1),  
        'total_human_rabies':total_human_rabies,

    }
    template = loader.get_template('monitoring/report_pdf4.html')
    html = template.render(context)

    """ # Specify path to wkhtmltopdf executable
    path_to_wkhtmltopdf = r"C:\wkhtmltox\bin\wkhtmltopdf.exe"
    
    # Configure pdfkit with the path
    config = pdfkit.configuration(wkhtmltopdf=path_to_wkhtmltopdf) """

    config = pdfkit.configuration(wkhtmltopdf=WKHTMLTOPDF_PATH)

    # Options for landscape orientation
    options = {
        'orientation': 'Landscape',  # Set to Landscape
        'page-size': 'A4',  # A4 paper size, you can adjust if needed
        'footer-left': center_label,  # Add page numbers
        'footer-right': 'Page [page] of [toPage]',  # Add page numbers
        'footer-font-size': '10',  # Adjust font size for the footer
        'footer-spacing': '5',  # Space between footer and content
        'margin-bottom': '15mm',  # Ensure space for the footer
    }
    
    # Generate PDF from HTML using pdfkit with configuration
    pdf = pdfkit.from_string(html, configuration=config,options=options)

    # Return the PDF as a responsev 
    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment;filename="4th Quarter Report.pdf"'
    return response

def pdf_report_create_annual(request):
    user = request.user
    municipality_map = {
        "MAR": "Maripipi",
        "KAW": "Kawayan",
        "NAV": "Naval",
        "CAIB": "Caibiran",
        "ALM": "Almeria",
        "BIL": "Biliran",
        "CUL": "Culaba",
        "CABUC": "Cabucgayan"
    }
    if user.is_superuser and user.code == "NAV":
        municipality_name = "BPH"
    else:
        municipality_name = municipality_map.get(user.code, "Province")

    if user.is_superuser:
        municipal = "Province of Biliran"
    else:
        munici = municipality_map.get(user.code)
        municipal = (f"Municipality of {munici}")
    # Only set for 2nd quarter (April 1 to June 30)
    selected_quarter = 'annual'
    year = date.today().year
    #Annual
    start_date = date(year, 1, 1)
    end_date = date(year, 12, 31)

    data = []
    total_male = 0
    total_female = 0
    total_all = 0
    total_age_15_below = 0
    total_age_above_15 = 0
    total_sex_percentage = 0
    total_animal_bite_I = 0
    total_animal_bite_II = 0
    total_animal_bite_III = 0
    total_category_percentage = 0
    total_tcv_given = 0
    total_hrig_given = 0
    total_erig_given = 0  
    total_tcv_percentage = 0
    total_hrig_percentage = 0
    total_rig_percentage = 0
    total_dog_bites = 0
    total_cat_bites = 0
    total_other_bites = 0
    total_animal_type_percentage = 0
    total_animal_bite_I_percentage = 0
    total_animal_bite_II_percentage = 0
    total_animal_bite_III_percentage = 0
    total_immunized = 0
    total_unimmunized = 0

    if user.is_superuser:
        abtcs = User.objects.filter(is_superuser=False).distinct()
        for abtc_user in abtcs:
            if not abtc_user.code:
                continue

            male_count = History.objects.filter(
                patient_id__user=abtc_user,
                date_registered__range=(start_date, end_date),
                patient_id__sex='male'
            ).count()
            female_count = History.objects.filter(
                patient_id__user=abtc_user,
                date_registered__range=(start_date, end_date),
                patient_id__sex='female'
            ).count()

            patients = Patient.objects.filter(
                user=abtc_user,  
                histories__date_registered__range=(start_date, end_date)  
            ).distinct()
            age_15_below_count = sum(1 for patient in patients if calculate_age(patient.birthday) <= 15)
            age_above_15_count = sum(1 for patient in patients if calculate_age(patient.birthday) > 15)

            user_animal_bite_I = 0
            user_animal_bite_II = 0
            user_animal_bite_III = 0

            animal_bite_counts = History.objects.filter(
                patient_id__user=abtc_user,
                date_registered__range=(start_date, end_date)
            ).values('category_of_exposure').annotate(count=models.Count('category_of_exposure'))

            for count in animal_bite_counts:
                if count['category_of_exposure'] == 'I':
                    user_animal_bite_I = count['count']
                elif count['category_of_exposure'] == 'II':
                    user_animal_bite_II = count['count']
                elif count['category_of_exposure'] == 'III':
                    user_animal_bite_III = count['count']

            tcv_count = Treatment.objects.filter(
                patient_id__user=abtc_user,
                tcv_given__range=(start_date, end_date)
            ).count()

            hrig_count = Treatment.objects.filter(
                patient_id__user=abtc_user,
                hrig_given__range=(start_date, end_date)
            ).count()

            erig_count = Treatment.objects.filter(
                patient_id__user=abtc_user,
                rig_given__range=(start_date, end_date) 
            ).count()

            animal_type_counts = History.objects.filter(
                patient_id__user=abtc_user,
                date_registered__range=(start_date, end_date)
            ).values('source_of_exposure').annotate(count=models.Count('source_of_exposure'))

            user_dog_bites = 0
            user_cat_bites = 0
            user_other_bites = 0

            for count in animal_type_counts:
                if count['source_of_exposure'] == 'Dog':
                    user_dog_bites = count['count']
                elif count['source_of_exposure'] == 'Cat':
                    user_cat_bites = count['count']
                elif count['source_of_exposure'] == 'Others':
                    user_other_bites = count['count']

            user_immunized_count = History.objects.filter(
                patient_id__user=abtc_user,
                immunization_status='Immunized',
                date_registered__range=(start_date, end_date)
            ).count()
            
            user_unimmunized_count = History.objects.filter(
                patient_id__user=abtc_user,
                immunization_status='Unimmunized',
                date_registered__range=(start_date, end_date)
            ).count()

            user_human_rabies_count = History.objects.filter(
                patient_id__user=abtc_user,
                human_rabies=True,
                date_registered__range=(start_date, end_date)
            ).count()
                    
            total_immunized += user_immunized_count
            total_unimmunized += user_unimmunized_count
            total_dog_bites += user_dog_bites
            total_cat_bites += user_cat_bites
            total_other_bites += user_other_bites
            total_tcv_given += tcv_count
            total_hrig_given += hrig_count
            total_erig_given += erig_count

            total_count = male_count + female_count
            total_male += male_count
            total_female += female_count
            total_all += total_count
            total_age_15_below += age_15_below_count
            total_age_above_15 += age_above_15_count
            total_animal_bite_I += user_animal_bite_I
            total_animal_bite_II += user_animal_bite_II
            total_animal_bite_III += user_animal_bite_III

            # Determine barangay field for this ABTC user
            if abtc_user.code == "NAV":
                barangay_name = "BPH-ABTC"
            else:
                barangay_name = f"{municipality_map.get(abtc_user.code, 'Unknown')}-ABTC"

            data.append({
                'barangay': barangay_name,
                'data_male': male_count,
                'data_female': female_count,
                'data_total': total_count,
                'age_15_below': age_15_below_count,
                'age_above_15': age_above_15_count,
                'age_total': age_15_below_count + age_above_15_count,
                'total_animal_bite_I': user_animal_bite_I,
                'total_animal_bite_II': user_animal_bite_II,
                'total_animal_bite_III': user_animal_bite_III,
                'total_animal': user_animal_bite_I + user_animal_bite_II + user_animal_bite_III,
                'total_tcv_given': tcv_count,
                'total_hrig_given': hrig_count,
                'total_erig_given': erig_count,
                'total_dog_bites': user_dog_bites,
                'total_cat_bites': user_cat_bites,
                'total_other_bites': user_other_bites,
                'total_animal_bites': user_dog_bites + user_cat_bites + user_other_bites,
                'immunized_count': user_immunized_count,
                'unimmunized_count': user_unimmunized_count,  
                'human_rabies_count': user_human_rabies_count,
            })
    else:
        # For non-superuser
        patients = Patient.objects.filter(user=user)
        barangays = Barangay.objects.filter(patients_brgy__in=patients).distinct()

        # Filter barangays that have records in the 2nd quarter
        barangays_with_data = []
        for barangay in barangays:
            # Check if there's any patient in the barangay with a history in the 2nd quarter
            patient_data_exists = History.objects.filter(
                patient_id__brgy_id=barangay,
                patient_id__user=user,
                date_registered__range=(start_date, end_date)
            ).exists()
            
            if patient_data_exists:
                barangays_with_data.append(barangay)

        # Loop through barangays that have data in the 2nd quarter
        for barangay in barangays_with_data:
            male_count = History.objects.filter(
                patient_id__brgy_id=barangay,
                patient_id__user=user,
                date_registered__range=(start_date, end_date),
                patient_id__sex='male'
            ).count()
            female_count = History.objects.filter(
                patient_id__brgy_id=barangay,
                patient_id__user=user,
                date_registered__range=(start_date, end_date),
                patient_id__sex='female'
            ).count()

            patients = Patient.objects.filter(
                brgy_id=barangay,
                user=user,
                histories__date_registered__range=(start_date, end_date)  # Filter by registration date within the second quarter
            ).distinct()
            
            age_15_below_count = sum(1 for patient in patients if calculate_age(patient.birthday) <= 15)
            age_above_15_count = sum(1 for patient in patients if calculate_age(patient.birthday) > 15)

            # Initialize counts for animal bite categories
            barangay_animal_bite_I = 0
            barangay_animal_bite_II = 0
            barangay_animal_bite_III = 0

            animal_bite_counts = History.objects.filter(
                patient_id__brgy_id=barangay,
                patient_id__user=user,
                date_registered__range=(start_date, end_date)
            ).values('category_of_exposure').annotate(count=models.Count('category_of_exposure'))

            tcv_count = Treatment.objects.filter(
                patient_id__brgy_id=barangay,
                patient_id__user=user,
                tcv_given__range=(start_date, end_date)
            ).count()

            hrig_count = Treatment.objects.filter(
                patient_id__brgy_id=barangay,
                patient_id__user=user,
                hrig_given__range=(start_date, end_date)
            ).count()

            erig_count = Treatment.objects.filter(
                patient_id__brgy_id=barangay,
                patient_id__user=user,
                rig_given__range=(start_date, end_date)
            ).count()

            total_tcv_given += tcv_count
            total_hrig_given += hrig_count
            total_erig_given += erig_count

            for count in animal_bite_counts:
                if count['category_of_exposure'] == 'I':
                    barangay_animal_bite_I = count['count']
                elif count['category_of_exposure'] == 'II':
                    barangay_animal_bite_II = count['count']
                elif count['category_of_exposure'] == 'III':
                    barangay_animal_bite_III = count['count']

            animal_type_counts = History.objects.filter(
                patient_id__brgy_id=barangay,
                patient_id__user=user,
                date_registered__range=(start_date, end_date)
            ).values('source_of_exposure').annotate(count=models.Count('source_of_exposure'))

            barangay_dog_bites = 0
            barangay_cat_bites = 0
            barangay_other_bites = 0

            for count in animal_type_counts:
                if count['source_of_exposure'] == 'Dog':
                    barangay_dog_bites = count['count']
                elif count['source_of_exposure'] == 'Cat':
                    barangay_cat_bites = count['count']
                elif count['source_of_exposure'] == 'Others':
                    barangay_other_bites = count['count']

            barangay_immunized_count = History.objects.filter(
                patient_id__brgy_id=barangay,
                patient_id__user=user,
                immunization_status='Immunized',
                date_registered__range=(start_date, end_date)
            ).count()

            barangay_unimmunized_count = History.objects.filter(
                patient_id__brgy_id=barangay,
                patient_id__user=user,
                immunization_status='Unimmunized',
                date_registered__range=(start_date, end_date)
            ).count()

            barangay_human_rabies_count = History.objects.filter(
                patient_id__brgy_id=barangay,
                patient_id__user=user,
                human_rabies=True,
                date_registered__range=(start_date, end_date)
            ).count()

            total_immunized += barangay_immunized_count
            total_unimmunized += barangay_unimmunized_count
            total_dog_bites += barangay_dog_bites
            total_cat_bites += barangay_cat_bites
            total_other_bites += barangay_other_bites
            total_animal_bite_I += barangay_animal_bite_I
            total_animal_bite_II += barangay_animal_bite_II
            total_animal_bite_III += barangay_animal_bite_III

            # Add to total counts
            total_count = male_count + female_count
            total_male += male_count
            total_female += female_count
            total_all += total_count
            total_age_15_below += age_15_below_count
            total_age_above_15 += age_above_15_count

            data.append({
                'barangay': barangay.brgy_name,
                'data_male': male_count,
                'data_female': female_count,
                'data_total': total_count,
                'age_15_below': age_15_below_count,
                'age_above_15': age_above_15_count,
                'age_total': age_15_below_count + age_above_15_count,
                'total_animal_bite_I': barangay_animal_bite_I,
                'total_animal_bite_II': barangay_animal_bite_II,
                'total_animal_bite_III': barangay_animal_bite_III,
                'total_animal': barangay_animal_bite_I + barangay_animal_bite_II + barangay_animal_bite_III,
                'total_tcv_given': tcv_count,
                'total_hrig_given': hrig_count,
                'total_erig_given': erig_count,
                'total_dog_bites': barangay_dog_bites,
                'total_cat_bites': barangay_cat_bites,
                'total_other_bites': barangay_other_bites,
                'total_animal_bites': barangay_dog_bites + barangay_cat_bites + barangay_other_bites,
                'immunized_count': barangay_immunized_count,
                'unimmunized_count': barangay_unimmunized_count,
                'human_rabies_count': barangay_human_rabies_count,
            })
    if total_all > 0:
        male_percentage = (total_male / total_all) * 100
        female_percentage = (total_female / total_all) * 100
        total_sex_percentage = (male_percentage + female_percentage)
        age_15_below_percentage = (total_age_15_below / total_all) * 100
        age_above_15_percentage = (total_age_above_15 / total_all) * 100
        total_age_percentage = (total_age_15_below + total_age_above_15) / total_all * 100
        total_animal_bite_I_percentage = (total_animal_bite_I / total_all ) * 100
        total_animal_bite_II_percentage = (total_animal_bite_II / total_all ) * 100
        total_animal_bite_III_percentage = (total_animal_bite_III / total_all ) * 100
        total_category_percentage = (total_animal_bite_I_percentage + total_animal_bite_II_percentage + total_animal_bite_III_percentage)
        total_tcv_percentage = (total_tcv_given / total_all) * 100
        total_hrig_percentage = (total_hrig_given / total_all) * 100
        total_rig_percentage = (total_erig_given / total_all) * 100  
        dog_bite_percentage = (total_dog_bites / total_all) * 100
        cat_bite_percentage = (total_cat_bites / total_all) * 100
        other_bite_percentage = (total_other_bites / total_all) * 100
        total_animal_type_percentage = dog_bite_percentage + cat_bite_percentage + other_bite_percentage
        if total_immunized + total_unimmunized > 0:
            immunized_percentage = (total_immunized / (total_immunized + total_unimmunized)) * 100
            unimmunized_percentage = (total_unimmunized / (total_immunized + total_unimmunized)) * 100
        else:
            immunized_percentage = 0
            unimmunized_percentage = 0
    else:
        male_percentage = female_percentage = age_15_below_percentage = age_above_15_percentage = total_age_percentage = 0
        total_tcv_percentage = total_hrig_percentage = total_rig_percentage = 0
        dog_bite_percentage = cat_bite_percentage = other_bite_percentage = 0
        immunized_percentage = 0
        unimmunized_percentage = 0

    overall_total = sum(entry.get('data_total', 0) for entry in data)
    overall_total_tcv = sum(entry.get('total_tcv_given', 0) for entry in data)
    overall_total_hrig = sum(entry.get('total_hrig_given', 0) for entry in data)
    overall_total_erig = sum(entry.get('total_erig_given', 0) for entry in data)

    for entry in data:
        entry['percent_total'] = round((entry['data_total'] / overall_total) * 100, 1) if overall_total > 0 else 0
        entry['percent_tcv'] = round((entry.get('total_tcv_given', 0) / overall_total_tcv) * 100, 1) if overall_total_tcv > 0 else 0
        entry['percent_hrig'] = round((entry.get('total_hrig_given', 0) / overall_total_hrig) * 100, 1) if overall_total_hrig > 0 else 0
        entry['percent_erig'] = round((entry.get('total_erig_given', 0) / overall_total_erig) * 100, 1) if overall_total_erig > 0 else 0
    total_percent = sum(entry['percent_total'] for entry in data)
    total_tcv_percent = sum(entry['percent_tcv'] for entry in data)
    total_hrig_percent = sum(entry['percent_hrig'] for entry in data)
    total_erig_percent = sum(entry['percent_erig'] for entry in data)   
    total_human_rabies = sum(entry.get('human_rabies_count', 0) for entry in data)  # Add this line
    municipalities = Municipality.objects.all()
    selected_municipality_id = request.GET.get('municipality_id')
    selected_municipality = municipalities.filter(muni_id=selected_municipality_id).first() if selected_municipality_id else None
    if user.is_superuser:
        table = "ABTC"
    else:
        table = "Barangay"

    if user.first_name or user.last_name:
        signature_name = f"{user.first_name} {user.last_name}".strip()
    else:
        signature_name = user.username

    if user.logo_image:  # Ensure the user has a logo image
        logo_url = request.build_absolute_uri(user.logo_image.url)
    else:
        logo_url = None  # If no logo is available, set logo_url to None
    
    karon = date.today().year
    
    code_to_muni_name = {
        'ALM': 'Almeria',
        'BIL': 'Biliran',
        'CABUC': 'Cabucgayan',
        'CAIB': 'Caibiran',
        'CUL': 'Culaba',
        'KAW': 'Kawayan',
        'MAR': 'Maripipi',
        'NAV': 'Naval',
    }

    if user.is_superuser:
        coordinator = "Provincial Rabies Coordinator"
        pho = "PHO II"
        # Get the first lead doctor
        doct = Doctor.objects.filter(is_superdoctor=True).first()
        doctor = doct.full_name() if doct else "No Doctor Assigned"
        center = "Biliran Province Hospital"
        center_label = f"Animal Bite Treatment Center: {center}"  # Add the label
    else:
        # Map user code to municipality name
        muni_name = code_to_muni_name.get(user.code)

        if muni_name:
            try:
                # Find the municipality with this name
                municipality = Municipality.objects.get(muni_name=muni_name)
                # Filter doctors by this municipality
                doct = Doctor.objects.filter(muni_id=municipality).first()
                doctor = doct.full_name() if doct else "No Doctor Assigned"
                coordinator = f"Municipality of {municipality.muni_name}"
                pho = f"Doctor of {municipality.muni_name}"
                center = muni_name  # Use the user's code to determine the center
                center_label = f"Animal Bite Treatment Center: {center} Animal Bite Treatment Center"  # Add the label
            except Municipality.DoesNotExist:
                # Handle case where the municipality name does not exist
                coordinator = "Unknown Municipality"
                pho = "No Doctor Assigned"
                doctor = "No Doctor Assigned"
                center = "Unknown Center"
                center_label = f"Animal Bite Treatment Center: {center}"  # Add the label
        else:
            # Handle case where the user's code does not match any municipality
            coordinator = "Unknown Municipality"
            pho = "No Doctor Assigned"
            doctor = "No Doctor Assigned"
            center = "Unknown Center"
            center_label = f"Animal Bite Treatment Center: {center} Animal Bite Treatment Center"  # Add the label


    context = {
        'doctor': doctor,
        'center': center,
        'center_label': center_label,
        'coordinator': coordinator,
        'pho': pho,
        'karon': karon,    
        'logo_url': logo_url,    
        'signature_name':signature_name,
        'table': table,
        'municipalities': municipalities,
        'selected_municipality': selected_municipality or municipalities.first(),  # Default to the first municipality if none is selected
        'municipality_name': municipality_name,
        'municipal': municipal,
        'selected_quarter': selected_quarter,
        'barangay_list': [d['barangay'] for d in data],
        'data': data,
        'total_male': total_male,
        'total_female': total_female,
        'total_all': total_all,
        'total_age_15_below': total_age_15_below,
        'total_age_above_15': total_age_above_15,
        'male_percentage': round(male_percentage, 1),
        'female_percentage': round(female_percentage, 1),
        'total_sex_percentage':round(total_sex_percentage),
        'age_15_below_percentage': round(age_15_below_percentage, 1),
        'age_above_15_percentage': round(age_above_15_percentage, 1),
        'total_age_percentage': round(total_age_percentage, 1),
        'total_animal_bite_I': total_animal_bite_I,
        'total_animal_bite_II': total_animal_bite_II,
        'total_animal_bite_III': total_animal_bite_III,
        'total_animal_bite_I_percentage': round(total_animal_bite_I_percentage,1),
        'total_animal_bite_II_percentage': round(total_animal_bite_II_percentage,1),
        'total_animal_bite_III_percentage': round(total_animal_bite_III_percentage,1),
        'total_category_percentage': round(total_category_percentage,1),
        'overall_total':overall_total,
        'total_percent':round(total_percent, ),
        'total_tcv_given': total_tcv_given,
        'total_hrig_given': total_hrig_given,
        'total_rig_given': total_erig_given,
        'total_tcv_percentage': round(total_tcv_percentage, 1),
        'total_hrig_percentage': round(total_hrig_percentage, 1),
        'total_rig_percentage': round(total_rig_percentage, 1),
        'total_tcv_percent':round(total_tcv_percent, ),
        'total_erig_percent':round(total_erig_percent, ),
        'total_hrig_percent':round(total_hrig_percent, ),
        'total_dog_bites': total_dog_bites,
        'total_cat_bites': total_cat_bites,
        'total_other_bites': total_other_bites,
        'dog_bite_percentage': round(dog_bite_percentage, 1),
        'cat_bite_percentage': round(cat_bite_percentage, 1),
        'other_bite_percentage': round(other_bite_percentage, 1),
        'total_animal_type_percentage': round(total_animal_type_percentage, 1),
        'total_tcv_given': sum(entry.get('total_tcv_given', 0) for entry in data),
        'total_hrig_given': sum(entry.get('total_hrig_given', 0) for entry in data),
        'total_erig_given': sum(entry.get('total_erig_given', 0) for entry in data),
        'total_immunized': total_immunized,
        'total_unimmunized': total_unimmunized,
        'immunized_percentage': round(immunized_percentage, 1), 
        'unimmunized_percentage': round(unimmunized_percentage, 1),  
        'total_human_rabies':total_human_rabies,

    }
    template = loader.get_template('monitoring/report_pdf_annual.html')
    html = template.render(context)

    """ # Specify path to wkhtmltopdf executable
    path_to_wkhtmltopdf = r"C:\wkhtmltox\bin\wkhtmltopdf.exe"
    
    # Configure pdfkit with the path
    config = pdfkit.configuration(wkhtmltopdf=path_to_wkhtmltopdf) """

    config = pdfkit.configuration(wkhtmltopdf=WKHTMLTOPDF_PATH)

    # Options for landscape orientation
    options = {
        'orientation': 'Landscape',  # Set to Landscape
        'page-size': 'A4',  # A4 paper size, you can adjust if needed
        'footer-left': center_label,  # Add page numbers
        'footer-right': 'Page [page] of [toPage]',  # Add page numbers
        'footer-font-size': '10',  # Adjust font size for the footer
        'footer-spacing': '5',  # Space between footer and content
        'margin-bottom': '15mm',  # Ensure space for the footer
    }
    
    # Generate PDF from HTML using pdfkit with configuration
    pdf = pdfkit.from_string(html, configuration=config,options=options)

    # Return the PDF as a responsev 
    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment;filename="Annual Report.pdf"'

    return response



@login_required
def tables(request):
    user = request.user

    # Fetch all municipalities
    all_municipalities = Municipality.objects.all()

    # Filter municipalities based on the user
    if user.is_superuser:
        municipalities = Municipality.objects.filter(patients_muni__histories__isnull=False).distinct()
    else:
        municipalities = Municipality.objects.filter(
            patients_muni__user=user,
            patients_muni__histories__isnull=False
        ).distinct()

    # Get the selected municipality and barangay search from the POST parameters
    if request.method == 'POST':
        selected_municipality = request.POST.get('municipality', None)
        barangay_search = request.POST.get('barangay', '').strip()  # Get barangay search input and remove extra spaces
    else:
        selected_municipality = "ALL"
        barangay_search = ""  # Default to an empty search

    # Initialize variables for summary
    total_barangays = 0
    total_patients = 0
    barangay_summary = []
    municipality_summary = []

    if selected_municipality == "ALL" or selected_municipality is None:
        # If "ALL" is selected, show the municipality-level summary
        if user.is_superuser:
            municipality_summary = (
                History.objects.values('muni_id__muni_name')  # Group by municipality
                .annotate(total_barangays=Count('brgy_id', distinct=True),  # Count distinct barangays
                          total_cases=Count('history_id'))  # Count total cases (patients)
                .order_by('muni_id__muni_name')
            )
        else:
            municipality_summary = (
                History.objects.filter(patient_id__user=user)  # Only histories for this user
                .values('muni_id__muni_name')
                .annotate(total_barangays=Count('brgy_id', distinct=True),  # Count distinct barangays
                          total_cases=Count('history_id'))  # Count total cases (patients)
                .order_by('muni_id__muni_name')
            )

        # Calculate the total number of barangays and patients across all municipalities
        total_barangays = sum(record['total_barangays'] for record in municipality_summary)
        total_patients = sum(record['total_cases'] for record in municipality_summary)

    else:
        # Get all barangays for the selected municipality
        barangays_in_municipality = Barangay.objects.filter(muni_id__muni_name=selected_municipality)
        
        # Filter barangays based on the search input (partial match)
        if barangay_search:
            barangays_in_municipality = barangays_in_municipality.filter(brgy_name__icontains=barangay_search)

        # Get barangay-level case summary
        if user.is_superuser:
            cases_in_barangays = History.objects.filter(muni_id__muni_name=selected_municipality) \
                .values('brgy_id__brgy_name') \
                .annotate(total_patients=Count('history_id')) \
                .order_by('brgy_id__brgy_name')
        else:
            cases_in_barangays = History.objects.filter(
                patient_id__user=user,
                muni_id__muni_name=selected_municipality
            ).values('brgy_id__brgy_name') \
                .annotate(total_patients=Count('history_id')) \
                .order_by('brgy_id__brgy_name')

        # Create a dictionary for easy lookup
        cases_dict = {record['brgy_id__brgy_name']: record['total_patients'] for record in cases_in_barangays}

        # Populate barangay summary: if no cases, set total patients to 0
        barangay_summary = [
            {
                'brgy_name': barangay.brgy_name,
                'total_patients': cases_dict.get(barangay.brgy_name, 0)  # Get patients, default to 0 if not found
            }
            for barangay in barangays_in_municipality
        ]

        # Calculate totals
        total_barangays = barangays_in_municipality.count()
        total_patients = sum(record['total_patients'] for record in barangay_summary)

        # Summary for the selected municipality
        municipality_summary = [
            {
                'muni_id__muni_name': selected_municipality,
                'total_barangays': total_barangays,
                'total_cases': total_patients,
            }
        ]

    # Context to pass to the template
    context = {
        'barangay_summary': barangay_summary,
        'municipality_summary': municipality_summary,
        'municipalities': municipalities,
        'selected_municipality': selected_municipality,
        'barangay_search': barangay_search,  # Pass the search input to the template
        'total_barangays': total_barangays,
        'total_patients': total_patients,
        'all_municipalities': all_municipalities,
    }

    return render(request, 'monitoring/table.html', context)

""" @login_required
def tables(request):
    user = request.user

    # Fetch all municipalities
    all_municipalities = Municipality.objects.all()

    # Filter municipalities based on the user
    if user.is_superuser:
        municipalities = Municipality.objects.filter(patients_muni__histories__isnull=False).distinct()
    else:
        municipalities = Municipality.objects.filter(
            patients_muni__user=user,
            patients_muni__histories__isnull=False
        ).distinct()

    # Get the selected municipality and barangay search from the POST parameters
    if request.method == 'POST':
        selected_municipality = request.POST.get('municipality', None)
        barangay_search = request.POST.get('barangay', '').strip()  # Get barangay search input and remove extra spaces
    else:
        selected_municipality = "ALL"
        barangay_search = ""  # Default to an empty search

    # Initialize variables for summary
    total_barangays = 0
    total_patients = 0
    barangay_summary = []
    municipality_summary = []

    # If there's a search term for barangay, show barangays matching the search term
    if barangay_search:
        # Get all barangays matching the search term, including abbreviations
        barangays = Barangay.objects.filter(
            Q(brgy_name__icontains=barangay_search) | 
            Q(brgy_name__icontains='Pob') & Q(brgy_name__icontains=barangay_search) |
            Q(brgy_name__icontains='Poblacion') & Q(brgy_name__icontains=barangay_search)
        )

        # Create a summary with the total patients for each barangay across all municipalities
        cases_in_barangays = History.objects.filter(brgy_id__in=barangays) \
            .values('brgy_id__brgy_name', 'muni_id__muni_name') \
            .annotate(total_patients=Count('history_id')) \
            .order_by('muni_id__muni_name', 'brgy_id__brgy_name')

        # Create a dictionary for easy lookup
        cases_dict = {record['brgy_id__brgy_name']: record['total_patients'] for record in cases_in_barangays}

        # Populate barangay summary with corresponding municipalities
        barangay_summary = [
            {
                'muni_name': record['muni_id__muni_name'],
                'brgy_name': record['brgy_id__brgy_name'],
                'total_patients': cases_dict.get(record['brgy_id__brgy_name'], 0)  # Get patients, default to 0 if not found
            }
            for record in cases_in_barangays
        ]

        # Calculate totals
        total_barangays = len(barangay_summary)
        total_patients = sum(record['total_patients'] for record in barangay_summary)

    else:
        # Existing logic for when no search term is provided
        if selected_municipality == "ALL" or selected_municipality is None:
            # If "ALL" is selected, show the municipality-level summary
            if user.is_superuser:
                municipality_summary = (
                    History.objects.values('muni_id__muni_name')  # Group by municipality
                    .annotate(total_barangays=Count('brgy_id', distinct=True),  # Count distinct barangays
                              total_cases=Count('history_id'))  # Count total cases (patients)
                    .order_by('muni_id__muni_name')
                )
            else:
                municipality_summary = (
                    History.objects.filter(patient_id__user=user)  # Only histories for this user
                    .values('muni_id__muni_name')
                    .annotate(total_barangays=Count('brgy_id', distinct=True),  # Count distinct barangays
                              total_cases=Count('history_id'))  # Count total cases (patients)
                    .order_by('muni_id__muni_name')
                )

            # Calculate the total number of barangays and patients across all municipalities
            total_barangays = sum(record['total_barangays'] for record in municipality_summary)
            total_patients = sum(record['total_cases'] for record in municipality_summary)

        else:
            # Get all barangays for the selected municipality
            barangays_in_municipality = Barangay.objects.filter(muni_id__muni_name=selected_municipality)

            # Get barangay-level case summary
            if user.is_superuser:
                cases_in_barangays = History.objects.filter(muni_id__muni_name=selected_municipality) \
                    .values('brgy_id__brgy_name') \
                    .annotate(total_patients=Count('history_id')) \
                    .order_by('brgy_id__brgy_name')
            else:
                cases_in_barangays = History.objects.filter(
                    patient_id__user=user,
                    muni_id__muni_name=selected_municipality
                ).values('brgy_id__brgy_name') \
                    .annotate(total_patients=Count('history_id')) \
                    .order_by('brgy_id__brgy_name')

            # Create a dictionary for easy lookup
            cases_dict = {record['brgy_id__brgy_name']: record['total_patients'] for record in cases_in_barangays}

            # Populate barangay summary: if no cases, set total patients to 0
            barangay_summary = [
                {
                    'brgy_name': barangay.brgy_name,
                    'total_patients': cases_dict.get(barangay.brgy_name, 0)  # Get patients, default to 0 if not found
                }
                for barangay in barangays_in_municipality
            ]

            # Calculate totals
            total_barangays = barangays_in_municipality.count()
            total_patients = sum(record['total_patients'] for record in barangay_summary)

            # Summary for the selected municipality
            municipality_summary = [
                {
                    'muni_id__muni_name': selected_municipality,
                    'total_barangays': total_barangays,
                    'total_cases': total_patients,
                }
            ]

    # Context to pass to the template
    context = {
        'barangay_summary': barangay_summary,
        'municipality_summary': municipality_summary,
        'municipalities': municipalities,
        'selected_municipality': selected_municipality,
        'barangay_search': barangay_search,  # Pass the search input to the template
        'total_barangays': total_barangays,
        'total_patients': total_patients,
        'all_municipalities': all_municipalities,
    }

    return render(request, 'monitoring/table.html', context)
 """


@login_required
def download(request):
    user = request.user

    municipality_map = {
        "MAR": "Maripipi",
        "KAW": "Kawayan",
        "NAV": "Naval",
        "CAIB": "Caibiran",
        "ALM": "Almeria",
        "BIL": "Biliran",
        "CUL": "Culaba",
        "CABUC": "Cabucgayan"
    }
    # Get selected filters from request
    selected_municipality = request.GET.get('municipality')
    selected_barangay = request.GET.get('barangay')
    selected_user = request.GET.get('searchUsername') 
    start_month = request.GET.get('startMonth')
    end_month = request.GET.get('endMonth')
    search_name = request.GET.get('searchName')
    
    # Fetch the histories with related patient, municipality, and barangay data
    histories = History.objects.select_related('patient_id', 'muni_id', 'brgy_id').order_by('-registration_no')

    if not user.is_superuser:
        # Filter histories for the current user if not a superuser
        histories = histories.filter(patient_id__user=user)
    
    patients = Patient.objects.all()

    # Apply filters based on selected municipality and barangay
    if selected_municipality:
        histories = histories.filter(muni_id=selected_municipality)
    if selected_barangay:
        histories = histories.filter(brgy_id=selected_barangay)

    # Apply filter based on username search only if user is a superuser
    if user.is_superuser and selected_user:
        histories = histories.filter(patient_id__user__username=selected_user)
    
    # Fetch unique users from the Patient model for the dropdown if user is superuser
    if user.is_superuser:
        patient_users = Patient.objects.select_related('user').values_list('user', flat=True).distinct()
        users = User.objects.filter(id__in=patient_users)
    else:
        users = []

    # Apply filter based on selected start and end months
    if start_month and end_month:
        try:
            current_year = datetime.now().year
            start_date = datetime.strptime(f"{start_month} {current_year}", "%B %Y").replace(day=1)
            end_date = datetime.strptime(f"{end_month} {current_year}", "%B %Y").replace(day=1) + relativedelta(months=1) - relativedelta(days=1)
            histories = histories.filter(date_registered__gte=start_date, date_registered__lte=end_date)
        except Exception as e:
            print(f"Error parsing date: {e}")  # Log the error and continue

     # Apply filter based on name search
    if search_name:
        histories = histories.filter(Q(patient_id__first_name__icontains=search_name) | Q(patient_id__last_name__icontains=search_name))

    months = month_name[1:]
    
    # Calculate age and attach to each history instance
    for history in histories:
        history.treatment = Treatment.objects.filter(patient_id=history.patient_id).first()
        history.age = calculate_age(history.patient_id.birthday)

    # Count the number of male and female patients
    male = histories.filter(patient_id__sex__iexact='Male').count()
    female = histories.filter(patient_id__sex__iexact='Female').count()

    # Calculate the number of age
    age_15_or_less_count = 0
    age_above_15_count = 0

    for history in histories:
        age = calculate_age(history.patient_id.birthday)
        if age <= 15:
            age_15_or_less_count += 1
        else:
            age_above_15_count += 1

    # Calculate counts for different animal bites   
    source_of_exposure_counter = Counter(histories.values_list('source_of_exposure', flat=True))
    dog_count = source_of_exposure_counter.get('Dog', 0)
    cat_count = source_of_exposure_counter.get('Cat', 0)
    other_animal_count = source_of_exposure_counter.get('Others', 0)

    
    # Calculate total number of distinct patients
    total_patients = histories.values('patient_id').distinct().count()
    

    paginator = Paginator(histories,10)  
    page_number = request.GET.get('page',1)
    try:
        histories = paginator.get_page(page_number)
    except PageNotAnInteger:
        histories = paginator.get_page(1)
    except EmptyPage:
        histories = paginator.get_page(paginator.num_pages)

    # Collect current query parameters
    query_params = request.GET.dict()
    if 'page' in query_params:
        del query_params['page']  # Remove the page parameter
    query_string = urlencode(query_params)

    municipalities = Municipality.objects.all()
    barangays = Barangay.objects.all()

    if user.first_name or user.last_name:
        signature_name = f"{user.first_name} {user.last_name}".strip()
    else:
        signature_name = user.username

    code_to_muni_name = {
        'ALM': 'Almeria',
        'BIL': 'Biliran',
        'CABUC': 'Cabucgayan',
        'CAIB': 'Caibiran',
        'CUL': 'Culaba',
        'KAW': 'Kawayan',
        'MAR': 'Maripipi',
        'NAV': 'Naval',
    }
    
    if user.is_superuser:
        coordinator = "Provincial Rabies Coordinator"
        pho = "PHO II"
        # Get the first lead doctor
        doct = Doctor.objects.filter(is_superdoctor=True).first()
        doctor = doct.full_name() if doct else "No Doctor Assigned"
    else:
        # Map user code to municipality name
        muni_name = code_to_muni_name.get(user.code)

        if muni_name:
            try:
                # Find the municipality with this name
                municipality = Municipality.objects.get(muni_name=muni_name)
                # Filter doctors by this municipality
                doct = Doctor.objects.filter(muni_id=municipality).first()
                doctor = doct.full_name() if doct else "No Doctor Assigned"
                coordinator = f"Municipality of {municipality.muni_name}"
                pho = f"Doctor of {municipality.muni_name}"
            except Municipality.DoesNotExist:
                # Handle case where the municipality name does not exist
                coordinator = "Unknown Municipality"
                pho = "No Doctor Assigned"
                doctor = "No Doctor Assigned"
        else:
            # Handle case where the user's code does not match any municipality
            coordinator = "Unknown Municipality"
            pho = "No Doctor Assigned"
            doctor = "No Doctor Assigned"

        

    context = {
        'doctor':doctor,
        'signature_name':signature_name,
        'histories': histories,
        'municipalities': municipalities,
        'coordinator': coordinator,
        'pho': pho,
        'barangays': barangays,
        'selected_municipality': selected_municipality,
        'selected_barangay': selected_barangay,
        'selected_user':selected_user,
        'start_month': start_month,
        'end_month': end_month,
        'search_name': search_name,
        'months': months,
        'male' : male,
        'female' : female,
        'dog_count' : dog_count,
        'cat_count' : cat_count,
        'other_animal_count' : other_animal_count,
        'age_15_or_less_count' : age_15_or_less_count,
        'age_above_15_count' : age_above_15_count,
        'query_string' : query_string,
        'users':users,
        'total_patients': total_patients,
    }

    return render(request, 'monitoring/download.html',context)


@login_required
def pdf_masterlist_create(request):
    
    user = request.user

    municipality_map = {
        "MAR": "Maripipi",
        "KAW": "Kawayan",
        "NAV": "Naval",
        "CAIB": "Caibiran",
        "ALM":"Almeria",
        "BIL":"Biliran",
        "CUL":"Culaba",
        "CABUC":"Cabucgayan"
    }

    # Get filters from the request
    selected_municipality = request.GET.get('municipality')
    selected_barangay = request.GET.get('barangay')
    selected_user = request.GET.get('searchUsername') 
    start_month = request.GET.get('startMonth', '')
    end_month = request.GET.get('endMonth', '')
    search_name = request.GET.get('searchName')
    
    # Base queryset with filters applied
    histories = History.objects.select_related('patient_id', 'muni_id', 'brgy_id').order_by('-registration_no')
    if not user.is_superuser:
        histories = histories.filter(patient_id__user=user)
    
    # Apply filters
    histories = histories.filter(
        Q(muni_id=selected_municipality) if selected_municipality else Q(),
        Q(brgy_id=selected_barangay) if selected_barangay else Q(),
        Q(patient_id__user__username=selected_user) if user.is_superuser and selected_user else Q(),
        Q(patient_id__first_name__icontains=search_name) | Q(patient_id__last_name__icontains=search_name) if search_name else Q()
    )

    # Date filter
    try:
        if start_month and end_month:
            current_year = datetime.now().year
            start_date = datetime.strptime(f"{start_month} {current_year}", "%B %Y").replace(day=1)
            end_date = datetime.strptime(f"{end_month} {current_year}", "%B %Y").replace(day=1) + relativedelta(months=1) - relativedelta(days=1)
            histories = histories.filter(date_registered__range=(start_date, end_date))
    except ValueError as e:
        messages.error(request, "Invalid date range selected.")
    
    # Calculate statistics
    male = histories.filter(patient_id__sex__iexact='Male').count()
    female = histories.filter(patient_id__sex__iexact='Female').count()

    age_15_or_less_count = age_above_15_count = 0
    for history in histories:
        history.age = calculate_age(history.patient_id.birthday)
        history.treatment = Treatment.objects.filter(patient_id=history.patient_id).first()
        if history.age <= 15:
            age_15_or_less_count += 1
        else:
            age_above_15_count += 1

    animal_counts = Counter(histories.values_list('source_of_exposure', flat=True))
    dog_count, cat_count, other_animal_count = animal_counts.get('Dog', 0), animal_counts.get('Cat', 0), animal_counts.get('Others', 0)

    # Calculate total number of distinct patients
    total_patients = histories.values('patient_id').distinct().count()

    code_to_muni_name = {
        'ALM': 'Almeria',
        'BIL': 'Biliran',
        'CABUC': 'Cabucgayan',
        'CAIB': 'Caibiran',
        'CUL': 'Culaba',
        'KAW': 'Kawayan',
        'MAR': 'Maripipi',
        'NAV': 'Naval',
    }

    if user.is_superuser:
        coordinator = "Provincial Rabies Coordinator"
        pho = "PHO II"
        # Get the first lead doctor
        doct = Doctor.objects.filter(is_superdoctor=True).first()
        doctor = doct.full_name() if doct else "No Doctor Assigned"
        center = "Biliran Province Hospital"
        center_label = f"Animal Bite Treatment Center: {center}"  # Add the label
    else:
        # Map user code to municipality name
        muni_name = code_to_muni_name.get(user.code)

        if muni_name:
            try:
                # Find the municipality with this name
                municipality = Municipality.objects.get(muni_name=muni_name)
                # Filter doctors by this municipality
                doct = Doctor.objects.filter(muni_id=municipality).first()
                doctor = doct.full_name() if doct else "No Doctor Assigned"
                coordinator = f"Municipality of {municipality.muni_name}"
                pho = f"Doctor of {municipality.muni_name}"
                center = muni_name  # Use the user's code to determine the center
                center_label = f"Animal Bite Treatment Center: {center} Animal Bite Treatment Center"  # Add the label
            except Municipality.DoesNotExist:
                # Handle case where the municipality name does not exist
                coordinator = "Unknown Municipality"
                pho = "No Doctor Assigned"
                doctor = "No Doctor Assigned"
                center = "Unknown Center"
                center_label = f"Animal Bite Treatment Center: {center}"  # Add the label
        else:
            # Handle case where the user's code does not match any municipality
            coordinator = "Unknown Municipality"
            pho = "No Doctor Assigned"
            doctor = "No Doctor Assigned"
            center = "Unknown Center"
            center_label = f"Animal Bite Treatment Center: {center} Animal Bite Treatment Center"  # Add the label
        
    
    
    karon = date.today().year
    # Set context for the PDF template
    context = {
        
        'signature_name': f"{user.first_name} {user.last_name}".strip() or user.username,
        'center_label': center_label,
        'center': center,
        'doctor': doctor,
        'karon': karon,
        'coordinator': coordinator,
        'pho': pho,
        'histories': histories,
        'total_patients': total_patients,
        'male': male,
        'female': female,
        'dog_count': dog_count,
        'cat_count': cat_count,
        'other_animal_count': other_animal_count,
        'age_15_or_less_count': age_15_or_less_count,
        'age_above_15_count': age_above_15_count,
        'municipalities': Municipality.objects.all(),
        'barangays': Barangay.objects.all(),
        'start_month': start_month,
        'end_month': end_month,
    }
    template = loader.get_template('monitoring/download_pdf.html')
    html = template.render(context)

    """ # Specify path to wkhtmltopdf executable
    path_to_wkhtmltopdf = r"C:\wkhtmltox\bin\wkhtmltopdf.exe"
    
    # Configure pdfkit with the path
    config = pdfkit.configuration(wkhtmltopdf=path_to_wkhtmltopdf) """

    config = pdfkit.configuration(wkhtmltopdf=WKHTMLTOPDF_PATH)

    options = {
        'orientation': 'Landscape',  # Set to Landscape
        'page-size': 'A4',  # A4 paper size, you can adjust if neededs
        'footer-left': center_label,
        'footer-right': 'Page [page] of [toPage]',  # Add page numbers
        'footer-font-size': '10',  # Adjust font size for the footer
        'footer-spacing': '5',  # Space between footer and content
        'margin-bottom': '15mm',  # Ensure space for the footer
    }
    
    # Generate PDF from HTML using pdfkit with configuration
    pdf = pdfkit.from_string(html, configuration=config,options=options)

    # Return the PDF as a responsev 
    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment;filename="Master List.pdf"'

    return response


@login_required
def cohort(request):
    user = request.user
    municipality_map = {
        "MAR": "Maripipi",
        "KAW": "Kawayan",
        "NAV": "Naval",
        "CAIB": "Caibiran",
        "ALM": "Almeria",
        "BIL": "Biliran",
        "CUL": "Culaba",
        "CABUC": "Cabucgayan"
    }
    if user.is_superuser and user.code == "NAV":
        municipality_name = "BPH"
    else:
        municipality_name = municipality_map.get(user.code, "Province")

    selected_quarter = request.GET.get('quarter', '1')  # Default to '1' if no quarter is selected
    year = date.today().year

    quarter_ranges = {
        '1': (date(year, 1, 1), date(year, 3, 31)),
        '2': (date(year, 4, 1), date(year, 6, 30)),
        '3': (date(year, 7, 1), date(year, 9, 30)),
        '4': (date(year, 10, 1), date(year, 12, 31)),
    }

    # Generate the title and date range based on the selected quarter
    if selected_quarter == 'annual':
        start_date, end_date = date(year, 1, 1), date(year, 12, 31)
        quarter_select = "Annual"
        report_title = f"{quarter_select} Report {year}"  # Annual report title
    else:
        start_date, end_date = quarter_ranges[selected_quarter]
        if selected_quarter == '1':
            quarter_select = '1st'
        elif selected_quarter == '2':
            quarter_select = '2nd'
        elif selected_quarter == '3':
            quarter_select = '3rd'
        elif selected_quarter == '4':
            quarter_select = '4th'
        report_title = f"{quarter_select} Quarter Report {year}"

    # Retrieve data based on the user type and date range
    if user.is_superuser:
        municipalities = Municipality.objects.all()
        barangays = Barangay.objects.all()
        patients = Patient.objects.all()
        histories = History.objects.filter(date_registered__range=(start_date, end_date))
        treatments = Treatment.objects.filter(
            patient_id__histories__date_registered__range=(start_date, end_date)
        ).distinct()
    else:
        patients = Patient.objects.filter(user=user)
        histories = History.objects.filter(patient_id__in=patients, date_registered__range=(start_date, end_date))
        treatments = Treatment.objects.filter(
            patient_id__histories__date_registered__range=(start_date, end_date)
        ).distinct()
        municipalities = Municipality.objects.filter(muni_id__in=patients.values_list('muni_id', flat=True))
        barangays = Barangay.objects.filter(brgy_id__in=patients.values_list('brgy_id', flat=True))

    # Calculate counts for each category of exposure
    category_i_count = histories.filter(category_of_exposure='I').count()
    category_ii_count = histories.filter(category_of_exposure='II').count()
    category_iii_count = histories.filter(category_of_exposure='III').count()
    total_count = category_i_count + category_ii_count + category_iii_count

    # RIG counts by category
    category_ii_with_rig = histories.filter(
        category_of_exposure='II',
        patient_id__treatments_patient__rig_given__isnull=False,
        date_registered__range=(start_date, end_date)
    ).distinct().count()

    category_iii_with_rig = histories.filter(
        category_of_exposure='III',
        patient_id__treatments_patient__rig_given__isnull=False,
        date_registered__range=(start_date, end_date)
    ).distinct().count()

    total_count_rig = category_ii_with_rig + category_iii_with_rig

    # Outcome counters for each category and date range
    category_ii_complete = category_ii_incomplete = category_ii_none = category_ii_died = 0
    category_iii_complete = category_iii_incomplete = category_iii_none = category_iii_died = 0

    # Process Category II outcomes
    for history in histories.filter(category_of_exposure='II'):
        treatments = Treatment.objects.filter(
            patient_id=history.patient_id,
            patient_id__histories__date_registered__range=(start_date, end_date)
        ).distinct()

        # Check if the patient has a "Died" remark in the treatment records
        if treatments.filter(remarks='Died').exists():
            category_ii_died += 1
            continue

        if treatments.exists():
            # Check if all of day0, day3, and day7 are marked True (Complete)
            if treatments.filter(day0_arrived=True, day3_arrived=True, day7_arrived=True).exists():
                category_ii_complete += 1
            # Check if any one of day0, day3, or day7 is False (Incomplete)
            elif treatments.filter(day0_arrived=False).exists() or \
                treatments.filter(day3_arrived=False).exists() or \
                treatments.filter(day7_arrived=False).exists():
                category_ii_incomplete += 1
            # If all day0, day3, and day7 are False (None)
            elif treatments.filter(day0_arrived=False, day3_arrived=False, day7_arrived=False).exists():
                category_ii_none += 1
        else:
            category_ii_none += 1

    # Process Category III outcomes
    for history in histories.filter(category_of_exposure='III'):
        treatments = Treatment.objects.filter(
            patient_id=history.patient_id,
            patient_id__histories__date_registered__range=(start_date, end_date)
        ).distinct()

        # Check if the patient has a "Died" remark in the treatment records
        if treatments.filter(remarks='Died').exists():
            category_iii_died += 1
            continue

        if treatments.exists():
            # Check if all of day0, day3, and day7 are marked True (Complete)
            if treatments.filter(day0_arrived=True, day3_arrived=True, day7_arrived=True).exists():
                category_iii_complete += 1
            # Check if any one of day0, day3, or day7 is False (Incomplete)
            elif treatments.filter(day0_arrived=False).exists() or \
                treatments.filter(day3_arrived=False).exists() or \
                treatments.filter(day7_arrived=False).exists():
                category_iii_incomplete += 1
            # If all day0, day3, and day7 are False (None)
            elif treatments.filter(day0_arrived=False, day3_arrived=False, day7_arrived=False).exists():
                category_iii_none += 1
        else:
            category_iii_none += 1

    
    total_complete = category_ii_complete + category_iii_complete
    total_incomplete = category_ii_incomplete + category_iii_incomplete
    total_none = category_ii_none + category_iii_none
    total_died = category_ii_died + category_iii_died

    signature_name = f"{user.first_name} {user.last_name}".strip() if user.first_name or user.last_name else user.username

    if user.logo_image:  # Ensure the user has a logo image
        logo_url = request.build_absolute_uri(user.logo_image.url)
    else:
        logo_url = None  # If no logo is available, set logo_url to None

    karon = date.today().year
        
    code_to_muni_name = {
        'ALM': 'Almeria',
        'BIL': 'Biliran',
        'CABUC': 'Cabucgayan',
        'CAIB': 'Caibiran',
        'CUL': 'Culaba',
        'KAW': 'Kawayan',
        'MAR': 'Maripipi',
        'NAV': 'Naval',
    }
    
    if user.is_superuser:
        coordinator = "Provincial Rabies Coordinator"
        pho = "PHO II"
        # Get the first lead doctor
        doct = Doctor.objects.filter(is_superdoctor=True).first()
        doctor = doct.full_name() if doct else "No Doctor Assigned"
    else:
        # Map user code to municipality name
        muni_name = code_to_muni_name.get(user.code)

        if muni_name:
            try:
                # Find the municipality with this name
                municipality = Municipality.objects.get(muni_name=muni_name)
                # Filter doctors by this municipality
                doct = Doctor.objects.filter(muni_id=municipality).first()
                doctor = doct.full_name() if doct else "No Doctor Assigned"
                coordinator = f"Municipality of {municipality.muni_name}"
                pho = f"Doctor of {municipality.muni_name}"
            except Municipality.DoesNotExist:
                # Handle case where the municipality name does not exist
                coordinator = "Unknown Municipality"
                pho = "No Doctor Assigned"
                doctor = "No Doctor Assigned"
        else:
            # Handle case where the user's code does not match any municipality
            coordinator = "Unknown Municipality"
            pho = "No Doctor Assigned"
            doctor = "No Doctor Assigned"

    context = {
        'coordinator': coordinator,
        'pho': pho,
        'doctor': doctor,
        'report_title': report_title,
        'karon': karon,
        'logo_url': logo_url,
        'municipality_name': municipality_name,
        'selected_quarter': selected_quarter,
        'quarter_select': quarter_select,
        'signature_name': signature_name,
        'municipalities': municipalities,
        'barangays': barangays,
        'patients': patients,
        'histories': histories,
        'treatments': treatments,
        'category_ii_count': category_ii_count,
        'category_iii_count': category_iii_count,
        'total_count': total_count,
        'total_count_rig': total_count_rig,
        'category_ii_with_rig': category_ii_with_rig,
        'category_iii_with_rig': category_iii_with_rig,
        'category_ii_complete': category_ii_complete,
        'category_ii_incomplete': category_ii_incomplete,
        'category_ii_none': category_ii_none,
        'category_ii_died': category_ii_died,
        'category_iii_complete': category_iii_complete,
        'category_iii_incomplete': category_iii_incomplete,
        'category_iii_none': category_iii_none,
        'category_iii_died': category_iii_died,
        'total_complete': total_complete,
        'total_incomplete': total_incomplete,
        'total_none': total_none,
        'total_died': total_died,
    }

    return render(request, 'monitoring/cohort.html', context)

def pdf_cohort_create1(request):
    user = request.user

    municipality_map = {
        "MAR": "Maripipi",
        "KAW": "Kawayan",
        "NAV": "Naval",
        "CAIB": "Caibiran",
        "ALM": "Almeria",
        "BIL": "Biliran",
        "CUL": "Culaba",
        "CABUC": "Cabucgayan"
    }
    if user.is_superuser and user.code == "NAV":
        municipality_name = "BPH"
    else:
        municipality_name = municipality_map.get(user.code, "Province")

    if user.is_superuser:
        municipal = "Province of Biliran"
    else:
        munici = municipality_map.get(user.code)
        municipal = (f"Municipality of {munici}")



    selected_quarter = '1'
    year = date.today().year
    start_date = date(year, 1, 1)
    end_date = date(year, 3, 31)

    # Retrieve data based on the user type and date range
    if user.is_superuser:
        municipalities = Municipality.objects.all()
        barangays = Barangay.objects.all()
        patients = Patient.objects.all()
        histories = History.objects.filter(date_registered__range=(start_date, end_date))
        treatments = Treatment.objects.filter(
            patient_id__histories__date_registered__range=(start_date, end_date)
        ).distinct()
    else:
        patients = Patient.objects.filter(user=user)
        histories = History.objects.filter(patient_id__in=patients, date_registered__range=(start_date, end_date))
        treatments = Treatment.objects.filter(
            patient_id__histories__date_registered__range=(start_date, end_date)
        ).distinct()
        municipalities = Municipality.objects.filter(muni_id__in=patients.values_list('muni_id', flat=True))
        barangays = Barangay.objects.filter(brgy_id__in=patients.values_list('brgy_id', flat=True))

    # Calculate counts for each category of exposure
    category_i_count = histories.filter(category_of_exposure='I').count()
    category_ii_count = histories.filter(category_of_exposure='II').count()
    category_iii_count = histories.filter(category_of_exposure='III').count()
    total_count = category_i_count + category_ii_count + category_iii_count

    # RIG counts by category
    category_ii_with_rig = histories.filter(
        category_of_exposure='II',
        patient_id__treatments_patient__rig_given__isnull=False,
        date_registered__range=(start_date, end_date)
    ).distinct().count()

    category_iii_with_rig = histories.filter(
        category_of_exposure='III',
        patient_id__treatments_patient__rig_given__isnull=False,
        date_registered__range=(start_date, end_date)
    ).distinct().count()

    total_count_rig = category_ii_with_rig + category_iii_with_rig

    # Outcome counters for each category and date range
    category_ii_complete = category_ii_incomplete = category_ii_none = category_ii_died = 0
    category_iii_complete = category_iii_incomplete = category_iii_none = category_iii_died = 0

    # Process Category II outcomes
    for history in histories.filter(category_of_exposure='II'):
        # Filter treatments associated with the patient within the date range
        treatments = Treatment.objects.filter(
            patient_id=history.patient_id,
            patient_id__histories__date_registered__range=(start_date, end_date)
        ).distinct()

        if history.human_rabies:
            category_ii_died += 1
            continue

        if treatments.exists():
            # Check if all of day0, day3, and day7 are marked True (Complete)
            if treatments.filter(day0_arrived=True, day3_arrived=True, day7_arrived=True).exists():
                category_ii_complete += 1
            # Check if any one of day0, day3, or day7 is False (Incomplete)
            elif treatments.filter(day0_arrived=False).exists() or \
                treatments.filter(day3_arrived=False).exists() or \
                treatments.filter(day7_arrived=False).exists():
                category_ii_incomplete += 1
            # If all day0, day3, and day7 are False (None)
            elif treatments.filter(day0_arrived=False, day3_arrived=False, day7_arrived=False).exists():
                category_ii_none += 1
        else:
            category_ii_none += 1

    # Process Category III outcomes
    for history in histories.filter(category_of_exposure='III'):
        treatments = Treatment.objects.filter(
            patient_id=history.patient_id,
            patient_id__histories__date_registered__range=(start_date, end_date)
        ).distinct()

        if history.human_rabies:
            category_iii_died += 1
            continue

        if treatments.exists():
            # Check if all of day0, day3, and day7 are marked True (Complete)
            if treatments.filter(day0_arrived=True, day3_arrived=True, day7_arrived=True).exists():
                category_iii_complete += 1
            # Check if any one of day0, day3, or day7 is False (Incomplete)
            elif treatments.filter(day0_arrived=False).exists() or \
                treatments.filter(day3_arrived=False).exists() or \
                treatments.filter(day7_arrived=False).exists():
                category_iii_incomplete += 1
            # If all day0, day3, and day7 are False (None)
            elif treatments.filter(day0_arrived=False, day3_arrived=False, day7_arrived=False).exists():
                category_iii_none += 1
        else:
            category_iii_none += 1


    total_complete = category_ii_complete + category_iii_complete
    total_incomplete = category_ii_incomplete + category_iii_incomplete
    total_none = category_ii_none + category_iii_none
    total_died = category_ii_died + category_iii_died

    signature_name = f"{user.first_name} {user.last_name}".strip() if user.first_name or user.last_name else user.username

    if user.logo_image:  # Ensure the user has a logo image
        logo_url = request.build_absolute_uri(user.logo_image.url)
    else:
        logo_url = None  # If no logo is available, set logo_url to None

    karon =  date.today().year

    code_to_muni_name = {
        'ALM': 'Almeria',
        'BIL': 'Biliran',
        'CABUC': 'Cabucgayan',
        'CAIB': 'Caibiran',
        'CUL': 'Culaba',
        'KAW': 'Kawayan',
        'MAR': 'Maripipi',
        'NAV': 'Naval',
    }

    if user.is_superuser:
        coordinator = "Provincial Rabies Coordinator"
        pho = "PHO II"
        # Get the first lead doctor
        doct = Doctor.objects.filter(is_superdoctor=True).first()
        doctor = doct.full_name() if doct else "No Doctor Assigned"
        center = "Biliran Province Hospital"
        center_label = f"Animal Bite Treatment Center: {center}"  # Add the label
    else:
        # Map user code to municipality name
        muni_name = code_to_muni_name.get(user.code)

        if muni_name:
            try:
                # Find the municipality with this name
                municipality = Municipality.objects.get(muni_name=muni_name)
                # Filter doctors by this municipality
                doct = Doctor.objects.filter(muni_id=municipality).first()
                doctor = doct.full_name() if doct else "No Doctor Assigned"
                coordinator = f"Municipality of {municipality.muni_name}"
                pho = f"Doctor of {municipality.muni_name}"
                center = muni_name  # Use the user's code to determine the center
                center_label = f"Animal Bite Treatment Center: {center} Animal Bite Treatment Center"  # Add the label
            except Municipality.DoesNotExist:
                # Handle case where the municipality name does not exist
                coordinator = "Unknown Municipality"
                pho = "No Doctor Assigned"
                doctor = "No Doctor Assigned"
                center = "Unknown Center"
                center_label = f"Animal Bite Treatment Center: {center}"  # Add the label
        else:
            # Handle case where the user's code does not match any municipality
            coordinator = "Unknown Municipality"
            pho = "No Doctor Assigned"
            doctor = "No Doctor Assigned"
            center = "Unknown Center"
            center_label = f"Animal Bite Treatment Center: {center} Animal Bite Treatment Center"  # Add the label

    context = {
        'doctor': doctor,
        'coordinator': coordinator,
        'center': center,
        'center_label':center_label,
        'pho': pho,
        'karon': karon,
        'logo_url': logo_url,
        'municipality_name': municipality_name,
        'selected_quarter': selected_quarter,
        'signature_name': signature_name,
        'municipalities': municipalities,
        'municipal': municipal,
        'barangays': barangays,
        'patients': patients,
        'histories': histories,
        'treatments': treatments,
        'category_ii_count': category_ii_count,
        'category_iii_count': category_iii_count,
        'total_count': total_count,
        'total_count_rig': total_count_rig,
        'category_ii_with_rig': category_ii_with_rig,
        'category_iii_with_rig': category_iii_with_rig,
        'category_ii_complete': category_ii_complete,
        'category_ii_incomplete': category_ii_incomplete,
        'category_ii_none': category_ii_none,
        'category_ii_died': category_ii_died,
        'category_iii_complete': category_iii_complete,
        'category_iii_incomplete': category_iii_incomplete,
        'category_iii_none': category_iii_none,
        'category_iii_died': category_iii_died,
        'total_complete': total_complete,
        'total_incomplete': total_incomplete,
        'total_none': total_none,
        'total_died': total_died,
    }
    template = loader.get_template('monitoring/cohort_pdf1.html')
    html = template.render(context)

    """ # Specify path to wkhtmltopdf executable
    path_to_wkhtmltopdf = r"C:\wkhtmltox\bin\wkhtmltopdf.exe"
    
    # Configure pdfkit with the path
    config = pdfkit.configuration(wkhtmltopdf=path_to_wkhtmltopdf)
 """
    config = pdfkit.configuration(wkhtmltopdf=WKHTMLTOPDF_PATH)

    # Options for landscape orientation
    options = {
        'orientation': 'Landscape',  # Set to Landscape
        'page-size': 'A4',  # A4 paper size, you can adjust if needed
        'footer-left': center_label,  # Add page numbers
        'footer-right': 'Page [page] of [toPage]',  # Add page numbers
        'footer-font-size': '10',  # Adjust font size for the footer
        'footer-spacing': '5',  # Space between footer and content
        'margin-bottom': '15mm',  # Ensure space for the footer
    }
    
    # Generate PDF from HTML using pdfkit with configuration
    pdf = pdfkit.from_string(html, configuration=config,options=options)

    # Return the PDF as a responsev 
    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment;filename="1st Quarter Cohort.pdf"'

    return response

def pdf_cohort_create2(request):
    user = request.user

    municipality_map = {
        "MAR": "Maripipi",
        "KAW": "Kawayan",
        "NAV": "Naval",
        "CAIB": "Caibiran",
        "ALM": "Almeria",
        "BIL": "Biliran",
        "CUL": "Culaba",
        "CABUC": "Cabucgayan"
    }
    if user.is_superuser and user.code == "NAV":
        municipality_name = "BPH"
    else:
        municipality_name = municipality_map.get(user.code, "Province")
    
    if user.is_superuser:
        municipal = "Province of Biliran"
    else:
        munici = municipality_map.get(user.code)
        municipal = (f"Municipality of {munici}")

    selected_quarter = '2'
    year = date.today().year
    start_date = date(year, 4, 1)
    end_date = date(year, 6, 30)

    # Retrieve data based on the user type and date range
    if user.is_superuser:
        municipalities = Municipality.objects.all()
        barangays = Barangay.objects.all()
        patients = Patient.objects.all()
        histories = History.objects.filter(date_registered__range=(start_date, end_date))
        treatments = Treatment.objects.filter(
            patient_id__histories__date_registered__range=(start_date, end_date)
        ).distinct()
    else:
        patients = Patient.objects.filter(user=user)
        histories = History.objects.filter(patient_id__in=patients, date_registered__range=(start_date, end_date))
        treatments = Treatment.objects.filter(
            patient_id__histories__date_registered__range=(start_date, end_date)
        ).distinct()
        municipalities = Municipality.objects.filter(muni_id__in=patients.values_list('muni_id', flat=True))
        barangays = Barangay.objects.filter(brgy_id__in=patients.values_list('brgy_id', flat=True))

    # Calculate counts for each category of exposure
    category_i_count = histories.filter(category_of_exposure='I').count()
    category_ii_count = histories.filter(category_of_exposure='II').count()
    category_iii_count = histories.filter(category_of_exposure='III').count()
    total_count = category_i_count + category_ii_count + category_iii_count

    # RIG counts by category
    category_ii_with_rig = histories.filter(
        category_of_exposure='II',
        patient_id__treatments_patient__rig_given__isnull=False,
        date_registered__range=(start_date, end_date)
    ).distinct().count()

    category_iii_with_rig = histories.filter(
        category_of_exposure='III',
        patient_id__treatments_patient__rig_given__isnull=False,
        date_registered__range=(start_date, end_date)
    ).distinct().count()

    total_count_rig = category_ii_with_rig + category_iii_with_rig

    # Outcome counters for each category and date range
    category_ii_complete = category_ii_incomplete = category_ii_none = category_ii_died = 0
    category_iii_complete = category_iii_incomplete = category_iii_none = category_iii_died = 0

    # Process Category II outcomes
    for history in histories.filter(category_of_exposure='II'):
        # Filter treatments associated with the patient within the date range
        treatments = Treatment.objects.filter(
            patient_id=history.patient_id,
            patient_id__histories__date_registered__range=(start_date, end_date)
        ).distinct()

        if history.human_rabies:
            category_ii_died += 1
            continue

        if treatments.exists():
            # Check if all of day0, day3, and day7 are marked True (Complete)
            if treatments.filter(day0_arrived=True, day3_arrived=True, day7_arrived=True).exists():
                category_ii_complete += 1
            # Check if any one of day0, day3, or day7 is False (Incomplete)
            elif treatments.filter(day0_arrived=False).exists() or \
                treatments.filter(day3_arrived=False).exists() or \
                treatments.filter(day7_arrived=False).exists():
                category_ii_incomplete += 1
            # If all day0, day3, and day7 are False (None)
            elif treatments.filter(day0_arrived=False, day3_arrived=False, day7_arrived=False).exists():
                category_ii_none += 1
        else:
            category_ii_none += 1

    # Process Category III outcomes
    for history in histories.filter(category_of_exposure='III'):
        treatments = Treatment.objects.filter(
            patient_id=history.patient_id,
            patient_id__histories__date_registered__range=(start_date, end_date)
        ).distinct()

        if history.human_rabies:
            category_iii_died += 1
            continue

        if treatments.exists():
            # Check if all of day0, day3, and day7 are marked True (Complete)
            if treatments.filter(day0_arrived=True, day3_arrived=True, day7_arrived=True).exists():
                category_iii_complete += 1
            # Check if any one of day0, day3, or day7 is False (Incomplete)
            elif treatments.filter(day0_arrived=False).exists() or \
                treatments.filter(day3_arrived=False).exists() or \
                treatments.filter(day7_arrived=False).exists():
                category_iii_incomplete += 1
            # If all day0, day3, and day7 are False (None)
            elif treatments.filter(day0_arrived=False, day3_arrived=False, day7_arrived=False).exists():
                category_iii_none += 1
        else:
            category_iii_none += 1


    total_complete = category_ii_complete + category_iii_complete
    total_incomplete = category_ii_incomplete + category_iii_incomplete
    total_none = category_ii_none + category_iii_none
    total_died = category_ii_died + category_iii_died

    signature_name = f"{user.first_name} {user.last_name}".strip() if user.first_name or user.last_name else user.username

    if user.logo_image:  # Ensure the user has a logo image
        logo_url = request.build_absolute_uri(user.logo_image.url)
    else:
        logo_url = None  # If no logo is available, set logo_url to None

    karon =  date.today().year

    code_to_muni_name = {
        'ALM': 'Almeria',
        'BIL': 'Biliran',
        'CABUC': 'Cabucgayan',
        'CAIB': 'Caibiran',
        'CUL': 'Culaba',
        'KAW': 'Kawayan',
        'MAR': 'Maripipi',
        'NAV': 'Naval',
    }

    if user.is_superuser:
        coordinator = "Provincial Rabies Coordinator"
        pho = "PHO II"
        # Get the first lead doctor
        doct = Doctor.objects.filter(is_superdoctor=True).first()
        doctor = doct.full_name() if doct else "No Doctor Assigned"
        center = "Biliran Province Hospital"
        center_label = f"Animal Bite Treatment Center: {center}"  # Add the label
    else:
        # Map user code to municipality name
        muni_name = code_to_muni_name.get(user.code)

        if muni_name:
            try:
                # Find the municipality with this name
                municipality = Municipality.objects.get(muni_name=muni_name)
                # Filter doctors by this municipality
                doct = Doctor.objects.filter(muni_id=municipality).first()
                doctor = doct.full_name() if doct else "No Doctor Assigned"
                coordinator = f"Municipality of {municipality.muni_name}"
                pho = f"Doctor of {municipality.muni_name}"
                center = muni_name  # Use the user's code to determine the center
                center_label = f"Animal Bite Treatment Center: {center} Animal Bite Treatment Center"  # Add the label
            except Municipality.DoesNotExist:
                # Handle case where the municipality name does not exist
                coordinator = "Unknown Municipality"
                pho = "No Doctor Assigned"
                doctor = "No Doctor Assigned"
                center = "Unknown Center"
                center_label = f"Animal Bite Treatment Center: {center}"  # Add the label
        else:
            # Handle case where the user's code does not match any municipality
            coordinator = "Unknown Municipality"
            pho = "No Doctor Assigned"
            doctor = "No Doctor Assigned"
            center = "Unknown Center"
            center_label = f"Animal Bite Treatment Center: {center} Animal Bite Treatment Center"  # Add the label

    context = {
        'doctor': doctor,
        'coordinator': coordinator,
        'center': center,
        'center_label': center_label,
        'pho': pho,
        'karon': karon,
        'logo_url': logo_url,
        'municipality_name': municipality_name,
        'selected_quarter': selected_quarter,
        'signature_name': signature_name,
        'municipalities': municipalities,
        'municipal': municipal,
        'barangays': barangays,
        'patients': patients,
        'histories': histories,
        'treatments': treatments,
        'category_ii_count': category_ii_count,
        'category_iii_count': category_iii_count,
        'total_count': total_count,
        'total_count_rig': total_count_rig,
        'category_ii_with_rig': category_ii_with_rig,
        'category_iii_with_rig': category_iii_with_rig,
        'category_ii_complete': category_ii_complete,
        'category_ii_incomplete': category_ii_incomplete,
        'category_ii_none': category_ii_none,
        'category_ii_died': category_ii_died,
        'category_iii_complete': category_iii_complete,
        'category_iii_incomplete': category_iii_incomplete,
        'category_iii_none': category_iii_none,
        'category_iii_died': category_iii_died,
        'total_complete': total_complete,
        'total_incomplete': total_incomplete,
        'total_none': total_none,
        'total_died': total_died,
    }
    template = loader.get_template('monitoring/cohort_pdf2.html')
    html = template.render(context)

    """ # Specify path to wkhtmltopdf executable
    path_to_wkhtmltopdf = r"C:\wkhtmltox\bin\wkhtmltopdf.exe"
    
    # Configure pdfkit with the path
    config = pdfkit.configuration(wkhtmltopdf=path_to_wkhtmltopdf) """

    config = pdfkit.configuration(wkhtmltopdf=WKHTMLTOPDF_PATH)

    # Options for landscape orientation
    options = {
        'orientation': 'Landscape',  # Set to Landscape
        'page-size': 'A4',  # A4 paper size, you can adjust if needed
        'footer-left': center_label,  # Add page numbers
        'footer-right': 'Page [page] of [toPage]',  # Add page numbers
        'footer-font-size': '10',  # Adjust font size for the footer
        'footer-spacing': '5',  # Space between footer and content
        'margin-bottom': '15mm',  # Ensure space for the footer
    }
    
    # Generate PDF from HTML using pdfkit with configuration
    pdf = pdfkit.from_string(html, configuration=config,options=options)

    # Return the PDF as a responsev 
    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment;filename="2nd Quarter Cohort.pdf"'

    return response

def pdf_cohort_create3(request):
    user = request.user

    municipality_map = {
        "MAR": "Maripipi",
        "KAW": "Kawayan",
        "NAV": "Naval",
        "CAIB": "Caibiran",
        "ALM": "Almeria",
        "BIL": "Biliran",
        "CUL": "Culaba",
        "CABUC": "Cabucgayan"
    }
    if user.is_superuser and user.code == "NAV":
        municipality_name = "BPH"
    else:
        municipality_name = municipality_map.get(user.code, "Province")

    if user.is_superuser:
        municipal = "Province of Biliran"
    else:
        munici = municipality_map.get(user.code)
        municipal = (f"Municipality of {munici}")
        
    selected_quarter = '3'
    year = date.today().year
    start_date = date(year, 7, 1)
    end_date = date(year, 9, 30)

    # Retrieve data based on the user type and date range
    if user.is_superuser:
        municipalities = Municipality.objects.all()
        barangays = Barangay.objects.all()
        patients = Patient.objects.all()
        histories = History.objects.filter(date_registered__range=(start_date, end_date))
        treatments = Treatment.objects.filter(
            patient_id__histories__date_registered__range=(start_date, end_date)
        ).distinct()
    else:
        patients = Patient.objects.filter(user=user)
        histories = History.objects.filter(patient_id__in=patients, date_registered__range=(start_date, end_date))
        treatments = Treatment.objects.filter(
            patient_id__histories__date_registered__range=(start_date, end_date)
        ).distinct()
        municipalities = Municipality.objects.filter(muni_id__in=patients.values_list('muni_id', flat=True))
        barangays = Barangay.objects.filter(brgy_id__in=patients.values_list('brgy_id', flat=True))

    # Calculate counts for each category of exposure
    category_i_count = histories.filter(category_of_exposure='I').count()
    category_ii_count = histories.filter(category_of_exposure='II').count()
    category_iii_count = histories.filter(category_of_exposure='III').count()
    total_count = category_i_count + category_ii_count + category_iii_count

    # RIG counts by category
    category_ii_with_rig = histories.filter(
        category_of_exposure='II',
        patient_id__treatments_patient__rig_given__isnull=False,
        date_registered__range=(start_date, end_date)
    ).distinct().count()

    category_iii_with_rig = histories.filter(
        category_of_exposure='III',
        patient_id__treatments_patient__rig_given__isnull=False,
        date_registered__range=(start_date, end_date)
    ).distinct().count()

    total_count_rig = category_ii_with_rig + category_iii_with_rig

    # Outcome counters for each category and date range
    category_ii_complete = category_ii_incomplete = category_ii_none = category_ii_died = 0
    category_iii_complete = category_iii_incomplete = category_iii_none = category_iii_died = 0

    # Process Category II outcomes
    for history in histories.filter(category_of_exposure='II'):
        # Filter treatments associated with the patient within the date range
        treatments = Treatment.objects.filter(
            patient_id=history.patient_id,
            patient_id__histories__date_registered__range=(start_date, end_date)
        ).distinct()

        if history.human_rabies:
            category_ii_died += 1
            continue

        if treatments.exists():
            # Check if all of day0, day3, and day7 are marked True (Complete)
            if treatments.filter(day0_arrived=True, day3_arrived=True, day7_arrived=True).exists():
                category_ii_complete += 1
            # Check if any one of day0, day3, or day7 is False (Incomplete)
            elif treatments.filter(day0_arrived=False).exists() or \
                treatments.filter(day3_arrived=False).exists() or \
                treatments.filter(day7_arrived=False).exists():
                category_ii_incomplete += 1
            # If all day0, day3, and day7 are False (None)
            elif treatments.filter(day0_arrived=False, day3_arrived=False, day7_arrived=False).exists():
                category_ii_none += 1
        else:
            category_ii_none += 1

    # Process Category III outcomes
    for history in histories.filter(category_of_exposure='III'):
        treatments = Treatment.objects.filter(
            patient_id=history.patient_id,
            patient_id__histories__date_registered__range=(start_date, end_date)
        ).distinct()

        if history.human_rabies:
            category_iii_died += 1
            continue

        if treatments.exists():
            # Check if all of day0, day3, and day7 are marked True (Complete)
            if treatments.filter(day0_arrived=True, day3_arrived=True, day7_arrived=True).exists():
                category_iii_complete += 1
            # Check if any one of day0, day3, or day7 is False (Incomplete)
            elif treatments.filter(day0_arrived=False).exists() or \
                treatments.filter(day3_arrived=False).exists() or \
                treatments.filter(day7_arrived=False).exists():
                category_iii_incomplete += 1
            # If all day0, day3, and day7 are False (None)
            elif treatments.filter(day0_arrived=False, day3_arrived=False, day7_arrived=False).exists():
                category_iii_none += 1
        else:
            category_iii_none += 1


    total_complete = category_ii_complete + category_iii_complete
    total_incomplete = category_ii_incomplete + category_iii_incomplete
    total_none = category_ii_none + category_iii_none
    total_died = category_ii_died + category_iii_died

    signature_name = f"{user.first_name} {user.last_name}".strip() if user.first_name or user.last_name else user.username

    if user.logo_image:  # Ensure the user has a logo image
        logo_url = request.build_absolute_uri(user.logo_image.url)
    else:
        logo_url = None  # If no logo is available, set logo_url to None

    karon =  date.today().year

    code_to_muni_name = {
        'ALM': 'Almeria',
        'BIL': 'Biliran',
        'CABUC': 'Cabucgayan',
        'CAIB': 'Caibiran',
        'CUL': 'Culaba',
        'KAW': 'Kawayan',
        'MAR': 'Maripipi',
        'NAV': 'Naval',
    }

    if user.is_superuser:
        coordinator = "Provincial Rabies Coordinator"
        pho = "PHO II"
        # Get the first lead doctor
        doct = Doctor.objects.filter(is_superdoctor=True).first()
        doctor = doct.full_name() if doct else "No Doctor Assigned"
        center = "Biliran Province Hospital"
        center_label = f"Animal Bite Treatment Center: {center}"  # Add the label
    else:
        # Map user code to municipality name
        muni_name = code_to_muni_name.get(user.code)

        if muni_name:
            try:
                # Find the municipality with this name
                municipality = Municipality.objects.get(muni_name=muni_name)
                # Filter doctors by this municipality
                doct = Doctor.objects.filter(muni_id=municipality).first()
                doctor = doct.full_name() if doct else "No Doctor Assigned"
                coordinator = f"Municipality of {municipality.muni_name}"
                pho = f"Doctor of {municipality.muni_name}"
                center = muni_name  # Use the user's code to determine the center
                center_label = f"Animal Bite Treatment Center: {center} Animal Bite Treatment Center"  # Add the label
            except Municipality.DoesNotExist:
                # Handle case where the municipality name does not exist
                coordinator = "Unknown Municipality"
                pho = "No Doctor Assigned"
                doctor = "No Doctor Assigned"
                center = "Unknown Center"
                center_label = f"Animal Bite Treatment Center: {center}"  # Add the label
        else:
            # Handle case where the user's code does not match any municipality
            coordinator = "Unknown Municipality"
            pho = "No Doctor Assigned"
            doctor = "No Doctor Assigned"
            center = "Unknown Center"
            center_label = f"Animal Bite Treatment Center: {center} Animal Bite Treatment Center"  # Add the label


    context = {
        'doctor': doctor,
        'coordinator': coordinator,
        'center': center,
        'center_label': center_label,
        'pho': pho,
        'karon': karon,
        'logo_url': logo_url,
        'municipality_name': municipality_name,
        'signature_name': signature_name,
        'municipalities': municipalities,
        'municipal': municipal,
        'barangays': barangays,
        'patients': patients,
        'histories': histories,
        'treatments': treatments,
        'category_ii_count': category_ii_count,
        'category_iii_count': category_iii_count,
        'total_count': total_count,
        'total_count_rig': total_count_rig,
        'category_ii_with_rig': category_ii_with_rig,
        'category_iii_with_rig': category_iii_with_rig,
        'category_ii_complete': category_ii_complete,
        'category_ii_incomplete': category_ii_incomplete,
        'category_ii_none': category_ii_none,
        'category_ii_died': category_ii_died,
        'category_iii_complete': category_iii_complete,
        'category_iii_incomplete': category_iii_incomplete,
        'category_iii_none': category_iii_none,
        'category_iii_died': category_iii_died,
        'total_complete': total_complete,
        'total_incomplete': total_incomplete,
        'total_none': total_none,
        'total_died': total_died,
    }
    template = loader.get_template('monitoring/cohort_pdf3.html')
    html = template.render(context)

    """ # Specify path to wkhtmltopdf executable
    path_to_wkhtmltopdf = r"C:\wkhtmltox\bin\wkhtmltopdf.exe"
    
    # Configure pdfkit with the path
    config = pdfkit.configuration(wkhtmltopdf=path_to_wkhtmltopdf) """

    config = pdfkit.configuration(wkhtmltopdf=WKHTMLTOPDF_PATH)

    # Options for landscape orientation
    options = {
        'orientation': 'Landscape',  # Set to Landscape
        'page-size': 'A4',  # A4 paper size, you can adjust if needed
        'footer-left': center_label,  # Add page numbers
        'footer-right': 'Page [page] of [toPage]',  # Add page numbers
        'footer-font-size': '10',  # Adjust font size for the footer
        'footer-spacing': '5',  # Space between footer and content
        'margin-bottom': '15mm',  # Ensure space for the footer
    }
    
    # Generate PDF from HTML using pdfkit with configuration
    pdf = pdfkit.from_string(html, configuration=config,options=options)

    # Return the PDF as a responsev 
    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment;filename="3rd Quarter Cohort.pdf"'

    return response

def pdf_cohort_create4(request):
    user = request.user

    municipality_map = {
        "MAR": "Maripipi",
        "KAW": "Kawayan",
        "NAV": "Naval",
        "CAIB": "Caibiran",
        "ALM": "Almeria",
        "BIL": "Biliran",
        "CUL": "Culaba",
        "CABUC": "Cabucgayan"
    }
    if user.is_superuser and user.code == "NAV":
        municipality_name = "BPH"
    else:
        municipality_name = municipality_map.get(user.code, "Province")

    if user.is_superuser:
        municipal = "Province of Biliran"
    else:
        munici = municipality_map.get(user.code)
        municipal = (f"Municipality of {munici}")
        
    # Only set for 2nd quarter (April 1 to June 30)
    selected_quarter = '4'
    year = date.today().year
    # 4th Quarter
    start_date = date(year, 10, 1)
    end_date = date(year, 12, 31)

    # Retrieve data based on the user type and date range
    if user.is_superuser:
        municipalities = Municipality.objects.all()
        barangays = Barangay.objects.all()
        patients = Patient.objects.all()
        histories = History.objects.filter(date_registered__range=(start_date, end_date))
        treatments = Treatment.objects.filter(
            patient_id__histories__date_registered__range=(start_date, end_date)
        ).distinct()
    else:
        patients = Patient.objects.filter(user=user)
        histories = History.objects.filter(patient_id__in=patients, date_registered__range=(start_date, end_date))
        treatments = Treatment.objects.filter(
            patient_id__histories__date_registered__range=(start_date, end_date)
        ).distinct()
        municipalities = Municipality.objects.filter(muni_id__in=patients.values_list('muni_id', flat=True))
        barangays = Barangay.objects.filter(brgy_id__in=patients.values_list('brgy_id', flat=True))

    # Calculate counts for each category of exposure
    category_i_count = histories.filter(category_of_exposure='I').count()
    category_ii_count = histories.filter(category_of_exposure='II').count()
    category_iii_count = histories.filter(category_of_exposure='III').count()
    total_count = category_i_count + category_ii_count + category_iii_count

    # RIG counts by category
    category_ii_with_rig = histories.filter(
        category_of_exposure='II',
        patient_id__treatments_patient__rig_given__isnull=False,
        date_registered__range=(start_date, end_date)
    ).distinct().count()

    category_iii_with_rig = histories.filter(
        category_of_exposure='III',
        patient_id__treatments_patient__rig_given__isnull=False,
        date_registered__range=(start_date, end_date)
    ).distinct().count()

    total_count_rig = category_ii_with_rig + category_iii_with_rig

    # Outcome counters for each category and date range
    category_ii_complete = category_ii_incomplete = category_ii_none = category_ii_died = 0
    category_iii_complete = category_iii_incomplete = category_iii_none = category_iii_died = 0

    # Process Category II outcomes
    for history in histories.filter(category_of_exposure='II'):
        # Filter treatments associated with the patient within the date range
        treatments = Treatment.objects.filter(
            patient_id=history.patient_id,
            patient_id__histories__date_registered__range=(start_date, end_date)
        ).distinct()

        if history.human_rabies:
            category_ii_died += 1
            continue

        if treatments.exists():
            # Check if all of day0, day3, and day7 are marked True (Complete)
            if treatments.filter(day0_arrived=True, day3_arrived=True, day7_arrived=True).exists():
                category_ii_complete += 1
            # Check if any one of day0, day3, or day7 is False (Incomplete)
            elif treatments.filter(day0_arrived=False).exists() or \
                treatments.filter(day3_arrived=False).exists() or \
                treatments.filter(day7_arrived=False).exists():
                category_ii_incomplete += 1
            # If all day0, day3, and day7 are False (None)
            elif treatments.filter(day0_arrived=False, day3_arrived=False, day7_arrived=False).exists():
                category_ii_none += 1
        else:
            category_ii_none += 1

    # Process Category III outcomes
    for history in histories.filter(category_of_exposure='III'):
        treatments = Treatment.objects.filter(
            patient_id=history.patient_id,
            patient_id__histories__date_registered__range=(start_date, end_date)
        ).distinct()

        if history.human_rabies:
            category_iii_died += 1
            continue

        if treatments.exists():
            # Check if all of day0, day3, and day7 are marked True (Complete)
            if treatments.filter(day0_arrived=True, day3_arrived=True, day7_arrived=True).exists():
                category_iii_complete += 1
            # Check if any one of day0, day3, or day7 is False (Incomplete)
            elif treatments.filter(day0_arrived=False).exists() or \
                treatments.filter(day3_arrived=False).exists() or \
                treatments.filter(day7_arrived=False).exists():
                category_iii_incomplete += 1
            # If all day0, day3, and day7 are False (None)
            elif treatments.filter(day0_arrived=False, day3_arrived=False, day7_arrived=False).exists():
                category_iii_none += 1
        else:
            category_iii_none += 1

    total_complete = category_ii_complete + category_iii_complete
    total_incomplete = category_ii_incomplete + category_iii_incomplete
    total_none = category_ii_none + category_iii_none
    total_died = category_ii_died + category_iii_died

    signature_name = f"{user.first_name} {user.last_name}".strip() if user.first_name or user.last_name else user.username

    if user.logo_image:  # Ensure the user has a logo image
        logo_url = request.build_absolute_uri(user.logo_image.url)
    else:
        logo_url = None  # If no logo is available, set logo_url to None

    karon =  date.today().year

    code_to_muni_name = {
        'ALM': 'Almeria',
        'BIL': 'Biliran',
        'CABUC': 'Cabucgayan',
        'CAIB': 'Caibiran',
        'CUL': 'Culaba',
        'KAW': 'Kawayan',
        'MAR': 'Maripipi',
        'NAV': 'Naval',
    }

    if user.is_superuser:
        coordinator = "Provincial Rabies Coordinator"
        pho = "PHO II"
        # Get the first lead doctor
        doct = Doctor.objects.filter(is_superdoctor=True).first()
        doctor = doct.full_name() if doct else "No Doctor Assigned"
        center = "Biliran Province Hospital"
        center_label = f"Animal Bite Treatment Center: {center}"  # Add the label
    else:
        # Map user code to municipality name
        muni_name = code_to_muni_name.get(user.code)

        if muni_name:
            try:
                # Find the municipality with this name
                municipality = Municipality.objects.get(muni_name=muni_name)
                # Filter doctors by this municipality
                doct = Doctor.objects.filter(muni_id=municipality).first()
                doctor = doct.full_name() if doct else "No Doctor Assigned"
                coordinator = f"Municipality of {municipality.muni_name}"
                pho = f"Doctor of {municipality.muni_name}"
                center = muni_name  # Use the user's code to determine the center
                center_label = f"Animal Bite Treatment Center: {center} Animal Bite Treatment Center"  # Add the label
            except Municipality.DoesNotExist:
                # Handle case where the municipality name does not exist
                coordinator = "Unknown Municipality"
                pho = "No Doctor Assigned"
                doctor = "No Doctor Assigned"
                center = "Unknown Center"
                center_label = f"Animal Bite Treatment Center: {center}"  # Add the label
        else:
            # Handle case where the user's code does not match any municipality
            coordinator = "Unknown Municipality"
            pho = "No Doctor Assigned"
            doctor = "No Doctor Assigned"
            center = "Unknown Center"
            center_label = f"Animal Bite Treatment Center: {center} Animal Bite Treatment Center"  # Add the label


    context = {
        'doctor': doctor,
        'coordinator': coordinator,
        'center': center,
        'center_label': center_label,
        'pho': pho,
        'karon': karon,
        'logo_url': logo_url,
        'municipality_name': municipality_name,
        'selected_quarter': selected_quarter,
        'signature_name': signature_name,
        'municipalities': municipalities,
        'municipal': municipal,
        'barangays': barangays,
        'patients': patients,
        'histories': histories,
        'treatments': treatments,
        'category_ii_count': category_ii_count,
        'category_iii_count': category_iii_count,
        'total_count': total_count,
        'total_count_rig': total_count_rig,
        'category_ii_with_rig': category_ii_with_rig,
        'category_iii_with_rig': category_iii_with_rig,
        'category_ii_complete': category_ii_complete,
        'category_ii_incomplete': category_ii_incomplete,
        'category_ii_none': category_ii_none,
        'category_ii_died': category_ii_died,
        'category_iii_complete': category_iii_complete,
        'category_iii_incomplete': category_iii_incomplete,
        'category_iii_none': category_iii_none,
        'category_iii_died': category_iii_died,
        'total_complete': total_complete,
        'total_incomplete': total_incomplete,
        'total_none': total_none,
        'total_died': total_died,
    }
    template = loader.get_template('monitoring/cohort_pdf4.html')
    html = template.render(context)

    """ # Specify path to wkhtmltopdf executable
    path_to_wkhtmltopdf = r"C:\wkhtmltox\bin\wkhtmltopdf.exe"
    
    # Configure pdfkit with the path
    config = pdfkit.configuration(wkhtmltopdf=path_to_wkhtmltopdf) """


    config = pdfkit.configuration(wkhtmltopdf=WKHTMLTOPDF_PATH)

    # Options for landscape orientation
    options = {
        'orientation': 'Landscape',  # Set to Landscape
        'page-size': 'A4',  # A4 paper size, you can adjust if needed
        'footer-left': center_label,  # Add page numbers
        'footer-right': 'Page [page] of [toPage]',  # Add page numbers
        'footer-font-size': '10',  # Adjust font size for the footer
        'footer-spacing': '5',  # Space between footer and content
        'margin-bottom': '15mm',  # Ensure space for the footer
    }
    
    # Generate PDF from HTML using pdfkit with configuration
    pdf = pdfkit.from_string(html, configuration=config,options=options)

    # Return the PDF as a responsev 
    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment;filename="4th Quarter Cohort.pdf"'

    return response

def pdf_cohort_create_annual(request):
    user = request.user

    municipality_map = {
        "MAR": "Maripipi",
        "KAW": "Kawayan",
        "NAV": "Naval",
        "CAIB": "Caibiran",
        "ALM": "Almeria",
        "BIL": "Biliran",
        "CUL": "Culaba",
        "CABUC": "Cabucgayan"
    }
    if user.is_superuser and user.code == "NAV":
        municipality_name = "BPH"
    else:
        municipality_name = municipality_map.get(user.code, "Province")

    if user.is_superuser:
        municipal = "Province of Biliran"
    else:
        munici = municipality_map.get(user.code)
        municipal = (f"Municipality of {munici}")
            
    # Only set for 2nd quarter (April 1 to June 30)
    selected_quarter = 'annual'
    year = date.today().year
    # 4th Quarter
    start_date = date(year, 1, 1)
    end_date = date(year, 12, 31)

    # Retrieve data based on the user type and date range
    if user.is_superuser:
        municipalities = Municipality.objects.all()
        barangays = Barangay.objects.all()
        patients = Patient.objects.all()
        histories = History.objects.filter(date_registered__range=(start_date, end_date))
        treatments = Treatment.objects.filter(
            patient_id__histories__date_registered__range=(start_date, end_date)
        ).distinct()
    else:
        patients = Patient.objects.filter(user=user)
        histories = History.objects.filter(patient_id__in=patients, date_registered__range=(start_date, end_date))
        treatments = Treatment.objects.filter(
            patient_id__histories__date_registered__range=(start_date, end_date)
        ).distinct()
        municipalities = Municipality.objects.filter(muni_id__in=patients.values_list('muni_id', flat=True))
        barangays = Barangay.objects.filter(brgy_id__in=patients.values_list('brgy_id', flat=True))

    # Calculate counts for each category of exposure
    category_i_count = histories.filter(category_of_exposure='I').count()
    category_ii_count = histories.filter(category_of_exposure='II').count()
    category_iii_count = histories.filter(category_of_exposure='III').count()
    total_count = category_i_count + category_ii_count + category_iii_count

    # RIG counts by category
    category_ii_with_rig = histories.filter(
        category_of_exposure='II',
        patient_id__treatments_patient__rig_given__isnull=False,
        date_registered__range=(start_date, end_date)
    ).distinct().count()

    category_iii_with_rig = histories.filter(
        category_of_exposure='III',
        patient_id__treatments_patient__rig_given__isnull=False,
        date_registered__range=(start_date, end_date)
    ).distinct().count()

    total_count_rig = category_ii_with_rig + category_iii_with_rig

    # Outcome counters for each category and date range
    category_ii_complete = category_ii_incomplete = category_ii_none = category_ii_died = 0
    category_iii_complete = category_iii_incomplete = category_iii_none = category_iii_died = 0

    # Process Category II outcomes
    for history in histories.filter(category_of_exposure='II'):
        # Filter treatments associated with the patient within the date range
        treatments = Treatment.objects.filter(
            patient_id=history.patient_id,
            patient_id__histories__date_registered__range=(start_date, end_date)
        ).distinct()

        if history.human_rabies:
            category_ii_died += 1
            continue

        if treatments.exists():
            # Check if all of day0, day3, and day7 are marked True (Complete)
            if treatments.filter(day0_arrived=True, day3_arrived=True, day7_arrived=True).exists():
                category_ii_complete += 1
            # Check if any one of day0, day3, or day7 is False (Incomplete)
            elif treatments.filter(day0_arrived=False).exists() or \
                treatments.filter(day3_arrived=False).exists() or \
                treatments.filter(day7_arrived=False).exists():
                category_ii_incomplete += 1
            # If all day0, day3, and day7 are False (None)
            elif treatments.filter(day0_arrived=False, day3_arrived=False, day7_arrived=False).exists():
                category_ii_none += 1
        else:
            category_ii_none += 1

    # Process Category III outcomes
    for history in histories.filter(category_of_exposure='III'):
        treatments = Treatment.objects.filter(
            patient_id=history.patient_id,
            patient_id__histories__date_registered__range=(start_date, end_date)
        ).distinct()

        if history.human_rabies:
            category_iii_died += 1
            continue

        if treatments.exists():
            # Check if all of day0, day3, and day7 are marked True (Complete)
            if treatments.filter(day0_arrived=True, day3_arrived=True, day7_arrived=True).exists():
                category_iii_complete += 1
            # Check if any one of day0, day3, or day7 is False (Incomplete)
            elif treatments.filter(day0_arrived=False).exists() or \
                treatments.filter(day3_arrived=False).exists() or \
                treatments.filter(day7_arrived=False).exists():
                category_iii_incomplete += 1
            # If all day0, day3, and day7 are False (None)
            elif treatments.filter(day0_arrived=False, day3_arrived=False, day7_arrived=False).exists():
                category_iii_none += 1
        else:
            category_iii_none += 1

    total_complete = category_ii_complete + category_iii_complete
    total_incomplete = category_ii_incomplete + category_iii_incomplete
    total_none = category_ii_none + category_iii_none
    total_died = category_ii_died + category_iii_died

    signature_name = f"{user.first_name} {user.last_name}".strip() if user.first_name or user.last_name else user.username

    if user.logo_image:  # Ensure the user has a logo image
        logo_url = request.build_absolute_uri(user.logo_image.url)
    else:
        logo_url = None  # If no logo is available, set logo_url to None

    karon =  date.today().year
    
    code_to_muni_name = {
        'ALM': 'Almeria',
        'BIL': 'Biliran',
        'CABUC': 'Cabucgayan',
        'CAIB': 'Caibiran',
        'CUL': 'Culaba',
        'KAW': 'Kawayan',
        'MAR': 'Maripipi',
        'NAV': 'Naval',
    }

    if user.is_superuser:
        coordinator = "Provincial Rabies Coordinator"
        pho = "PHO II"
        # Get the first lead doctor
        doct = Doctor.objects.filter(is_superdoctor=True).first()
        doctor = doct.full_name() if doct else "No Doctor Assigned"
        center = "Biliran Province Hospital"
        center_label = f"Animal Bite Treatment Center: {center}"  # Add the label
    else:
        # Map user code to municipality name
        muni_name = code_to_muni_name.get(user.code)

        if muni_name:
            try:
                # Find the municipality with this name
                municipality = Municipality.objects.get(muni_name=muni_name)
                # Filter doctors by this municipality
                doct = Doctor.objects.filter(muni_id=municipality).first()
                doctor = doct.full_name() if doct else "No Doctor Assigned"
                coordinator = f"Municipality of {municipality.muni_name}"
                pho = f"Doctor of {municipality.muni_name}"
                center = muni_name  # Use the user's code to determine the center
                center_label = f"Animal Bite Treatment Center: {center} Animal Bite Treatment Center"  # Add the label
            except Municipality.DoesNotExist:
                # Handle case where the municipality name does not exist
                coordinator = "Unknown Municipality"
                pho = "No Doctor Assigned"
                doctor = "No Doctor Assigned"
                center = "Unknown Center"
                center_label = f"Animal Bite Treatment Center: {center}"  # Add the label
        else:
            # Handle case where the user's code does not match any municipality
            coordinator = "Unknown Municipality"
            pho = "No Doctor Assigned"
            doctor = "No Doctor Assigned"
            center = "Unknown Center"
            center_label = f"Animal Bite Treatment Center: {center} Animal Bite Treatment Center"  # Add the label

    
    context = {
        'doctor': doctor,
        'coordinator': coordinator,
        'center': center,
        'center_label': center_label,
        'pho': pho,
        'karon': karon,
        'logo_url': logo_url,
        'municipality_name': municipality_name,
        'selected_quarter': selected_quarter,
        'signature_name': signature_name,
        'municipalities': municipalities,
        'municipal': municipal,
        'barangays': barangays,
        'patients': patients,
        'histories': histories,
        'treatments': treatments,
        'category_ii_count': category_ii_count,
        'category_iii_count': category_iii_count,
        'total_count': total_count,
        'total_count_rig': total_count_rig,
        'category_ii_with_rig': category_ii_with_rig,
        'category_iii_with_rig': category_iii_with_rig,
        'category_ii_complete': category_ii_complete,
        'category_ii_incomplete': category_ii_incomplete,
        'category_ii_none': category_ii_none,
        'category_ii_died': category_ii_died,
        'category_iii_complete': category_iii_complete,
        'category_iii_incomplete': category_iii_incomplete,
        'category_iii_none': category_iii_none,
        'category_iii_died': category_iii_died,
        'total_complete': total_complete,
        'total_incomplete': total_incomplete,
        'total_none': total_none,
        'total_died': total_died,
        'year': year,
    }
    template = loader.get_template('monitoring/cohort_pdf_annual.html')
    html = template.render(context)

    """ # Specify path to wkhtmltopdf executable
    path_to_wkhtmltopdf = r"C:\wkhtmltox\bin\wkhtmltopdf.exe"
    
    # Configure pdfkit with the path
    config = pdfkit.configuration(wkhtmltopdf=path_to_wkhtmltopdf) """


    config = pdfkit.configuration(wkhtmltopdf=WKHTMLTOPDF_PATH)

    # Options for landscape orientation
    options = {
        'orientation': 'Landscape',  # Set to Landscape
        'page-size': 'A4',  # A4 paper size, you can adjust if needed
        'footer-left': center_label,  # Add page numbers
        'footer-right': 'Page [page] of [toPage]',  # Add page numbers
        'footer-font-size': '10',  # Adjust font size for the footer
        'footer-spacing': '5',  # Space between footer and content
        'margin-bottom': '15mm',  # Ensure space for the footer
    }
    
    # Generate PDF from HTML using pdfkit with configuration
    pdf = pdfkit.from_string(html, configuration=config,options=options)

    # Return the PDF as a responsev 
    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment;filename="Annual Cohort.pdf"'

    return response


@login_required
def download_excel(request):
    # Create a workbook and a worksheet
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Animal'


    # Save the workbook to a BytesIO stream
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)

    # Create the response
    response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="animal_bite_report.xlsx"'
    return response

def export_excel(request):
    # Create a workbook and get the active worksheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Animal"

    # Define the table headers based on your HTML table structure
    headers = [
        # Row 1
        ["Barangay/ABTC", "Human Case", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "Biting Animals", "", "PEP Coverage", "", "Vaccination Status of Biting Animal", ""],
        # Row 2
        ["", "", "SEX", "", "", "AGE", "", "", "", "Animal Bite", "", "", "HR", "", "Post Exposure Prophylaxis", "", "", "", "", "", "", "", "", "", ""],
        # Row 3
        ["", "", "", "", "", "", "", "", "By Category", "", "", "", "", "", "", "", "", "", "", "", "", "", "", ""],
        # Row 4
        ["", "M", "F", "Total", "<=15", ">15", "Total", "I", "II", "III", "Total", "%/Total", "No.", "TCV", "HRIG", "ERIG", "D", "C", "O", "Total", "%TCV", "%ERIG", "Yes", "No"]
    ]

    # Write headers to the sheet
    for row_num, row in enumerate(headers, 1):
        for col_num, value in enumerate(row, 1):
            cell = sheet.cell(row=row_num, column=col_num, value=value)
            cell.alignment = Alignment(horizontal="center", vertical="center")  # Align text to center both horizontally and vertically

    # Merge the cells as per colspan and rowspan in the HTML table
    sheet.merge_cells('A1:A4')  # "Barangay/ABTC" merges across 4 rows
    sheet.merge_cells('B1:P1')  # "Human Case" merges from B1 to P1
    sheet.merge_cells('Q1:T1')  # "Biting Animals" merges from Q1 to T1
    sheet.merge_cells('U1:V1')  # "PEP Coverage" merges from U1 to V1
    sheet.merge_cells('W1:X1')  # "Vaccination Status of Biting Animal" merges from W1 to X1

    sheet.merge_cells('B2:D3')  # "SEX" merges from B2 to D3
    sheet.merge_cells('E2:G3')  # "AGE" merges from E2 to G3
    sheet.merge_cells('H2:L2')  # "Animal Bite" merges from H2 to L2
    sheet.merge_cells('M2:M3')  # "HR" merges from M2 to M3
    sheet.merge_cells('N2:P3')  # "Post Exposure Prophylaxis" merges from N2 to P3

    # Individual column merging in the 4th row
    sheet.merge_cells('H3:H4')  # "By Category I"
    sheet.merge_cells('I3:I4')  # "By Category II"
    sheet.merge_cells('J3:J4')  # "By Category III"
    sheet.merge_cells('K3:K4')  # "Total"
    sheet.merge_cells('L3:L4')  # "%/Total"

    sheet.merge_cells('M4:M4')  # "No."
    sheet.merge_cells('N4:N4')  # "TCV"
    sheet.merge_cells('O4:O4')  # "HRIG"
    sheet.merge_cells('P4:P4')  # "ERIG"

    sheet.merge_cells('Q4:Q4')  # "D"
    sheet.merge_cells('R4:R4')  # "C"
    sheet.merge_cells('S4:S4')  # "O"
    sheet.merge_cells('T4:T4')  # "Total"
    sheet.merge_cells('U4:U4')  # "%TCV"
    sheet.merge_cells('V4:V4')  # "%ERIG"
    sheet.merge_cells('W4:W4')  # "Yes"
    sheet.merge_cells('X4:X4')  # "No"

    # Assuming you have data in a similar structure as your HTML
    data = [
        # Example row data
        ['Barangay 1', 10, 5, 15, 7, 8, 15, 3, 6, 1, 10, '100%', 2, 8, 3, 1, 4, 5, 1, 10, '80%', '90%', 8, 2],
        ['Barangay 2', 12, 6, 18, 9, 9, 18, 4, 5, 2, 11, '95%', 1, 7, 2, 1, 3, 6, 1, 10, '85%', '88%', 7, 3],
        # Add more rows as per your data
    ]

    # Write the data to the sheet
    for row_num, row_data in enumerate(data, len(headers) + 1):
        for col_num, value in enumerate(row_data, 1):
            cell = sheet.cell(row=row_num, column=col_num, value=value)
            cell.alignment = Alignment(horizontal="center", vertical="center")  # Align all data cells to the center

    # Adjust the column widths for readability
    for i in range(1, sheet.max_column + 1):
        sheet.column_dimensions[get_column_letter(i)].width = 15

    # Set the response to download the file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=animal_bite_report.xlsx'

    # Save the workbook to the response
    workbook.save(response)
    return response

def exp_excel(request):
    # Get the selected quarter from the URL parameters
    selected_quarter = request.GET.get('quarter', '1')  # Default to '1' if no quarter is selected

    # Determine the start and end dates based on the selected quarter
    if selected_quarter == '1':
        start_date = '2024-01-01'
        end_date = '2024-03-31'
    elif selected_quarter == '2':
        start_date = '2024-04-01'
        end_date = '2024-06-30'
    elif selected_quarter == '3':
        start_date = '2024-07-01'
        end_date = '2024-09-30'
    else:
        start_date = '2024-10-01'
        end_date = '2024-12-31'

    # Filter History data by date_registered based on the selected quarter
    data = History.objects.filter(date_registered__range=[start_date, end_date])

    # Create a new Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Q{selected_quarter} Data"

    # Define the headers based on your table structure
    headers = ['Patient Name', 'Date of Exposure', 'Barangay', 'Category', 'Animal Type', 'Bite Site', 'Human Rabies', 'Vaccine Given']
    ws.append(headers)

    # Populate the table rows with filtered data
    for history in data:
        patient = history.patient_id  # Get the patient linked to the history
        treatments = patient.treatments_patient.all()  # Get all treatments for the patient
        
        # Check if there are treatments, and get the first treatment's TCV date if it exists
        vaccine_given = treatments.first().tcv_given if treatments.exists() else 'N/A'

        # Prepare the row with the data
        row = [
            f"{patient.first_name} {patient.last_name}",  # Patient Name
            history.date_of_exposure,  # Date of Exposure
            history.brgy_id.brgy_name,  # Barangay Name
            history.category_of_exposure,  # Category
            history.source_of_exposure,  # Animal Type
            history.bite_site,  # Bite Site
            'Yes' if history.human_rabies else 'No',  # Human Rabies
            vaccine_given  # Vaccine Given (if exists, otherwise N/A)
        ]
        ws.append(row)

    # Set up HTTP response to serve the Excel file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=report_Q{selected_quarter}.xlsx'

    # Save the workbook to the response
    wb.save(response)

    return response

@login_required
# Function to generate the Excel report
def download_report_excel(request):
    user = request.user
    municipality_map = {
        "MAR": "Maripipi",
        "KAW": "Kawayan",
        "NAV": "Naval",
        "CAIB": "Caibiran"
    }

    selected_quarter = request.GET.get('quarter', '1')
    year = 2024

    quarter_ranges = {
        '1': (date(year, 1, 1), date(year, 3, 31)),
        '2': (date(year, 4, 1), date(year, 6, 30)),
        '3': (date(year, 7, 1), date(year, 9, 30)),
        '4': (date(year, 10, 1), date(year, 12, 31)),
    }

    start_date, end_date = quarter_ranges[selected_quarter]

    # Example data - Replace with your actual data
    data = [
        {
            'barangay': 'Barangay A',
            'data_male': 10,
            'data_female': 12,
            'data_total': 22,
            'age_15_below': 8,
            'age_above_15': 14,
            'total_animal_bite_I': 5,
            'total_animal_bite_II': 4,
            'total_animal_bite_III': 7,
            'total_tcv_given': 5,
            'total_hrig_given': 3,
            'total_erig_given': 2,
            'total_dog_bites': 9,
            'total_cat_bites': 5,
            'total_other_bites': 1,
            'immunized_count': 15,
            'unimmunized_count': 7,
            'human_rabies_count': 0
        },
    ]
    wb = Workbook()
    ws = wb.active
    ws.title = "Animal Bite and Rabies Report"
    # Set headers
    headers = [
        'Barangay', 'Male', 'Female', 'Total', 'Age 15 and Below', 'Age Above 15',
        'Total Animal Bite I', 'Total Animal Bite II', 'Total Animal Bite III', 
        'Total TCV Given', 'Total HRIG Given', 'Total ERIG Given', 'Dog Bites', 
        'Cat Bites', 'Other Bites', 'Immunized', 'Unimmunized', 'Human Rabies Cases'
    ]
    ws.append(headers)
    for row in data:
        ws.append([ 
            row['barangay'], row['data_male'], row['data_female'], row['data_total'],
            row['age_15_below'], row['age_above_15'], row['total_animal_bite_I'],
            row['total_animal_bite_II'], row['total_animal_bite_III'], 
            row['total_tcv_given'], row['total_hrig_given'], row['total_erig_given'],
            row['total_dog_bites'], row['total_cat_bites'], row['total_other_bites'],
            row['immunized_count'], row['unimmunized_count'], row['human_rabies_count']
        ])
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)  # Move the pointer to the beginning of the file

    # Create the response object for the Excel download
    response = HttpResponse(
        content=excel_file.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="animal_bite_report.xlsx"'

    return response

@login_required
def download_report_pdf(request):
    user = request.user
    municipality_map = {
        "MAR": "Maripipi",
        "KAW": "Kawayan",
        "NAV": "Naval",
        "CAIB": "Caibiran"
    }
    if user.is_superuser and user.code == "NAV":
        municipality_name = "BPH"
    else:
        municipality_name = municipality_map.get(user.code, "Naval")
    selected_quarter = request.GET.get('quarter', '1')
    year = 2024
    quarter_ranges = {
        '1': (date(year, 1, 1), date(year, 3, 31)),
        '2': (date(year, 4, 1), date(year, 6, 30)),
        '3': (date(year, 7, 1), date(year, 9, 30)),
        '4': (date(year, 10, 1), date(year, 12, 31)),
    }
    start_date, end_date = quarter_ranges[selected_quarter]
    data = []
    total_male = 0
    total_female = 0
    total_all = 0
    total_age_15_below = 0
    total_age_above_15 = 0
    total_sex_percentage = 0
    total_animal_bite_I = 0
    total_animal_bite_II = 0
    total_animal_bite_III = 0
    total_category_percentage = 0
    total_tcv_given = 0
    total_hrig_given = 0
    total_erig_given = 0  
    total_tcv_percentage = 0
    total_hrig_percentage = 0
    total_rig_percentage = 0
    total_dog_bites = 0
    total_cat_bites = 0
    total_other_bites = 0
    total_animal_type_percentage = 0
    total_animal_bite_I_percentage = 0
    total_animal_bite_II_percentage = 0
    total_animal_bite_III_percentage = 0
    total_immunized = 0
    total_unimmunized = 0

    if user.is_superuser:
        table_head = "ABTC"
        
        abtcs = User.objects.filter(is_superuser=False).distinct()
        for abtc_user in abtcs:
            if not abtc_user.code:
                continue
            male_count = History.objects.filter(
                patient_id__user=abtc_user,
                date_registered__range=(start_date, end_date),
                patient_id__sex='male'
            ).count()
            female_count = History.objects.filter(
                patient_id__user=abtc_user,
                date_registered__range=(start_date, end_date),
                patient_id__sex='female'
            ).count()

            patients = Patient.objects.filter(
                user=abtc_user,  
                histories__date_registered__range=(start_date, end_date)  
            ).distinct()
            age_15_below_count = sum(1 for patient in patients if calculate_age(patient.birthday) <= 15)
            age_above_15_count = sum(1 for patient in patients if calculate_age(patient.birthday) > 15)

            user_animal_bite_I = 0
            user_animal_bite_II = 0
            user_animal_bite_III = 0

            animal_bite_counts = History.objects.filter(
                patient_id__user=abtc_user,
                date_registered__range=(start_date, end_date)
            ).values('category_of_exposure').annotate(count=models.Count('category_of_exposure'))

            for count in animal_bite_counts:
                if count['category_of_exposure'] == 'I':
                    user_animal_bite_I = count['count']
                elif count['category_of_exposure'] == 'II':
                    user_animal_bite_II = count['count']
                elif count['category_of_exposure'] == 'III':
                    user_animal_bite_III = count['count']

            tcv_count = Treatment.objects.filter(
                patient_id__user=abtc_user,
                tcv_given__range=(start_date, end_date)
            ).count()

            hrig_count = Treatment.objects.filter(
                patient_id__user=abtc_user,
                hrig_given__range=(start_date, end_date)
            ).count()

            erig_count = Treatment.objects.filter(
                patient_id__user=abtc_user,
                rig_given__range=(start_date, end_date) 
            ).count()

            animal_type_counts = History.objects.filter(
                patient_id__user=abtc_user,
                date_registered__range=(start_date, end_date)
            ).values('source_of_exposure').annotate(count=models.Count('source_of_exposure'))

            user_dog_bites = 0
            user_cat_bites = 0
            user_other_bites = 0

            for count in animal_type_counts:
                if count['source_of_exposure'] == 'Dog':
                    user_dog_bites = count['count']
                elif count['source_of_exposure'] == 'Cat':
                    user_cat_bites = count['count']
                elif count['source_of_exposure'] == 'Others':
                    user_other_bites = count['count']

            user_immunized_count = History.objects.filter(
                patient_id__user=abtc_user,
                immunization_status='Immunized',
                date_registered__range=(start_date, end_date)
            ).count()
            
            user_unimmunized_count = History.objects.filter(
                patient_id__user=abtc_user,
                immunization_status='Unimmunized',
                date_registered__range=(start_date, end_date)
            ).count()

            user_human_rabies_count = History.objects.filter(
                patient_id__user=abtc_user,
                human_rabies=True,
                date_registered__range=(start_date, end_date)
            ).count()
                    
            total_immunized += user_immunized_count
            total_unimmunized += user_unimmunized_count
            total_dog_bites += user_dog_bites
            total_cat_bites += user_cat_bites
            total_other_bites += user_other_bites
            total_tcv_given += tcv_count
            total_hrig_given += hrig_count
            total_erig_given += erig_count

            total_count = male_count + female_count
            total_male += male_count
            total_female += female_count
            total_all += total_count
            total_age_15_below += age_15_below_count
            total_age_above_15 += age_above_15_count
            total_animal_bite_I += user_animal_bite_I
            total_animal_bite_II += user_animal_bite_II
            total_animal_bite_III += user_animal_bite_III

            data.append({
                'table_head':table_head,
                'barangay': f"{municipality_map.get(abtc_user.code, 'Unknown')}-ABTC",
                'data_male': male_count,
                'data_female': female_count,
                'data_total': total_count,
                'age_15_below': age_15_below_count,
                'age_above_15': age_above_15_count,
                'age_total': age_15_below_count + age_above_15_count,
                'total_animal_bite_I':user_animal_bite_I,
                'total_animal_bite_II':user_animal_bite_II,
                'total_animal_bite_III':user_animal_bite_III,
                'total_animal':user_animal_bite_I + user_animal_bite_II + user_animal_bite_III,
                'total_tcv_given': tcv_count,
                'total_hrig_given': hrig_count,
                'total_erig_given': erig_count,
                'total_dog_bites': user_dog_bites,
                'total_cat_bites': user_cat_bites,
                'total_other_bites': user_other_bites,
                'total_animal_bites': user_dog_bites + user_cat_bites + user_other_bites,
                'immunized_count': user_immunized_count,
                'unimmunized_count': user_unimmunized_count,  
                'human_rabies_count':user_human_rabies_count,

            })
    else:
        table_head = "Barangay"
        # For non-superuser
        patients = Patient.objects.filter(user=user)
        barangays = Barangay.objects.filter(patients_brgy__in=patients).distinct()

        for barangay in barangays:
            male_count = History.objects.filter(
                patient_id__brgy_id=barangay,
                patient_id__user=user,
                date_registered__range=(start_date, end_date),
                patient_id__sex='male'
            ).count()
            female_count = History.objects.filter(
                patient_id__brgy_id=barangay,
                patient_id__user=user,
                date_registered__range=(start_date, end_date),
                patient_id__sex='female'
            ).count()

            patients = Patient.objects.filter(
                brgy_id=barangay,
                user=user,
                histories__date_registered__range=(start_date, end_date)  # Filter by registration date within the quarter
            ).distinct()
            age_15_below_count = sum(1 for patient in patients if calculate_age(patient.birthday) <= 15)
            age_above_15_count = sum(1 for patient in patients if calculate_age(patient.birthday) > 15)

            barangay_animal_bite_I = 0
            barangay_animal_bite_II = 0
            barangay_animal_bite_III = 0

            animal_bite_counts = History.objects.filter(
                patient_id__brgy_id=barangay,
                patient_id__user=user,
                date_registered__range=(start_date, end_date)
            ).values('category_of_exposure').annotate(count=models.Count('category_of_exposure'))

            tcv_count = Treatment.objects.filter(
                patient_id__brgy_id=barangay,
                patient_id__user=user,
                tcv_given__range=(start_date, end_date)
            ).count()

            hrig_count = Treatment.objects.filter(
                patient_id__brgy_id=barangay,
                patient_id__user=user,
                hrig_given__range=(start_date, end_date)
            ).count()

            erig_count = Treatment.objects.filter(
                patient_id__brgy_id=barangay,
                patient_id__user=user,
                rig_given__range=(start_date, end_date)
            ).count()

            total_tcv_given += tcv_count
            total_hrig_given += hrig_count
            total_erig_given += erig_count

            for count in animal_bite_counts:
                if count['category_of_exposure'] == 'I':
                    barangay_animal_bite_I = count['count']
                elif count['category_of_exposure'] == 'II':
                    barangay_animal_bite_II = count['count']
                elif count['category_of_exposure'] == 'III':
                    barangay_animal_bite_III = count['count']

            animal_type_counts = History.objects.filter(
                patient_id__brgy_id=barangay,
                patient_id__user=user,
                date_registered__range=(start_date, end_date)
            ).values('source_of_exposure').annotate(count=models.Count('source_of_exposure'))

            barangay_dog_bites = 0
            barangay_cat_bites = 0
            barangay_other_bites = 0

            for count in animal_type_counts:
                if count['source_of_exposure'] == 'Dog':
                    barangay_dog_bites = count['count']
                elif count['source_of_exposure'] == 'Cat':
                    barangay_cat_bites = count['count']
                elif count['source_of_exposure'] == 'Others':
                    barangay_other_bites = count['count']

            barangay_immunized_count = History.objects.filter(
                patient_id__brgy_id=barangay,
                patient_id__user=user,
                immunization_status='Immunized',
                date_registered__range=(start_date, end_date)
            ).count()
            
            barangay_unimmunized_count = History.objects.filter(
                patient_id__brgy_id=barangay,
                patient_id__user=user,
                immunization_status='Unimmunized',
                date_registered__range=(start_date, end_date)
            ).count()

            barangay_human_rabies_count = History.objects.filter(
                patient_id__brgy_id=barangay,
                patient_id__user=user,
                human_rabies=True,
                date_registered__range=(start_date, end_date)
            ).count()

            total_immunized += barangay_immunized_count
            total_unimmunized += barangay_unimmunized_count
            total_dog_bites += barangay_dog_bites
            total_cat_bites += barangay_cat_bites
            total_other_bites += barangay_other_bites
            total_dog_bites += barangay_dog_bites
            total_cat_bites += barangay_cat_bites
            total_other_bites += barangay_other_bites
            total_animal_bite_I += barangay_animal_bite_I
            total_animal_bite_II += barangay_animal_bite_II
            total_animal_bite_III += barangay_animal_bite_III

            # Add to total counts
            total_count = male_count + female_count
            total_male += male_count
            total_female += female_count
            total_all += total_count
            total_age_15_below += age_15_below_count
            total_age_above_15 += age_above_15_count

            data.append({
                'table_head':table_head,
                'barangay': barangay.brgy_name,
                'data_male': male_count,
                'data_female': female_count,
                'data_total': total_count,
                'age_15_below': age_15_below_count,
                'age_above_15': age_above_15_count,
                'age_total': age_15_below_count + age_above_15_count,
                'total_animal_bite_I': barangay_animal_bite_I,
                'total_animal_bite_II': barangay_animal_bite_II,
                'total_animal_bite_III': barangay_animal_bite_III,
                'total_animal': barangay_animal_bite_I + barangay_animal_bite_II + barangay_animal_bite_III,
                'total_tcv_given': tcv_count,
                'total_hrig_given': hrig_count,
                'total_erig_given': erig_count,
                'total_dog_bites': barangay_dog_bites,
                'total_cat_bites': barangay_cat_bites,
                'total_other_bites': barangay_other_bites,
                'total_animal_bites': barangay_dog_bites + barangay_cat_bites + barangay_other_bites,
                'immunized_count': barangay_immunized_count,  # Change made here
                'unimmunized_count': barangay_unimmunized_count,  # Change made here
                'human_rabies_count':barangay_human_rabies_count,

            })

    if total_all > 0:
        male_percentage = (total_male / total_all) * 100
        female_percentage = (total_female / total_all) * 100
        total_sex_percentage = (male_percentage + female_percentage)
        age_15_below_percentage = (total_age_15_below / total_all) * 100
        age_above_15_percentage = (total_age_above_15 / total_all) * 100
        total_age_percentage = (total_age_15_below + total_age_above_15) / total_all * 100
        total_animal_bite_I_percentage = (total_animal_bite_I / total_all ) * 100
        total_animal_bite_II_percentage = (total_animal_bite_II / total_all ) * 100
        total_animal_bite_III_percentage = (total_animal_bite_III / total_all ) * 100
        total_category_percentage = (total_animal_bite_I_percentage + total_animal_bite_II_percentage + total_animal_bite_III_percentage)
        total_tcv_percentage = (total_tcv_given / total_all) * 100
        total_hrig_percentage = (total_hrig_given / total_all) * 100
        total_rig_percentage = (total_erig_given / total_all) * 100  
        dog_bite_percentage = (total_dog_bites / total_all) * 100
        cat_bite_percentage = (total_cat_bites / total_all) * 100
        other_bite_percentage = (total_other_bites / total_all) * 100
        total_animal_type_percentage = dog_bite_percentage + cat_bite_percentage + other_bite_percentage
        if total_immunized + total_unimmunized > 0:
            immunized_percentage = (total_immunized / (total_immunized + total_unimmunized)) * 100
            unimmunized_percentage = (total_unimmunized / (total_immunized + total_unimmunized)) * 100
        else:
            immunized_percentage = 0
            unimmunized_percentage = 0
    else:
        male_percentage = female_percentage = age_15_below_percentage = age_above_15_percentage = total_age_percentage = 0
        total_tcv_percentage = total_hrig_percentage = total_rig_percentage = 0
        dog_bite_percentage = cat_bite_percentage = other_bite_percentage = 0
        immunized_percentage = 0
        unimmunized_percentage = 0

    overall_total = sum(entry.get('data_total', 0) for entry in data)
    overall_total_tcv = sum(entry.get('total_tcv_given', 0) for entry in data)
    overall_total_hrig = sum(entry.get('total_hrig_given', 0) for entry in data)
    overall_total_erig = sum(entry.get('total_erig_given', 0) for entry in data)

    for entry in data:
        entry['percent_total'] = round((entry['data_total'] / overall_total) * 100, 1) if overall_total > 0 else 0
        entry['percent_tcv'] = round((entry.get('total_tcv_given', 0) / overall_total_tcv) * 100, 1) if overall_total_tcv > 0 else 0
        entry['percent_hrig'] = round((entry.get('total_hrig_given', 0) / overall_total_hrig) * 100, 1) if overall_total_hrig > 0 else 0
        entry['percent_erig'] = round((entry.get('total_erig_given', 0) / overall_total_erig) * 100, 1) if overall_total_erig > 0 else 0
    #all percent totals
    total_percent = sum(entry['percent_total'] for entry in data)
    total_tcv_percent = sum(entry['percent_tcv'] for entry in data)
    total_hrig_percent = sum(entry['percent_hrig'] for entry in data)
    total_erig_percent = sum(entry['percent_erig'] for entry in data)   
    total_human_rabies = sum(entry.get('human_rabies_count', 0) for entry in data)  # Add this line

    municipalities = Municipality.objects.all()
    selected_municipality_id = request.GET.get('municipality_id')
    selected_municipality = municipalities.filter(muni_id=selected_municipality_id).first() if selected_municipality_id else None

    if user.is_superuser:
        logo_url = 'assets/images/report_logo/province.png'
    elif user.code == 'MAR':
        logo_url = 'assets/images/report_logo/maripipi.png'
    elif user.code == 'KAW':
        logo_url = 'assets/images/report_logo/kawayan.png'
    elif user.code == 'NAV':
        logo_url = 'assets/images/report_logo/naval.png'
    elif user.code == 'CAIB':
        logo_url = 'assets/images/report_logo/caibiran.png'
    else:
        logo_url = None  # or some default logo

    full_name = f"{user.first_name} {user.last_name}" if user.first_name or user.last_name else user.username

    if user.is_superuser:
        table = "ABTC"
    else:
        table = "Barangay"

    # Pass totals and data to the template
    html_content = render_to_string('monitoring/report_pdf.html', {
        'table':table,
        'table_head':table_head,
        'logo_url': logo_url,
        'full_name': full_name,
        'municipalities': municipalities,
        'selected_municipality': selected_municipality or municipalities.first(),  # Default to the first municipality if none is selected
        'municipality_name': municipality_name,
        'selected_quarter': selected_quarter,
        'barangay_list': [d['barangay'] for d in data],
        'data': data,
        'total_male': total_male,
        'total_female': total_female,
        'total_all': total_all,
        'total_age_15_below': total_age_15_below,
        'total_age_above_15': total_age_above_15,
        'male_percentage': round(male_percentage, 1),
        'female_percentage': round(female_percentage, 1),
        'total_sex_percentage':round(total_sex_percentage),
        'age_15_below_percentage': round(age_15_below_percentage, 1),
        'age_above_15_percentage': round(age_above_15_percentage, 1),
        'total_age_percentage': round(total_age_percentage, 1),
        'total_animal_bite_I': total_animal_bite_I,
        'total_animal_bite_II': total_animal_bite_II,
        'total_animal_bite_III': total_animal_bite_III,
        'total_animal_bite_I_percentage': round(total_animal_bite_I_percentage,1),
        'total_animal_bite_II_percentage': round(total_animal_bite_II_percentage,1),
        'total_animal_bite_III_percentage': round(total_animal_bite_III_percentage,1),
        'total_category_percentage': round(total_category_percentage,1),
        'overall_total':overall_total,
        'total_percent':round(total_percent, ),
        'total_tcv_given': total_tcv_given,
        'total_hrig_given': total_hrig_given,
        'total_rig_given': total_erig_given,
        'total_tcv_percentage': round(total_tcv_percentage, 1),
        'total_hrig_percentage': round(total_hrig_percentage, 1),
        'total_rig_percentage': round(total_rig_percentage, 1),
        'total_tcv_percent':round(total_tcv_percent, ),
        'total_erig_percent':round(total_erig_percent, ),
        'total_hrig_percent':round(total_hrig_percent, ),
        'total_dog_bites': total_dog_bites,
        'total_cat_bites': total_cat_bites,
        'total_other_bites': total_other_bites,
        'dog_bite_percentage': round(dog_bite_percentage, 1),
        'cat_bite_percentage': round(cat_bite_percentage, 1),
        'other_bite_percentage': round(other_bite_percentage, 1),
        'total_animal_type_percentage': round(total_animal_type_percentage, 1),
        'total_tcv_given': sum(entry.get('total_tcv_given', 0) for entry in data),
        'total_hrig_given': sum(entry.get('total_hrig_given', 0) for entry in data),
        'total_erig_given': sum(entry.get('total_erig_given', 0) for entry in data),
        'total_immunized': total_immunized,
        'total_unimmunized': total_unimmunized,
        'immunized_percentage': round(immunized_percentage, 1), 
        'unimmunized_percentage': round(unimmunized_percentage, 1),  
        'total_human_rabies':total_human_rabies,
    })

    # Create a response object for the PDF download
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'filename="Report_pdf.pdf"' #attachment;

    # Generate the PDF from HTML
    pisa_status = pisa.CreatePDF(html_content, dest=response)

    # If the PDF creation fails, return an error message
    if pisa_status.err:
        return HttpResponse('We had some errors <pre>' + html_content + '</pre>', content_type='text/html')

    return response

# PDF Generation Function
@login_required
def download_masterlist_pdf(request):
    user = request.user

    # Get selected filters from request
    selected_municipality = request.GET.get('municipality')
    selected_barangay = request.GET.get('barangay')
    selected_user = request.GET.get('searchUsername') 
    start_month = request.GET.get('startMonth')
    end_month = request.GET.get('endMonth')
    search_name = request.GET.get('searchName')
    
    # Fetch the histories with related patient, municipality, and barangay data
    histories = History.objects.select_related('patient_id', 'muni_id', 'brgy_id').order_by('-registration_no')

    if not user.is_superuser:
        # Filter histories for the current user if not a superuser
        histories = histories.filter(patient_id__user=user)
    
    patients = Patient.objects.all()

    # Apply filters based on selected municipality and barangay
    if selected_municipality:
        histories = histories.filter(muni_id=selected_municipality)
    if selected_barangay:
        histories = histories.filter(brgy_id=selected_barangay)

    # Apply filter based on username search only if user is a superuser
    if user.is_superuser and selected_user:
        histories = histories.filter(patient_id__user__username=selected_user)
    
    # Fetch unique users from the Patient model for the dropdown if user is superuser
    if user.is_superuser:
        patient_users = Patient.objects.select_related('user').values_list('user', flat=True).distinct()
        users = User.objects.filter(id__in=patient_users)
    else:
        users = []

    # Apply filter based on selected start and end months
    if start_month and end_month:
        try:
            current_year = datetime.now().year
            start_date = datetime.strptime(f"{start_month} {current_year}", "%B %Y").replace(day=1)
            end_date = datetime.strptime(f"{end_month} {current_year}", "%B %Y").replace(day=1) + relativedelta(months=1) - relativedelta(days=1)
            histories = histories.filter(date_registered__gte=start_date, date_registered__lte=end_date)
        except Exception as e:
            print(f"Error parsing date: {e}")  # Log the error and continue

     # Apply filter based on name search
    if search_name:
        histories = histories.filter(Q(patient_id__first_name__icontains=search_name) | Q(patient_id__last_name__icontains=search_name))

    months = month_name[1:]
    
    # Calculate age and attach to each history instance
    for history in histories:
        history.treatment = Treatment.objects.filter(patient_id=history.patient_id).first()
        history.age = calculate_age(history.patient_id.birthday)

    # Count the number of male and female patients
    male = histories.filter(patient_id__sex__iexact='Male').count()
    female = histories.filter(patient_id__sex__iexact='Female').count()

    # Calculate the number of age
    age_15_or_less_count = 0
    age_above_15_count = 0

    for history in histories:
        age = calculate_age(history.patient_id.birthday)
        if age <= 15:
            age_15_or_less_count += 1
        else:
            age_above_15_count += 1

    # Calculate counts for different animal bites   
    source_of_exposure_counter = Counter(histories.values_list('source_of_exposure', flat=True))
    dog_count = source_of_exposure_counter.get('Dog', 0)
    cat_count = source_of_exposure_counter.get('Cat', 0)
    other_animal_count = source_of_exposure_counter.get('Others', 0)

    paginator = Paginator(histories,10)  
    page_number = request.GET.get('page',1)
    try:
        histories = paginator.get_page(page_number)
    except PageNotAnInteger:
        histories = paginator.get_page(1)
    except EmptyPage:
        histories = paginator.get_page(paginator.num_pages)

    # Collect current query parameters
    query_params = request.GET.dict()
    if 'page' in query_params:
        del query_params['page']  # Remove the page parameter
    query_string = urlencode(query_params)

    municipalities = Municipality.objects.all()
    barangays = Barangay.objects.all()

    # Pass totals and data to the template
    html_content = render_to_string('monitoring/download_pdf.html', {
        'histories': histories,
        'municipalities': municipalities,
        'barangays': barangays,
        'selected_municipality': selected_municipality,
        'selected_barangay': selected_barangay,
        'selected_user':selected_user,
        'start_month': start_month,
        'end_month': end_month,
        'search_name': search_name,
        'months': months,
        'male' : male,
        'female' : female,
        'dog_count' : dog_count,
        'cat_count' : cat_count,
        'other_animal_count' : other_animal_count,
        'age_15_or_less_count' : age_15_or_less_count,
        'age_above_15_count' : age_above_15_count,
        'query_string' : query_string,
        'users':users
    })

    # Create a response object for the PDF download
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment;filename="Master_list_pdf.pdf"' #attachment;

    # Generate the PDF from HTML
    pisa_status = pisa.CreatePDF(html_content, dest=response)

    # If the PDF creation fails, return an error message
    if pisa_status.err:
        return HttpResponse('We had some errors <pre>' + html_content + '</pre>', content_type='text/html')

    return response

# Excel Generation Function
@login_required
def download_masterlist_excel(request):
    # Fetch data similar to your existing download view
    user = request.user
    selected_municipality = request.GET.get('municipality')
    selected_barangay = request.GET.get('barangay')
    selected_user = request.GET.get('searchUsername') 
    start_month = request.GET.get('startMonth')
    end_month = request.GET.get('endMonth')
    search_name = request.GET.get('searchName')

    # Query for histories
    histories = History.objects.select_related('patient_id', 'muni_id', 'brgy_id').order_by('-registration_no')
    
    if selected_municipality:
        histories = histories.filter(muni_id=selected_municipality)
    if selected_barangay:
        histories = histories.filter(brgy_id=selected_barangay)
    if selected_user and user.is_superuser:
        histories = histories.filter(patient_id__user__username=selected_user)
    if start_month and end_month:
        start_date = datetime.strptime(f"{start_month} {datetime.now().year}", "%B %Y").replace(day=1)
        end_date = datetime.strptime(f"{end_month} {datetime.now().year}", "%B %Y").replace(day=1) + relativedelta(months=1) - relativedelta(days=1)
        histories = histories.filter(date_registered__gte=start_date, date_registered__lte=end_date)
    if search_name:
        histories = histories.filter(Q(patient_id__first_name__icontains=search_name) | Q(patient_id__last_name__icontains=search_name))

    # Create Excel file using openpyxl
    wb = Workbook()
    ws = wb.active
    ws.title = "Masterlist Report"

    # Add headers
    headers = ['Registration No.', 'Name', 'Barangay', 'Municipality', 'Sex', 'Date of Exposure']
    ws.append(headers)

    # Populate data rows
    for history in histories:
        ws.append([
            history.registration_no,
            f"{history.patient_id.first_name} {history.patient_id.last_name}",
            history.brgy_id.brgy_name,
            history.muni_id.muni_name,
            history.patient_id.sex,
            history.date_of_exposure,
        ])

    # Create Excel response
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = 'attachment; filename="rabies_masterlist.xlsx"'

    # Save the workbook to the response
    wb.save(response)

    return response

